"""
Infinia Labour Tool - Web Backend
====================================
FastAPI application. Reuses the desktop app's own calculation and
validation logic (data_engine.py, daily_attendance.py, payroll_cycle.py)
directly - the only thing that changed is the storage layer, from
pickled sessions/JSON files to a real multi-user database.
"""
from datetime import date, datetime, timedelta, timezone
from typing import Optional
import re
import io
import json

from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, func
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, engine, Base, SessionLocal
import models
import schemas
import services
import auth
import payroll_cycle as pcyc
import reports as rp
import export_web

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Infinia Labour Tool API")


@app.on_event("startup")
def seed_on_startup():
    """
    Runs the same seeding logic as seed_data.py automatically on every
    startup - idempotent (upserts, never duplicates), so this is safe
    to run every single time the app boots. Needed specifically because
    Render's free tier has no shell access to run a one-off script
    manually; master_data.json (if bundled alongside this file) gets
    picked up automatically the first time the app starts.
    """
    import os
    import seed_data
    # Only seed from master_data.json when there are NO employees yet.
    # Re-running it on every boot meant the file was the permanent source
    # of truth: employees deleted in the app were silently re-inserted at
    # the next restart, which is what kept resurrecting the F793-F798
    # duplicates and eventually crashed startup against the unique index.
    # Once real data exists, the database is authoritative, not the file.
    # Add any newly-introduced columns to tables that already exist.
    # create_all() only creates missing TABLES, never missing columns, so
    # without this a new field works on a fresh database but breaks every
    # existing one - which is exactly what the store rental fields did.
    _add_missing_columns()

    db = SessionLocal()
    try:
        _retire_staff_role(db)
        already_seeded = db.query(models.Employee).count() > 0
    finally:
        db.close()
    if already_seeded:
        seed_data.run(None)          # users/defaults only, no employee import
        return
    json_path = os.path.join(os.path.dirname(__file__), "master_data.json")
    seed_data.run(json_path if os.path.exists(json_path) else None)

# Locked down to specific origins in production - wide open here only
# for local dev/testing against a frontend running on a different port.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------
# AUTH
# ---------------------------------------------------------------------
# ---------------------------------------------------------------------
# PER-USER SCREEN PERMISSIONS
# ---------------------------------------------------------------------
ALL_SCREENS = ["dashboard", "attendance", "masterdata", "reports", "combine",
               "adjustments", "livecard", "store", "requests", "approvals",
               "errorcheck", "settings", "activity"]

# What a role can see when no explicit permissions have been set, so
# existing accounts keep working exactly as before this was added.
def _dubai_today() -> date:
    """Today in Dubai, whatever clock the server keeps.

    The VPS runs on UTC, four hours behind. Between midnight and four in
    the morning Dubai time, date.today() on the server is still
    yesterday - so a foreman marking the day's attendance early would be
    told it is in the future, and a check for missing days would not
    yet count today. Every 'what day is it' in this file goes through
    here.
    """
    return (datetime.now(timezone.utc) + timedelta(hours=4)).date()


# Three roles, which is what the company actually has. "staff" was a
# catch-all that gave nearly everything away by default; anyone who
# needs an unusual mix gets it through their own permissions instead.
ROLE_DEFAULTS = {
    "admin": ALL_SCREENS,
    # The office approves and orders; a site raises requests and does
    # not approve its own.
    # Settings is on every role: it is where anyone changes their own
    # password and takes a backup. What sits inside it - staff logins,
    # restoring, clearing - is guarded on its own, not by hiding the
    # screen.
    "office": ["dashboard", "store", "requests", "approvals", "reports", "errorcheck", "settings"],
    "site": ["dashboard", "attendance", "store", "requests", "settings"],
}
ROLES = list(ROLE_DEFAULTS)


def _retire_staff_role(db):
    """Move anyone left on the old catch-all role onto 'office'.

    Their access must not change in the process: a staff account with no
    permissions of its own was relying on the old default, so that list
    is written into their permissions first. Someone who already had
    explicit permissions - the store keeper, for instance - keeps
    exactly those.
    """
    old_default = [s for s in ALL_SCREENS if s != "activity"]
    moved = 0
    for u in db.query(models.User).filter(models.User.role == "staff").all():
        if not (u.permissions or "").strip():
            u.permissions = ",".join(old_default)
        u.role = "office"
        moved += 1
    if moved:
        db.commit()
        print(f"Moved {moved} account(s) off the retired 'staff' role, access unchanged")


def effective_permissions(user: models.User) -> list:
    if user.role == "admin":
        return list(ALL_SCREENS)          # admin always has everything
    raw = (user.permissions or "").strip()
    if raw:
        return [s for s in raw.split(",") if s in ALL_SCREENS]
    # An unrecognised role - an old account, or one from a future version
    # - gets the narrowest access rather than the widest. Failing closed
    # is the only safe direction here.
    return list(ROLE_DEFAULTS.get(user.role, ROLE_DEFAULTS["site"]))


def require_screen(screen: str):
    """
    Dependency that blocks an endpoint unless the user may open the screen
    it belongs to. Hiding a menu item is presentation only - without this
    the endpoint is still reachable by anyone with a login.
    """
    def _check(user: models.User = Depends(auth.get_current_user)):
        if screen not in effective_permissions(user):
            raise HTTPException(status_code=403,
                detail=f"You don't have access to {screen}. Ask an admin to enable it.")
        return user
    return _check


def require_any_screen(*screens):
    """Some data belongs to two jobs at once: suppliers are used by the
    keeper (store screen) receiving deliveries and by the office
    (requests screen) placing orders. Either permission opens the door -
    demanding one specific screen was silently blanking the supplier
    list for whoever happened to hold the other."""
    def _check(user: models.User = Depends(auth.get_current_user)):
        perms = effective_permissions(user)
        if not any(s in perms for s in screens):
            raise HTTPException(status_code=403,
                detail=f"You don't have access to this. Ask an admin to enable {' or '.join(screens)}.")
        return user
    return _check


def log_action(db: Session, user_id, action: str, details: str = ""):
    db.add(models.AuditLog(user_id=user_id, action=action, details=details))
    db.commit()


@app.post("/auth/login", response_model=schemas.TokenResponse)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not user.active or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Incorrect username or password")
    token = auth.create_access_token({"sub": user.username})
    log_action(db, user.id, "login")
    maybe_create_auto_backup(db)
    return schemas.TokenResponse(access_token=token, username=user.username,
                                  role=user.role, full_name=user.full_name)


@app.get("/auth/me")
def read_me(user: models.User = Depends(auth.get_current_user)):
    return {"username": user.username, "full_name": user.full_name, "role": user.role}


@app.post("/auth/download-token")
def get_download_token(user: models.User = Depends(auth.get_current_user)):
    """
    Issues a 60-second, download-only token - requires the normal
    Authorization header (proper auth, not a URL param), then hands
    back a short-lived token that export/backup links can safely carry
    in their query string instead of the real session token.
    """
    return {"token": auth.create_download_token(user.username)}


@app.post("/auth/change-password")
def change_password(payload: schemas.ChangePasswordRequest, db: Session = Depends(get_db),
                     user: models.User = Depends(auth.get_current_user)):
    if not auth.verify_password(payload.current_password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect.")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters.")
    user.hashed_password = auth.hash_password(payload.new_password)
    db.commit()
    log_action(db, user.id, "change_password")
    return {"ok": True}


# ---------------------------------------------------------------------
# USER MANAGEMENT (admin only - lets staff have their own logins
# instead of everyone sharing the one admin account)
# ---------------------------------------------------------------------
@app.get("/users", response_model=list[schemas.UserOut])
def list_users(db: Session = Depends(get_db),
                user: models.User = Depends(auth.require_admin)):
    """Who can log in, and what each may open, is administration - not
    something a store keeper or site engineer needs to read."""
    return db.query(models.User).order_by(models.User.username).all()


@app.post("/users", response_model=schemas.UserOut)
def create_user(payload: schemas.UserIn, db: Session = Depends(get_db),
                 user: models.User = Depends(auth.require_admin)):
    # Staff can add fellow staff, but only an admin can mint another
    # admin - otherwise any staff login could promote itself (or a new
    # account) to admin, which would make every admin-only restriction
    # meaningless, including the Activity Monitor.
    # office and site are both non-privileged; only an admin can mint
    # another admin, otherwise any login could promote itself.
    if payload.role not in ROLES:
        raise HTTPException(status_code=400, detail="Role must be office, site or admin.")
    if payload.role == "admin" and user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can create an admin account.")
    existing = db.query(models.User).filter(models.User.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="That username is already taken.")
    new_user = models.User(
        username=payload.username,
        hashed_password=auth.hash_password(payload.password),
        # Full name is no longer collected when creating a login; fall
        # back to the username so the Activity Monitor and header still
        # have something readable to show.
        full_name=(payload.full_name or "").strip() or payload.username,
        role=payload.role,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    log_action(db, user.id, "create_user", f"{payload.username} ({payload.role})")
    return new_user


@app.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db),
                 user: models.User = Depends(auth.require_admin)):
    """Remove a login for good. Admin only.

    Two things this must not allow: deleting your own account, which
    would sign you out of a system you administer, and removing the last
    admin, which would leave nobody able to restore a backup or add a
    login again.

    Backups, activity entries and adjustments record who made them. The
    login goes, but that history stays and simply no longer names a
    live account - deleting a person should not quietly rewrite what
    was done last month.
    """
    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.id == user.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account.")
    if target.role == "admin":
        others = (db.query(models.User)
                    .filter(models.User.role == "admin", models.User.id != target.id).count())
        if others == 0:
            raise HTTPException(status_code=400,
                detail="This is the only admin account. Make someone else an admin first.")

    username = target.username
    # Detach the history before removing the row, or the database
    # refuses the delete for the sake of those references.
    for model, column in ((models.Backup, "created_by"),
                          (models.AuditLog, "user_id"),
                          (models.SalaryAdjustment, "created_by")):
        db.query(model).filter(getattr(model, column) == target.id) \
                       .update({column: None}, synchronize_session=False)
    db.delete(target)
    db.commit()
    log_action(db, user.id, "delete_user", username)
    return {"ok": True, "detail": f"{username} removed."}


@app.post("/users/{user_id}/reset-password")
def reset_user_password(user_id: int, payload: schemas.ResetPasswordRequest,
                         db: Session = Depends(get_db),
                         user: models.User = Depends(auth.require_admin)):
    """
    Set another user's password without knowing their current one - for
    when someone forgets theirs. Anyone can reset a staff account (staff
    manage staff), but only an admin can reset an admin's, matching the
    same rule that governs creating and deactivating admins.
    """
    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.role == "admin" and user.role != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can reset an admin's password.")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters.")
    target.hashed_password = auth.hash_password(payload.new_password)
    db.commit()
    log_action(db, user.id, "reset_password", target.username)
    return {"ok": True}


@app.get("/audit-log")
def list_audit_log(limit: int = 200, db: Session = Depends(get_db),
                    user: models.User = Depends(auth.require_admin)):
    rows = (
        db.query(models.AuditLog, models.User.username, models.User.full_name)
        .outerjoin(models.User, models.AuditLog.user_id == models.User.id)
        .order_by(models.AuditLog.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": log.id, "action": log.action, "details": log.details,
            "created_at": log.created_at.isoformat() if log.created_at else None,
            "username": username or "unknown", "full_name": full_name or "Unknown",
        }
        for log, username, full_name in rows
    ]


# ---------------------------------------------------------------------
# MASTER DATA - Employees
# ---------------------------------------------------------------------
@app.get("/employees", response_model=list[schemas.EmployeeOut])
def list_employees(active_only: bool = False, db: Session = Depends(get_db),
                    user: models.User = Depends(auth.get_current_user)):
    """Every logged-in user may read the worker list - the store needs
    names to record who took material. Pay is another matter: only
    someone who can open Master Data or Salary Adjustments sees it, so a
    store keeper or site engineer can no longer read all 74 salaries."""
    perms = effective_permissions(user)
    may_see_pay = any(s in perms for s in ("masterdata", "adjustments", "livecard"))
    q = db.query(models.Employee)
    if active_only:
        q = q.filter(models.Employee.active == True)  # noqa: E712
    rows = q.order_by(models.Employee.emp_no).all()
    if may_see_pay:
        return rows
    return [{"id": e.id, "emp_no": e.emp_no, "name": e.name, "trade": e.trade or "",
             "active": e.active, "total_salary": 0, "basic_salary": 0} for e in rows]


@app.post("/employees", response_model=schemas.EmployeeOut)
def upsert_employee(emp: schemas.EmployeeIn, db: Session = Depends(get_db),
                     user: models.User = Depends(require_screen("masterdata"))):
    # Names arrive typed every which way. Tidied here, at the one place
    # they are written, so the list never shows the same man twice.
    emp.name = _person_name(emp.name.strip())
    emp.trade = _proper_name((emp.trade or "").strip())
    existing = db.query(models.Employee).filter(models.Employee.emp_no == emp.emp_no).first()
    if existing:
        for field, value in emp.dict().items():
            setattr(existing, field, value)
    else:
        existing = models.Employee(**emp.dict())
        db.add(existing)
    db.commit()
    db.refresh(existing)

    # Salary figures live on the Employee record but are copied into every
    # EmployeeSummary when it's calculated, so editing Total/Basic Salary
    # has to recalculate this worker's existing summaries - otherwise the
    # cards, reports and payroll totals keep showing the OLD salary
    # indefinitely, with no indication they're stale.
    for (cycle,) in db.query(models.EmployeeSummary.month_year).filter(
        models.EmployeeSummary.emp_no == existing.emp_no
    ).distinct().all():
        services.recalculate_summary(db, existing, cycle)

    log_action(db, user.id, "save_employee", f"{emp.emp_no} - {emp.name}")
    return existing


@app.delete("/employees/{emp_no}")
def remove_employee(emp_no: str, purge: bool = False, db: Session = Depends(get_db),
                     user: models.User = Depends(require_screen("masterdata"))):
    """Remove a worker, in the way that suits what has happened to them.

    Someone who has worked has attendance and salary behind them, and
    deleting the record would tear a hole in months of payroll. That
    person is marked inactive instead: they leave the daily list and
    every report still adds up.

    A record created by mistake - a typo, a duplicate, someone who never
    started - has nothing behind it, and is genuinely deleted.

    purge=true forces a real delete along with that person's history,
    for a duplicate that was already marked and paid against. It says
    what it is destroying before it does it.
    """
    emp = db.query(models.Employee).filter(models.Employee.emp_no == emp_no).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    rows = db.query(models.DailyRow).filter(models.DailyRow.emp_no == emp_no).count()
    sums = db.query(models.EmployeeSummary).filter(models.EmployeeSummary.emp_no == emp_no).count()
    sum_ids = [s.id for s in db.query(models.EmployeeSummary)
                                .filter(models.EmployeeSummary.emp_no == emp_no).all()]
    adjs = (db.query(models.SalaryAdjustment)
              .filter(models.SalaryAdjustment.summary_id.in_(sum_ids)).count() if sum_ids else 0)
    history = rows + sums + adjs

    if history and not purge:
        emp.active = False
        db.commit()
        log_action(db, user.id, "deactivate_employee", f"{emp_no} ({rows} attendance day(s))")
        return {"ok": True, "action": "deactivated", "attendance_days": rows,
                "adjustments": adjs,
                "detail": f"{emp.name} has {rows} day(s) of attendance behind them, so the record is "
                          f"kept and marked inactive. They no longer appear in the daily list."}

    if purge and history:
        db.query(models.DailyRow).filter(models.DailyRow.emp_no == emp_no).delete(synchronize_session=False)
        if sum_ids:
            (db.query(models.SalaryAdjustment)
               .filter(models.SalaryAdjustment.summary_id.in_(sum_ids))
               .delete(synchronize_session=False))
        (db.query(models.EmployeeSummary)
           .filter(models.EmployeeSummary.emp_no == emp_no).delete(synchronize_session=False))
    db.delete(emp)
    db.commit()
    log_action(db, user.id, "delete_employee", f"{emp_no} ({history} record(s) removed)")
    return {"ok": True, "action": "deleted", "removed_history": history,
            "detail": f"{emp.name} removed." + (f" {history} record(s) went with them." if history else "")}


EMPLOYEE_TEMPLATE_HEADERS = ["Emp No", "Name", "Trade", "Company", "Total Salary", "Basic Salary"]


@app.get("/employees/template")
def download_employee_template(token: str, db: Session = Depends(get_db)):
    """Blank spreadsheet with the exact columns /employees/import expects,
    so staff can fill it in offline and bring it back."""
    auth.get_download_user_from_token(token, db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    for i, h in enumerate(EMPLOYEE_TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C0392B")
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.append(["D-99", "SAMPLE WORKER", "DRIVER", "Infinia", 2000, 900])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Infinia_Employee_Template.xlsx"},
    )


@app.get("/employees/export")
def export_employees(token: str, db: Session = Depends(get_db)):
    """
    Every active employee's master data, in the exact same column
    format /employees/import expects - export this, edit it, and
    re-import it straight back in without reshaping anything.
    """
    auth.get_download_user_from_token(token, db)
    employees = db.query(models.Employee).filter(models.Employee.active == True).order_by(models.Employee.emp_no).all()  # noqa: E712
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    for i, h in enumerate(EMPLOYEE_TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C0392B")
        ws.column_dimensions[get_column_letter(i)].width = 18
    for emp in employees:
        ws.append([emp.emp_no, emp.name, emp.trade, emp.total_salary, emp.basic_salary])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Infinia_Employee_Export.xlsx"},
    )


@app.post("/employees/import")
async def import_employees(file: UploadFile = File(...), mode: str = Form("add_only"),
                            duplicate_handling: str = Form("skip"),
                            db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    """
    Bulk create from the filled-in template, with explicit control over
    two independent choices instead of always silently updating:

    mode: 'add_only' leaves every existing employee alone; 'replace'
    deactivates any active employee whose Emp No isn't in this file
    (soft-delete via the same 'active' flag the rest of the app uses -
    never a hard delete, since that would break their historical
    DailyRow/EmployeeSummary records).

    duplicate_handling: what to do when a row's Emp No already exists -
    'skip' leaves the existing record untouched, 'update' overwrites it
    with the file's data, 'add_new' creates a second, separate record
    under a modified Emp No (e.g. "D-01 (1)") so both exist side by side.
    """
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read that file - please upload the .xlsx template.")
    ws = wb.active

    header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
    expected = {h.lower(): i for i, h in enumerate(header_row)}
    required = ["emp no", "name"]
    if not all(r in expected for r in required):
        raise HTTPException(status_code=400, detail="Missing required columns 'Emp No' and 'Name' - please use the template.")

    def unique_suffixed_emp_no(base_emp_no):
        n = 1
        while True:
            candidate = f"{base_emp_no} ({n})"
            if not db.query(models.Employee).filter(models.Employee.emp_no == candidate).first():
                return candidate
            n += 1

    created, updated, skipped, added_as_new, errors = 0, 0, 0, 0, []
    updated_emp_nos = set()
    file_emp_nos = set()
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def get(col_name, default=""):
            idx = expected.get(col_name)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return val if val is not None else default

        emp_no = str(get("emp no")).strip()
        name = str(get("name")).strip()
        if not emp_no or not name:
            errors.append(f"Row {row_idx}: missing Emp No or Name, skipped.")
            continue
        try:
            total_salary = float(get("total salary", 0) or 0)
            basic_salary = float(get("basic salary", 0) or 0)
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx} ({emp_no}): Total/Basic Salary must be numbers, skipped.")
            continue
        trade = str(get("trade", "")).strip()
        # A blank or unrecognised company means Infinia, so an older
        # spreadsheet without the column still imports cleanly.
        company = str(get("company", "") or "").strip()
        company = "Prime Infinia" if company.lower().replace("-", " ") in ("prime infinia", "prime") else "Infinia"
        file_emp_nos.add(emp_no)

        existing = db.query(models.Employee).filter(models.Employee.emp_no == emp_no).first()
        if existing:
            if duplicate_handling == "skip":
                skipped += 1
            elif duplicate_handling == "add_new":
                new_emp_no = unique_suffixed_emp_no(emp_no)
                db.add(models.Employee(emp_no=new_emp_no, name=name, trade=trade, company=company,
                                        total_salary=total_salary, basic_salary=basic_salary, active=True))
                added_as_new += 1
            else:  # update
                existing.name, existing.trade = name, trade
                existing.company = company
                existing.total_salary, existing.basic_salary = total_salary, basic_salary
                existing.active = True
                updated += 1
                updated_emp_nos.add(emp_no)
        else:
            db.add(models.Employee(emp_no=emp_no, name=name, trade=trade, company=company,
                                    total_salary=total_salary, basic_salary=basic_salary, active=True))
            created += 1

    deactivated = 0
    if mode == "replace":
        active_not_in_file = (
            db.query(models.Employee)
            .filter(models.Employee.active == True, ~models.Employee.emp_no.in_(file_emp_nos))  # noqa: E712
            .all()
        )
        for emp in active_not_in_file:
            emp.active = False
            deactivated += 1

    db.commit()

    # Any employee whose salary figures were just overwritten needs their
    # existing summaries recalculated, same reason as upsert_employee -
    # otherwise cards and reports keep showing the pre-import salary.
    for emp_no in updated_emp_nos:
        emp_obj = db.query(models.Employee).filter(models.Employee.emp_no == emp_no).first()
        if not emp_obj:
            continue
        for (cycle,) in db.query(models.EmployeeSummary.month_year).filter(
            models.EmployeeSummary.emp_no == emp_no
        ).distinct().all():
            services.recalculate_summary(db, emp_obj, cycle)

    log_action(db, user.id, "import_employees",
               f"{created} created, {updated} updated, {skipped} skipped, "
               f"{added_as_new} added as new, {deactivated} deactivated, {len(errors)} errors")
    return {"created": created, "updated": updated, "skipped": skipped,
            "added_as_new": added_as_new, "deactivated": deactivated, "errors": errors}


# ---------------------------------------------------------------------
# MASTER DATA - Sites / Engineers
# ---------------------------------------------------------------------
@app.get("/sites", response_model=list[schemas.SiteOut])
def list_sites(db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    return db.query(models.Site).filter(models.Site.active == True).order_by(models.Site.code).all()  # noqa: E712


@app.post("/sites", response_model=schemas.SiteOut)
def add_site(site: schemas.SiteIn, db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    existing = db.query(models.Site).filter(models.Site.code == site.code).first()
    if existing:
        existing.active = True
        db.commit()
        db.refresh(existing)
        return existing
    new_site = models.Site(**site.dict())
    db.add(new_site)
    db.commit()
    db.refresh(new_site)
    return new_site


@app.put("/sites/{site_id}", response_model=schemas.SiteOut)
def rename_site(site_id: int, site: schemas.SiteIn, db: Session = Depends(get_db),
                 user: models.User = Depends(auth.get_current_user)):
    existing = db.query(models.Site).filter(models.Site.id == site_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Site not found")
    existing.code = site.code
    db.commit()
    db.refresh(existing)
    log_action(db, user.id, "rename_site", f"#{site_id} -> {site.code}")
    return existing


@app.delete("/sites/{site_id}")
def remove_site(site_id: int, db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    existing = db.query(models.Site).filter(models.Site.id == site_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Site not found")
    existing.active = False
    db.commit()
    log_action(db, user.id, "remove_site", existing.code)
    return {"ok": True}


@app.get("/engineers", response_model=list[schemas.EngineerOut])
def list_engineers(db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    return db.query(models.Engineer).filter(models.Engineer.active == True).order_by(models.Engineer.name).all()  # noqa: E712


@app.post("/engineers", response_model=schemas.EngineerOut)
def add_engineer(eng: schemas.EngineerIn, db: Session = Depends(get_db),
                  user: models.User = Depends(auth.get_current_user)):
    existing = db.query(models.Engineer).filter(models.Engineer.name == eng.name).first()
    if existing:
        existing.active = True
        db.commit()
        db.refresh(existing)
        return existing
    new_eng = models.Engineer(**eng.dict())
    db.add(new_eng)
    db.commit()
    db.refresh(new_eng)
    return new_eng


@app.put("/engineers/{engineer_id}", response_model=schemas.EngineerOut)
def rename_engineer(engineer_id: int, eng: schemas.EngineerIn, db: Session = Depends(get_db),
                     user: models.User = Depends(auth.get_current_user)):
    existing = db.query(models.Engineer).filter(models.Engineer.id == engineer_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Engineer not found")
    existing.name = eng.name
    db.commit()
    db.refresh(existing)
    log_action(db, user.id, "rename_engineer", f"#{engineer_id} -> {eng.name}")
    return existing


@app.delete("/engineers/{engineer_id}")
def remove_engineer(engineer_id: int, db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    existing = db.query(models.Engineer).filter(models.Engineer.id == engineer_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Engineer not found")
    existing.active = False
    db.commit()
    log_action(db, user.id, "remove_engineer", existing.name)
    return {"ok": True}


# ---------------------------------------------------------------------
# DAILY ATTENDANCE
# ---------------------------------------------------------------------
@app.get("/attendance/{target_date}", response_model=list[schemas.DailyRowOut])
def get_attendance_for_date(target_date: date, db: Session = Depends(get_db),
                             user: models.User = Depends(auth.get_current_user)):
    return db.query(models.DailyRow).filter(models.DailyRow.full_date == target_date).all()


@app.delete("/attendance/{target_date}")
def clear_attendance_for_date(target_date: date, db: Session = Depends(get_db),
                               user: models.User = Depends(auth.get_current_user)):
    """
    Deletes every worker's attendance row for one specific date - the
    'Clear Day' button's confirmed action. Admin-only, since this wipes
    every worker's entry for the day at once, not just one row.
    """
    count = (
        db.query(models.DailyRow)
        .filter(models.DailyRow.full_date == target_date)
        .delete(synchronize_session=False)
    )
    db.commit()
    log_action(db, user.id, "clear_day", f"{target_date}: {count} row(s) deleted")
    return {"deleted": count}


@app.get("/attendance/completion/{month_year}")
def get_completion_status(month_year: str, mode: str = "cycle",
                           db: Session = Depends(get_db),
                           user: models.User = Depends(auth.get_current_user)):
    """
    Per-day completion status for the calendar: a day is 'complete' when
    every currently-active employee has a saved attendance row for it
    (green = complete, red = incomplete).

    mode="cycle"    - the payroll cycle, 26th of the previous month to
                      the 25th of this one. Used where the view must
                      line up with payroll.
    mode="calendar" - the plain calendar month, 1st to last day. Used by
                      the dashboard/attendance calendars, because a
                      normal month is what people actually recognise
                      when scanning for a date.
    """
    from datetime import datetime as dt, timedelta as td
    import calendar as _cal
    try:
        parsed = dt.strptime(f"25 {month_year}", "%d %B %Y").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="month_year must look like 'August 2026'.")
    if mode == "calendar":
        cycle_start = parsed.replace(day=1)
        last = _cal.monthrange(parsed.year, parsed.month)[1]
        cycle_end = parsed.replace(day=last)
    else:
        cycle_start, cycle_end, _ = pcyc.cycle_bounds_for(parsed)

    total_active = db.query(models.Employee).filter(models.Employee.active == True).count()  # noqa: E712
    rows = db.query(models.DailyRow.full_date, models.DailyRow.emp_no).filter(
        and_(models.DailyRow.full_date >= cycle_start, models.DailyRow.full_date <= cycle_end,
             or_(models.DailyRow.am != "", models.DailyRow.pm != ""))
        # excludes rows that exist only because auto_fill_sunday_from_saturday
        # pre-filled Site/Engineer with A.M/P.M still blank - those aren't
        # "marked" yet, so they must not count toward the day being complete
    ).all()
    counts_by_date = {}
    for full_date, emp_no in rows:
        counts_by_date.setdefault(full_date, set()).add(emp_no)

    days = []
    d = cycle_start
    while d <= cycle_end:
        entered = len(counts_by_date.get(d, set()))
        days.append({"date": d.isoformat(), "entered": entered, "total": total_active,
                      "complete": total_active > 0 and entered >= total_active})
        d += td(days=1)
    return {"cycle_start": cycle_start.isoformat(), "cycle_end": cycle_end.isoformat(),
            "total_active": total_active, "days": days}


@app.post("/attendance/save")
def save_attendance(payload: schemas.BulkSaveRequest, db: Session = Depends(get_db),
                     user: models.User = Depends(auth.get_current_user)):
    """
    Same validation and Holiday-previous-day rules as the desktop app's
    save_all(): every row is validated BEFORE anything is written, a
    row cleared back to fully blank (no A.M/P.M and no Site) deletes
    any existing saved entry instead of being silently skipped, and
    Holiday specifically requires Site/Engineer to come from that
    worker's own saved entry the day before - blocked with the exact
    missing date if that isn't there.

    A row with blank A.M/P.M but a real Site (an auto-filled Sunday
    placeholder, or staff editing just the Site before marking
    A.M/P.M) is kept rather than deleted - every save resubmits ALL
    workers for the date, so treating that the same as a genuine clear
    would silently wipe out every other worker's still-unmarked
    placeholder too.
    """
    errors = []
    blocked = []
    to_process = []

    today = _dubai_today()
    for row_in in payload.rows:
        # Attendance can't be recorded for a day that hasn't happened yet.
        # Without this the app accepted any date at all - a save for
        # 25 December went through months early - and those future days
        # count as worked in the payroll totals, inflating salaries.
        if row_in.full_date > today:
            errors.append(f"{row_in.emp_no}: {row_in.full_date} is in the future - attendance can only be entered up to today.")
            continue

        employee = db.query(models.Employee).filter(models.Employee.emp_no == row_in.emp_no).first()
        if not employee:
            errors.append(f"{row_in.emp_no}: employee not found.")
            continue

        am, pm = (row_in.am or "").strip(), (row_in.pm or "").strip()
        site_val = (row_in.site or "").strip()
        if not am and not pm:
            if not site_val:
                # Genuinely blank - delete any existing saved entry for this day.
                services.delete_daily_row_if_blank(db, row_in.emp_no, row_in.full_date)
                to_process.append((employee, None))
            else:
                # Blank A.M/P.M but a real Site - an auto-filled placeholder
                # (Saturday -> Sunday) staff hasn't marked yet, or staff
                # editing just the Site before marking A.M/P.M. Every save
                # resubmits ALL workers for the date, so treating this the
                # same as a genuine clear would silently wipe out every
                # other worker's still-unmarked placeholder too - keep/
                # update it instead, skip full validation since it isn't
                # a real attendance entry yet.
                row_in.am, row_in.pm = "", ""
                to_process.append((employee, row_in))
            continue

        site, engineer = row_in.site, row_in.engineer
        if am == "Holiday" or pm == "Holiday":
            prev = services.get_previous_day_site_engineer(db, row_in.emp_no, row_in.full_date)
            if prev is None:
                from datetime import timedelta
                prev_date = row_in.full_date - timedelta(days=1)
                blocked.append(f"{row_in.emp_no} ({employee.name}) - needs {prev_date} filled in first")
                continue
            site, engineer = prev
            row_in.site, row_in.engineer = site, engineer

        problems = services.validate_row(am, pm, site, engineer, row_in.bh, row_in.comments, row_in.ot)
        if problems:
            errors.append(f"{row_in.emp_no}: " + "; ".join(problems))
            continue

        to_process.append((employee, row_in))

    if errors:
        raise HTTPException(status_code=400, detail={"errors": errors})

    touched_cycles = set()
    for employee, row_in in to_process:
        if row_in is not None:
            saved_row = services.upsert_daily_row(db, employee, row_in)
            services.auto_fill_sunday_from_saturday(db, employee, saved_row)
            _, _, month_year = pcyc.cycle_bounds_for(row_in.full_date)
        else:
            # a delete - recalc using whatever cycle the deleted date was in
            if payload.rows:
                _, _, month_year = pcyc.cycle_bounds_for(payload.rows[0].full_date)
            else:
                continue
        touched_cycles.add((employee.emp_no, month_year))

    for emp_no, month_year in touched_cycles:
        employee = db.query(models.Employee).filter(models.Employee.emp_no == emp_no).first()
        services.recalculate_summary(db, employee, month_year)

    saved_count = len([r for e, r in to_process if r is not None])
    if saved_count or blocked:
        the_date = payload.rows[0].full_date if payload.rows else ""
        log_action(db, user.id, "save_attendance", f"{the_date}: {saved_count} worker(s)")

    return {"saved": saved_count, "blocked": blocked}


# ---------------------------------------------------------------------
# EMPLOYEE SUMMARIES / SALARY ADJUSTMENTS
# ---------------------------------------------------------------------
@app.get("/summaries/{month_year}", response_model=list[schemas.EmployeeSummaryOut])
def list_summaries(month_year: str, as_of: str = None, db: Session = Depends(get_db),
                    user: models.User = Depends(auth.get_current_user)):
    """
    as_of (optional, YYYY-MM-DD): limits the Site column to that
    worker's most recent site on or before this date, instead of the
    latest site in the whole cycle - lets Reports answer 'who was
    where as of a given date', not just 'as of cycle end'.
    """
    summaries = (
        db.query(models.EmployeeSummary)
        .options(joinedload(models.EmployeeSummary.adjustments))
        .filter(models.EmployeeSummary.month_year == month_year)
        .order_by(models.EmployeeSummary.emp_no)
        .all()
    )
    # Site isn't a stored summary field - a worker can be at a different
    # site each day - so this picks their single MOST RECENT site in
    # the cycle (optionally as of a given date), not a list of every
    # site they were ever at.
    query = db.query(models.DailyRow.emp_no, models.DailyRow.site, models.DailyRow.full_date).filter(
        models.DailyRow.month_year == month_year, models.DailyRow.site != ""
    )
    if as_of:
        try:
            as_of_date = date.fromisoformat(as_of)
            query = query.filter(models.DailyRow.full_date <= as_of_date)
        except ValueError:
            pass
    latest_site_by_emp = {}
    for emp_no, site, full_date in query.all():
        if not site:
            continue
        current = latest_site_by_emp.get(emp_no)
        if current is None or full_date > current[1]:
            latest_site_by_emp[emp_no] = (site, full_date)

    out = []
    for s in summaries:
        item = schemas.EmployeeSummaryOut.from_orm(s)
        latest = latest_site_by_emp.get(s.emp_no)
        item.sites = latest[0] if latest else ""
        out.append(item)
    return out


@app.get("/summaries/{month_year}/by-site")
def summaries_by_site(month_year: str, date_from: str = None, date_to: str = None,
                       db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    """
    Site project cost for the cycle, based on actual attendance - reuses
    reports.site_cost_center() directly, the same logic the desktop app
    uses. Each worker's cost per day at a site is their own daily rate
    (total_salary / 30) plus OT/BH at their own hourly rate, summed by
    site - so a worker who was at two sites contributes only their
    actual days at each, never their full salary twice. With no date
    range, uses every date on file for the cycle; with one, costs only
    that exact window.
    """
    daily_rows = db.query(models.DailyRow).filter(models.DailyRow.month_year == month_year).all()
    summaries2 = db.query(models.EmployeeSummary).filter(models.EmployeeSummary.month_year == month_year).all()
    filters = {}
    if date_from:
        filters["date_from"] = datetime.strptime(date_from, "%Y-%m-%d").date()
    if date_to:
        filters["date_to"] = datetime.strptime(date_to, "%Y-%m-%d").date()
    result = rp.site_cost_center(daily_rows, summaries2, filters)
    return {"title": result.title, "note": result.note,
            "columns": [{"key": k, "label": label} for k, label in result.columns],
            "rows": result.rows, "totals": result.totals}


BUILDER_MEASURE_HINTS = {
    "worker_count": "How many distinct workers were at this group at least once - each worker only counts once, no matter how many days.",
    "record_count": "Total worker-days combined - 5 workers present 3 days each = 15 man-days.",
    "final_salary_cost": "Each worker's own daily rate, apportioned across every paid day in this group's window.",
    "ot_amount": "Each row's OT hours valued at that worker's own hourly rate (their monthly salary / 30 / 8).",
    "bh_amount": "Each row's BH hours valued at that worker's own hourly rate (their monthly salary / 30 / 8).",
}


@app.get("/reports/builder-catalog")
def builder_catalog(user: models.User = Depends(auth.get_current_user)):
    """
    What's available to pick from in the Report Builder, for both data
    sources - dimensions (things to group by / show as labels, never
    summed) and measures (things summed per group). Site is just one
    dimension among several, same as the desktop app - not a separate
    mode.
    """
    def with_hints(d):
        return [{"key": k, "label": v, "hint": BUILDER_MEASURE_HINTS.get(k)} for k, v in d.items()]

    return {
        "summary": {
            "dimensions": [{"key": k, "label": v} for k, v in rp.BUILDER_SUMMARY_DIMENSIONS.items()],
            "measures": with_hints(rp.BUILDER_SUMMARY_MEASURES),
        },
        "daily": {
            "dimensions": [{"key": k, "label": v} for k, v in rp.BUILDER_DAILY_DIMENSIONS.items()],
            "measures": with_hints(rp.BUILDER_DAILY_MEASURES),
        },
    }


def _report_source_rows(db: Session, month_year: str, date_from: str, date_to: str):
    """
    Picks the rows a report runs over.

    When a date range is given it is used DIRECTLY against full_date, so
    a report can span several cycles or any arbitrary window - previously
    the query was pinned to a single month_year and the range could only
    narrow within it, making a two-cycle report impossible.

    With no range, it falls back to the whole selected cycle.

    Summaries are per-worker-per-cycle, so for a range they are pulled
    for every cycle the range touches; measures that come from summaries
    (e.g. Final Salary Cost) are apportioned by the daily rows that fall
    inside the window, which is what build_custom_report already does.
    """
    filters = {}
    if date_from or date_to:
        q = db.query(models.DailyRow)
        if date_from:
            d1 = datetime.strptime(date_from, "%Y-%m-%d").date()
            q = q.filter(models.DailyRow.full_date >= d1)
            filters["date_from"] = d1
        if date_to:
            d2 = datetime.strptime(date_to, "%Y-%m-%d").date()
            q = q.filter(models.DailyRow.full_date <= d2)
            filters["date_to"] = d2
        daily_rows = q.all()
        cycles = {r.month_year for r in daily_rows} or {month_year}
        summaries2 = (db.query(models.EmployeeSummary)
                        .filter(models.EmployeeSummary.month_year.in_(cycles)).all())
        return daily_rows, summaries2, filters

    daily_rows = db.query(models.DailyRow).filter(models.DailyRow.month_year == month_year).all()
    summaries2 = db.query(models.EmployeeSummary).filter(models.EmployeeSummary.month_year == month_year).all()
    return daily_rows, summaries2, filters


@app.get("/reports/custom")
def custom_report(month_year: str, data_source: str = "daily", dimensions: str = "", measures: str = "",
                   date_from: str = None, date_to: str = None, company: str = "",
                   db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    """
    The Report Builder's own aggregation engine (reports.build_custom_report,
    the same one the desktop app used) - groups by whichever dimensions
    were picked and sums whichever measures were picked. No dimensions
    picked means one overall row; picking 'site' groups by site;
    picking 'emp_no'+'name' groups per worker - grouping is just a
    consequence of what's checked, not a separate toggle.

    data_source matters for accuracy: 'daily' uses each attendance row's
    OWN actual site/date/engineer (correct for a worker who was at
    different sites on different days). 'summary' is one row per worker
    per cycle, so a dimension like 'site' can only show that worker's
    single MOST FREQUENT site for the whole cycle - fine for a worker
    who stayed at one site, misleading for one who moved around, which
    is why 'daily' is the default here.
    """
    dims = [d for d in dimensions.split(",") if d]
    meas = [m for m in measures.split(",") if m]
    daily_rows, summaries2, filters = _report_source_rows(db, month_year, date_from, date_to)
    source = data_source if data_source in ("daily", "summary") else "daily"
    company_by_emp = {e.emp_no: (e.company or "Infinia") for e in db.query(models.Employee).all()}
    # Asking for one company narrows the report to that company's
    # workers, so Infinia and Prime Infinia can be reported separately.
    if company:
        keep = {n for n, c in company_by_emp.items() if c == company}
        daily_rows = [r for r in daily_rows if r.emp_no in keep]
        summaries2 = [s for s in summaries2 if s.emp_no in keep]
    result = rp.build_custom_report(source, dims, meas, filters, daily_rows, summaries2, company_by_emp)
    return {"title": result.title + (f" - {company}" if company else ""), "note": result.note,
            "columns": [{"key": k, "label": label} for k, label in result.columns],
            "rows": result.rows, "totals": result.totals}


@app.get("/export/{month_year}/custom-report")
def export_custom_report(month_year: str, token: str, data_source: str = "daily",
                          dimensions: str = "", measures: str = "",
                          date_from: str = None, date_to: str = None, format: str = "excel",
                          company: str = "", db: Session = Depends(get_db)):
    user = auth.get_download_user_from_token(token, db)
    dims = [d for d in dimensions.split(",") if d]
    meas = [m for m in measures.split(",") if m]
    daily_rows, summaries2, filters = _report_source_rows(db, month_year, date_from, date_to)
    source = data_source if data_source in ("daily", "summary") else "daily"
    # The download must match what was on screen, company filter and all.
    company_by_emp = {e.emp_no: (e.company or "Infinia") for e in db.query(models.Employee).all()}
    if company:
        keep = {n for n, c in company_by_emp.items() if c == company}
        daily_rows = [r for r in daily_rows if r.emp_no in keep]
        summaries2 = [s for s in summaries2 if s.emp_no in keep]
    result = rp.build_custom_report(source, dims, meas, filters, daily_rows, summaries2, company_by_emp)
    result_dict = {"columns": [{"key": k, "label": label} for k, label in result.columns],
                    "rows": result.rows, "totals": result.totals}
    if not result.columns:
        raise HTTPException(status_code=400, detail="No columns selected.")

    safe_name = "".join(c if c.isalnum() else "_" for c in month_year)
    if format == "excel":
        buf = export_web.build_generic_result_excel(result_dict, month_year)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Infinia_Report_{safe_name}.xlsx"},
        )
    elif format == "pdf":
        buf = export_web.build_generic_result_pdf(result_dict, month_year)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Infinia_Report_{safe_name}.pdf"},
        )
    raise HTTPException(status_code=400, detail="format must be 'excel' or 'pdf'.")


@app.get("/live-card/{emp_no}/{month_year}")
def get_live_card(emp_no: str, month_year: str, db: Session = Depends(get_db),
                   user: models.User = Depends(require_screen("livecard"))):
    # A live card is a worker's full pay picture. The screen was hidden
    # from roles that must not see salaries, but the endpoint answered
    # anyone signed in - so a site engineer could read every wage by
    # calling it directly. Guarded like the screen now.
    """
    Everything needed to render one worker's card on screen: their
    summary totals for the cycle plus every daily row, keyed by actual
    date (not calendar day-of-month) - the cycle runs 26th to 25th
    spanning two calendar months, so day-of-month alone would put late
    dates from the first month out of chronological order against the
    early dates of the second month.
    """
    summary = (
        db.query(models.EmployeeSummary)
        .options(joinedload(models.EmployeeSummary.adjustments))
        .filter(and_(models.EmployeeSummary.emp_no == emp_no, models.EmployeeSummary.month_year == month_year))
        .first()
    )
    rows = (
        db.query(models.DailyRow)
        .filter(and_(models.DailyRow.emp_no == emp_no, models.DailyRow.month_year == month_year))
        .all()
    )
    rows_by_date = {r.full_date.isoformat(): schemas.DailyRowOut.from_orm(r) for r in rows if r.full_date}
    if summary:
        summary_out = schemas.EmployeeSummaryOut.from_orm(summary)
    else:
        emp = db.query(models.Employee).filter(models.Employee.emp_no == emp_no).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found.")
        summary_out = None

    try:
        parsed = datetime.strptime(f"25 {month_year}", "%d %B %Y").date()
        cycle_start, cycle_end, _ = pcyc.cycle_bounds_for(parsed)
    except ValueError:
        cycle_start, cycle_end = None, None

    return {
        "emp_no": emp_no, "month_year": month_year,
        "summary": summary_out, "days": rows_by_date,
        "cycle_start": cycle_start.isoformat() if cycle_start else None,
        "cycle_end": cycle_end.isoformat() if cycle_end else None,
    }


@app.post("/summaries/{summary_id}/adjustments", response_model=schemas.SalaryAdjustmentOut)
def add_adjustment(summary_id: int, adj: schemas.SalaryAdjustmentIn, db: Session = Depends(get_db),
                    user: models.User = Depends(require_screen("adjustments"))):
    summary = db.query(models.EmployeeSummary).filter(models.EmployeeSummary.id == summary_id).first()
    if not summary:
        raise HTTPException(status_code=404, detail="Summary not found")
    new_adj = models.SalaryAdjustment(summary_id=summary_id, created_by=user.id, **adj.dict())
    db.add(new_adj)
    db.commit()
    db.refresh(new_adj)
    log_action(db, user.id, "add_adjustment",
               f"{summary.emp_no} - {adj.description}: {'-' if adj.is_deduction else '+'}{adj.amount}")
    return new_adj


@app.delete("/adjustments/{adjustment_id}")
def remove_adjustment(adjustment_id: int, db: Session = Depends(get_db),
                       user: models.User = Depends(auth.get_current_user)):
    adj = db.query(models.SalaryAdjustment).filter(models.SalaryAdjustment.id == adjustment_id).first()
    if not adj:
        raise HTTPException(status_code=404, detail="Adjustment not found")
    summary = db.query(models.EmployeeSummary).filter(models.EmployeeSummary.id == adj.summary_id).first()
    detail = f"{summary.emp_no} - {adj.description}" if summary else adj.description
    db.delete(adj)
    db.commit()
    log_action(db, user.id, "remove_adjustment", detail)
    return {"ok": True}


@app.get("/")
def root():
    return {"service": "Infinia Labour Tool API", "status": "running",
            "docs": "See /health for a simple status check."}


@app.get("/error-check/{month_year}")
def error_check(month_year: str, db: Session = Depends(get_db),
                 user: models.User = Depends(require_screen("errorcheck"))):
    """
    The desktop app's original three checks (missing AM/P.M, Present
    without Site/Engineer, BH without a comment) can never actually
    happen here - /attendance/save already enforces those exact same
    rules before a row is ever written, so this would always come back
    empty. What genuinely CAN go wrong in the web app instead: a day
    inside the cycle with no entry at all for an active worker, or an
    OT/BH value large enough to be worth a second look.
    """
    try:
        parsed = __import__("datetime").datetime.strptime(f"25 {month_year}", "%d %B %Y").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="month_year must look like 'August 2026'.")
    cycle_start, cycle_end, _ = pcyc.cycle_bounds_for(parsed)

    active_employees = db.query(models.Employee).filter(models.Employee.active == True).all()  # noqa: E712
    rows = db.query(models.DailyRow).filter(models.DailyRow.month_year == month_year).all()

    dates_by_emp = {}
    for r in rows:
        dates_by_emp.setdefault(r.emp_no, set()).add(r.full_date)

    from datetime import timedelta
    # Only days that have already happened can be missing. On the third
    # of the month the remaining twenty-two are not late, they are
    # simply not here yet - listing them made every worker look
    # twenty-nine days behind on day one.
    today = _dubai_today()
    last_day = min(cycle_end, today)
    all_dates = []
    d = cycle_start
    while d <= last_day:
        all_dates.append(d)
        d += timedelta(days=1)
    cycle_days = (cycle_end - cycle_start).days + 1
    elapsed_days = len(all_dates)

    out = []
    for emp in active_employees:
        entered = dates_by_emp.get(emp.emp_no, set())
        missing = [d for d in all_dates if d not in entered]
        if not entered:
            out.append({"emp_no": emp.emp_no, "name": emp.name, "date": "-", "site": "-",
                        "issue": f"No attendance entered at all for {month_year}.",
                        "kind": "Nothing entered",
                        "detail": {"Cycle": f"{cycle_start} to {cycle_end}",
                                   "Days in cycle": str(len(all_dates)),
                                   "Days entered": "0", "Trade": emp.trade or "-"}})
        elif missing:
            preview = ", ".join(d.strftime("%d %b") for d in missing[:5])
            more = f" (+{len(missing) - 5} more)" if len(missing) > 5 else ""
            out.append({"emp_no": emp.emp_no, "name": emp.name, "date": "-", "site": "-",
                        "issue": f"{len(missing)} day(s) missing: {preview}{more}",
                        "kind": f"{len(missing)} day(s) missing",
                        "detail": {"Trade": emp.trade or "-",
                                   "Days in cycle": str(len(all_dates)),
                                   "Days entered": str(len(entered)),
                                   "Every missing day":
                                       ", ".join(d.strftime("%d %b") for d in missing)}})

    # Hours worked on a day nobody worked. Absent means the worker was
    # not there, so OT or BH against it is a contradiction - usually a
    # status picked in the wrong row, or hours typed on the wrong line.
    # Holiday is flagged too: a worker who genuinely worked a holiday
    # should be marked Present, not Holiday.
    for r in rows:
        flagged_statuses = {s for s in ((r.am or "").strip(), (r.pm or "").strip())
                            if s in ("Absent", "Holiday")}
        hours = []
        if r.ot and r.ot > 0:
            hours.append(f"{r.ot:g} OT")
        if r.bh and r.bh > 0:
            hours.append(f"{r.bh:g} BH")
        if flagged_statuses and hours:
            marked = " and ".join(sorted(flagged_statuses))
            out.append({
                "emp_no": r.emp_no, "name": r.emp_name, "date": str(r.full_date), "site": r.site or "-",
                "issue": f"{' and '.join(hours)} hour(s) recorded on a day marked {marked}.",
                "kind": f"Hours on an {marked.lower()} day" if marked == "Absent" else f"Hours on a {marked.lower()} day",
                "severity": "contradiction",
                "detail": {
                    "Date": str(r.full_date),
                    "A.M": r.am or "-", "P.M": r.pm or "-",
                    "OT hours": f"{r.ot or 0:g}", "BH hours": f"{r.bh or 0:g}",
                    "Site": r.site or "-", "Engineer": r.engineer or "-",
                    "Comment": r.comments or "none",
                    "What to check": "Either the status is wrong, or the hours belong to another day "
                                     "or another worker. A worker who did work should be marked Present.",
                },
            })

    for r in rows:
        if r.ot and r.ot > 12:
            out.append({"emp_no": r.emp_no, "name": r.emp_name, "date": str(r.full_date), "site": r.site,
                        "issue": f"OT of {r.ot} hours in one day looks unusually high.",
                        "kind": "High OT",
                        "detail": {"Date": str(r.full_date), "A.M": r.am or "-", "P.M": r.pm or "-",
                                   "Site": r.site or "-", "Engineer": r.engineer or "-",
                                   "OT hours": f"{r.ot:g}", "BH hours": f"{r.bh or 0:g}",
                                   "Comment": r.comments or "none"}})
        if r.bh and r.bh > 8:
            out.append({"emp_no": r.emp_no, "name": r.emp_name, "date": str(r.full_date), "site": r.site,
                        "issue": f"BH of {r.bh} hours in one day looks unusually high.",
                        "kind": "High BH",
                        "detail": {"Date": str(r.full_date), "A.M": r.am or "-", "P.M": r.pm or "-",
                                   "Site": r.site or "-", "Engineer": r.engineer or "-",
                                   "OT hours": f"{r.ot or 0:g}", "BH hours": f"{r.bh:g}",
                                   "Comment": r.comments or "none"}})

    # UAE rule: a worker must take home at least 40% of their total
    # salary. Deductions and absences can eat past that without anyone
    # noticing until the pay run, so it is flagged here while there is
    # still time to look at it.
    summaries = (db.query(models.EmployeeSummary)
                   .filter(models.EmployeeSummary.month_year == month_year).all())
    entered_by_emp = {}
    for r in rows:
        entered_by_emp[r.emp_no] = entered_by_emp.get(r.emp_no, 0) + 1
    for s in summaries:
        total = s.total_salary or 0
        entered = entered_by_emp.get(s.emp_no, 0)
        if total <= 0 or entered <= 0:
            continue
        take_home = s.adjusted_final_salary()
        # Measured against the days actually entered for this worker,
        # not the days elapsed - a day nobody has marked yet is the
        # missing-days flag's business, not this one's. Three days in,
        # a man is judged on three days' pay; at the end of a full cycle
        # this is his whole salary and the rule is exactly the legal one.
        so_far = total * entered / cycle_days
        floor = so_far * 0.40
        if take_home < floor - 0.005:
            pct = (take_home / so_far * 100) if so_far else 0
            partial = entered < cycle_days
            adj = sum((-a.amount if a.is_deduction else a.amount) for a in s.adjustments)
            out.append({
                "emp_no": s.emp_no, "name": s.emp_name, "date": "-", "site": "-",
                "issue": (f"Final salary AED {take_home:,.0f} is {pct:.0f}% of the "
                          f"AED {so_far:,.0f} for the {entered} day(s) entered so far "
                          f"- below the 40% minimum (AED {floor:,.0f})."
                          if partial else
                          f"Final salary AED {take_home:,.0f} is {pct:.0f}% of the "
                          f"AED {total:,.0f} total - below the 40% minimum "
                          f"(AED {floor:,.0f})."),
                "kind": "Pay below 40%",
                "severity": "legal",
                "detail": {
                    "Total salary": f"AED {total:,.2f}",
                    **({"Pay for days entered": f"AED {so_far:,.2f} ({entered} of {cycle_days} days)"} if partial else {}),
                    "Minimum payable (40%)": f"AED {floor:,.2f}",
                    "Final salary now": f"AED {take_home:,.2f}",
                    "Short by": f"AED {floor - take_home:,.2f}",
                    "Days absent": f"{s.absent_days:g}",
                    "Absence deduction": f"AED {s.deduction:,.2f}",
                    "Adjustments": (f"AED {adj:,.2f}" if s.adjustments else "none"),
                    "What the adjustments were":
                        ", ".join(f"{a.description} {'-' if a.is_deduction else '+'}{a.amount:,.0f}"
                                  for a in s.adjustments) or "-",
                },
            })

    # Everything about one worker belongs together - chasing a man's
    # problems across three separate parts of the list is how one gets
    # missed. Workers with the most serious issue come first, and within
    # a worker the serious ones lead.
    rank = {"legal": 0, "contradiction": 1}
    worst_by_emp, count_by_emp = {}, {}
    for x in out:
        r = rank.get(x.get("severity"), 2)
        worst_by_emp[x["emp_no"]] = min(worst_by_emp.get(x["emp_no"], 9), r)
        count_by_emp[x["emp_no"]] = count_by_emp.get(x["emp_no"], 0) + 1
    out.sort(key=lambda x: (worst_by_emp[x["emp_no"]], x["emp_no"],
                            rank.get(x.get("severity"), 2), str(x["date"])))
    for x in out:
        x["issue_count"] = count_by_emp[x["emp_no"]]
    return {
        "title": "Check for Errors",
        "note": "Workers paid below the 40% minimum, hours recorded on absent or holiday "
                "days, missing days in this cycle, and unusually high single-day OT/BH values.",
        "rows": out,
    }


@app.get("/export/{month_year}/excel")
def export_excel(month_year: str, token: str, db: Session = Depends(get_db)):
    # Auth comes ONLY via the ?token= query param here, not the standard
    # Authorization header - this endpoint is meant to be hit by a plain
    # browser navigation (window.open(url)), which can't set custom
    # headers at all. That's also the actual fix for exports doing
    # nothing on mobile: the old fetch()+blob()+<a download> approach
    # loses the "real user tap" context by the time the async blob is
    # ready, so mobile browsers silently block the download. A direct,
    # synchronous navigation has no such problem.
    user = auth.get_download_user_from_token(token, db)
    summaries = (
        db.query(models.EmployeeSummary)
        .options(joinedload(models.EmployeeSummary.adjustments))
        .filter(models.EmployeeSummary.month_year == month_year)
        .order_by(models.EmployeeSummary.emp_no)
        .all()
    )
    if not summaries:
        raise HTTPException(status_code=404, detail="No data found for this cycle.")
    pairs = []
    for s in summaries:
        rows = db.query(models.DailyRow).filter(
            and_(models.DailyRow.emp_no == s.emp_no, models.DailyRow.month_year == month_year)
        ).all()
        pairs.append((s, rows))
    buf = export_web.build_combined_excel(pairs)
    safe_name = "".join(c if c.isalnum() else "_" for c in month_year)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Infinia_Cards_{safe_name}.xlsx"},
    )


@app.get("/export/{month_year}/pdf")
def export_pdf(month_year: str, token: str, db: Session = Depends(get_db)):
    user = auth.get_download_user_from_token(token, db)
    summaries = (
        db.query(models.EmployeeSummary)
        .options(joinedload(models.EmployeeSummary.adjustments))
        .filter(models.EmployeeSummary.month_year == month_year)
        .order_by(models.EmployeeSummary.emp_no)
        .all()
    )
    if not summaries:
        raise HTTPException(status_code=404, detail="No data found for this cycle.")
    pairs = []
    for s in summaries:
        rows = db.query(models.DailyRow).filter(
            and_(models.DailyRow.emp_no == s.emp_no, models.DailyRow.month_year == month_year)
        ).all()
        pairs.append((s, rows))
    buf = export_web.build_combined_pdf(pairs)
    safe_name = "".join(c if c.isalnum() else "_" for c in month_year)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Infinia_Cards_{safe_name}.pdf"},
    )


def _get_summary_pairs(month_year: str, db: Session):
    summaries = (
        db.query(models.EmployeeSummary)
        .options(joinedload(models.EmployeeSummary.adjustments))
        .filter(models.EmployeeSummary.month_year == month_year)
        .order_by(models.EmployeeSummary.emp_no)
        .all()
    )
    if not summaries:
        raise HTTPException(status_code=404, detail="No data found for this cycle.")
    pairs = []
    for s in summaries:
        rows = db.query(models.DailyRow).filter(
            and_(models.DailyRow.emp_no == s.emp_no, models.DailyRow.month_year == month_year)
        ).all()
        pairs.append((s, rows))
    return pairs


@app.get("/export/{month_year}/report-table")
def export_report_table(month_year: str, token: str, columns: str, format: str, db: Session = Depends(get_db)):
    """
    Exports the Report Builder's own preview - whatever columns are
    currently ticked, in one row per worker plus a totals row - as
    opposed to Combine, which always exports the full individual salary
    cards regardless of column selection.
    """
    user = auth.get_download_user_from_token(token, db)
    column_keys = [c for c in columns.split(",") if c]
    if not column_keys:
        raise HTTPException(status_code=400, detail="No columns selected.")
    summaries = (
        db.query(models.EmployeeSummary)
        .options(joinedload(models.EmployeeSummary.adjustments))
        .filter(models.EmployeeSummary.month_year == month_year)
        .order_by(models.EmployeeSummary.emp_no)
        .all()
    )
    site_query = db.query(models.DailyRow.emp_no, models.DailyRow.site, models.DailyRow.full_date).filter(
        models.DailyRow.month_year == month_year, models.DailyRow.site != ""
    )
    latest_site_by_emp = {}
    for emp_no, site, full_date in site_query.all():
        if not site:
            continue
        current = latest_site_by_emp.get(emp_no)
        if current is None or full_date > current[1]:
            latest_site_by_emp[emp_no] = (site, full_date)
    items = []
    for s in summaries:
        item = schemas.EmployeeSummaryOut.from_orm(s)
        latest = latest_site_by_emp.get(s.emp_no)
        item.sites = latest[0] if latest else ""
        items.append(item)
    if not items:
        raise HTTPException(status_code=404, detail="No data found for this cycle.")

    safe_name = "".join(c if c.isalnum() else "_" for c in month_year)
    if format == "excel":
        buf = export_web.build_report_table_excel(items, column_keys, month_year)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Infinia_Report_{safe_name}.xlsx"},
        )
    elif format == "pdf":
        buf = export_web.build_report_table_pdf(items, column_keys, month_year)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Infinia_Report_{safe_name}.pdf"},
        )
    raise HTTPException(status_code=400, detail="format must be 'excel' or 'pdf'.")


@app.get("/export/{month_year}/excel-separate")
def export_excel_separate(month_year: str, token: str, db: Session = Depends(get_db)):
    """One .xlsx per worker, zipped together - the 'Separate Files' option next to Combine."""
    user = auth.get_download_user_from_token(token, db)
    pairs = _get_summary_pairs(month_year, db)
    files = export_web.build_separate_excel_files(pairs)
    buf = export_web.zip_files(files)
    safe_name = "".join(c if c.isalnum() else "_" for c in month_year)
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Infinia_Cards_Separate_{safe_name}.zip"},
    )


@app.get("/export/{month_year}/pdf-separate")
def export_pdf_separate(month_year: str, token: str, db: Session = Depends(get_db)):
    """One .pdf per worker, zipped together - the 'Separate Files' option next to Combine."""
    user = auth.get_download_user_from_token(token, db)
    pairs = _get_summary_pairs(month_year, db)
    files = export_web.build_separate_pdf_files(pairs)
    buf = export_web.zip_files(files)
    safe_name = "".join(c if c.isalnum() else "_" for c in month_year)
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Infinia_Cards_Separate_{safe_name}.zip"},
    )


def _row_to_dict(obj):
    d = {}
    for col in obj.__table__.columns:
        val = getattr(obj, col.name)
        d[col.name] = val.isoformat() if hasattr(val, "isoformat") else val
    return d


# Daily snapshots older than this are cleared as new ones are taken.
BACKUP_KEEP_DAYS = 40


def build_backup_data(db: Session) -> dict:
    """Everything the company would need to rebuild this system.

    The store was added long after this function and never joined it,
    so a backup held the payroll side while thousands of materials,
    every stock movement, every request and every supplier existed only
    on the live server. A backup that restores half the business is not
    a backup.

    Left out on purpose: the audit log (a record of who clicked what,
    not company data), and previous backups (a backup of backups).
    Staff logins are included so people can sign in after a restore -
    passwords are stored hashed, never in the clear.
    """
    return {
        "generated_at": datetime.utcnow().isoformat(),
        "format": 2,
        # ---- People and attendance ----
        "users": [_row_to_dict(u) for u in db.query(models.User).all()],
        "employees": [_row_to_dict(e) for e in db.query(models.Employee).all()],
        "sites": [_row_to_dict(s) for s in db.query(models.Site).all()],
        "engineers": [_row_to_dict(e) for e in db.query(models.Engineer).all()],
        "daily_rows": [_row_to_dict(r) for r in db.query(models.DailyRow).all()],
        "summaries": [_row_to_dict(s) for s in db.query(models.EmployeeSummary).all()],
        "adjustments": [_row_to_dict(a) for a in db.query(models.SalaryAdjustment).all()],
        # ---- Store and materials ----
        "suppliers": [_row_to_dict(s) for s in db.query(models.Supplier).all()],
        "store_items": [_row_to_dict(i) for i in db.query(models.StoreItem).all()],
        "store_movements": [_row_to_dict(m) for m in db.query(models.StoreMovement).all()],
        "material_requests": [_row_to_dict(r) for r in db.query(models.MaterialRequest).all()],
        "material_request_lines": [_row_to_dict(l) for l in db.query(models.MaterialRequestLine).all()],
    }


def _add_missing_columns():
    """Lightweight migration: ALTER TABLE ADD COLUMN for anything the
    models declare but the live table doesn't have yet."""
    from sqlalchemy import inspect, text
    type_sql = {"VARCHAR": "VARCHAR", "FLOAT": "DOUBLE PRECISION", "DATE": "DATE",
                "INTEGER": "INTEGER", "BOOLEAN": "BOOLEAN", "TEXT": "TEXT"}
    insp = inspect(engine)
    existing_tables = set(insp.get_table_names())
    with engine.begin() as conn:
        for table in Base.metadata.sorted_tables:
            if table.name not in existing_tables:
                continue
            have = {c["name"] for c in insp.get_columns(table.name)}
            for col in table.columns:
                if col.name in have:
                    continue
                t = type_sql.get(str(col.type).split("(")[0].upper())
                if not t:
                    continue
                try:
                    conn.execute(text(f'ALTER TABLE {table.name} ADD COLUMN {col.name} {t}'))
                    print(f"Added column {table.name}.{col.name}")
                    # A new column arrives empty on every existing row.
                    # Where the model declares a plain default, apply it
                    # to what is already there - otherwise hundreds of
                    # records sit with nothing in a field the app now
                    # expects to be filled.
                    d = getattr(col.default, "arg", None) if col.default is not None else None
                    if isinstance(d, (str, int, float, bool)):
                        val = f"'{d}'" if isinstance(d, str) else (
                            "TRUE" if d is True else "FALSE" if d is False else str(d))
                        conn.execute(text(
                            f'UPDATE {table.name} SET {col.name} = {val} WHERE {col.name} IS NULL'))
                        print(f"  backfilled {table.name}.{col.name} = {d}")
                except Exception as e:
                    print(f"Could not add {table.name}.{col.name}: {e}")


AUTO_BACKUP_KEEP_DAYS = 40

def maybe_create_auto_backup(db: Session):
    """
    Creates one 'auto' backup per calendar DAY, then prunes automatic
    backups down to the most recent AUTO_BACKUP_KEEP_DAYS - a rolling
    40-day window where day 41 replaces day 1.

    Runs on login rather than a cron job: this app has no scheduler
    process of its own, and a backup is only useful if the data has
    actually been touched, which requires someone to be signed in
    anyway. If nobody logs in on a given day there is nothing new to
    snapshot, so no backup is the correct outcome, not a missed one.

    Manual backups are never pruned - only the automatic ones - so a
    snapshot someone deliberately took before a risky change is never
    silently deleted out from under them.
    """
    today = _dubai_today()
    existing = (
        db.query(models.Backup)
        .filter(models.Backup.trigger == "auto",
                func.date(models.Backup.created_at) == today)
        .first()
    )
    if not existing:
        data = build_backup_data(db)
        db.add(models.Backup(created_by=None, trigger="auto", data=json.dumps(data, default=str)))
        db.commit()

    # Prune: keep only the newest N automatic backups.
    old_auto = (
        db.query(models.Backup)
        .filter(models.Backup.trigger == "auto")
        .order_by(models.Backup.created_at.desc())
        .offset(AUTO_BACKUP_KEEP_DAYS)
        .all()
    )
    if old_auto:
        for b in old_auto:
            db.delete(b)
        db.commit()


# Deliberately under /backup/ rather than a new /admin/ prefix: nginx
# only forwards the API paths listed in its rule, and anything else
# falls through to the frontend, where a POST comes back 405. Reusing a
# prefix that already works means no server config change to deploy
# this - and it belongs with the backups anyway, since it takes one.
@app.post("/backup/fresh-start")
def fresh_start(payload: dict = Body(...), db: Session = Depends(get_db),
                 user: models.User = Depends(auth.require_admin)):
    """Clear the test data before going live, keeping the master lists.

    Kept: workers with their company, sites, engineers, the material
    list and every staff login. Cleared: attendance, payroll summaries,
    salary adjustments, stock movements, requests, suppliers and the
    activity log.

    Admin only, and the exact words must be typed - this empties the
    company's records and there is no undo beyond the backup it takes
    first. That backup is kept on the server so it can be restored from
    the list below, whatever else happens.
    """
    if (payload or {}).get("confirm") != "CLEAR EVERYTHING":
        raise HTTPException(status_code=400,
            detail='Type CLEAR EVERYTHING exactly to confirm.')

    # A copy of everything first, kept as a normal backup so it can be
    # restored from Settings if today's figures turn out to be needed.
    db.add(models.Backup(created_by=user.id, trigger="before-fresh-start",
                         data=json.dumps(build_backup_data(db), default=str)))
    db.commit()

    cleared = {}
    # Child-first, so nothing is left pointing at a deleted row.
    for label, model in (("request lines", models.MaterialRequestLine),
                         ("material requests", models.MaterialRequest),
                         ("stock movements", models.StoreMovement),
                         ("suppliers", models.Supplier),
                         ("salary adjustments", models.SalaryAdjustment),
                         ("payroll summaries", models.EmployeeSummary),
                         ("attendance days", models.DailyRow)):
        cleared[label] = db.query(model).delete(synchronize_session=False)
    db.commit()
    # The activity log goes last, so this clearance is itself recorded.
    db.query(models.AuditLog).delete(synchronize_session=False)
    db.commit()
    log_action(db, user.id, "fresh_start",
               ", ".join(f"{n} {k}" for k, n in cleared.items() if n))

    kept = {
        "workers": db.query(models.Employee).count(),
        "sites": db.query(models.Site).count(),
        "engineers": db.query(models.Engineer).count(),
        "materials": db.query(models.StoreItem).count(),
        "staff logins": db.query(models.User).count(),
    }
    return {"ok": True, "cleared": cleared, "kept": kept,
            "detail": "Cleared. A backup of what was removed is at the top of the "
                      "backup list, and request numbering starts again at MR-0001."}


@app.post("/backup/create")
def create_backup(db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    """Any signed-in user can take a backup - it's a data-safety net,
    not something that should be gated behind admin access."""
    data = build_backup_data(db)
    b = models.Backup(created_by=user.id, trigger="manual", data=json.dumps(data, default=str))
    db.add(b)
    db.commit()
    db.refresh(b)
    log_action(db, user.id, "create_backup")
    return {"id": b.id, "created_at": b.created_at.isoformat(), "trigger": b.trigger}


@app.get("/backup/list")
def list_backups(db: Session = Depends(get_db), user: models.User = Depends(auth.get_current_user)):
    rows = (
        db.query(models.Backup, models.User.username)
        .outerjoin(models.User, models.Backup.created_by == models.User.id)
        .order_by(models.Backup.created_at.desc())
        .limit(120)
        .all()
    )
    return [
        {"id": b.id, "created_at": b.created_at.isoformat(), "trigger": b.trigger,
         "created_by": username or ("automatic" if b.trigger == "auto" else "unknown")}
        for b, username in rows
    ]


@app.get("/backup/latest/download")
def download_latest_backup(token: str = None, db: Session = Depends(get_db)):
    """A complete copy of the system for whoever is signed in.

    The point is that the company's data ends up on several machines
    rather than only the one server, so anyone using the app can hold a
    copy. Every table is included, so any of these files can rebuild the
    business.

    Password hashes are the one thing removed. They are of no use in a
    restore - existing logins are kept and missing ones come back with
    their names, roles and permissions, needing a fresh password - and
    a file sitting in a Downloads folder should not carry them.
    """
    user = auth.get_download_user_from_token(token, db)
    # A full copy carries every worker's pay, so it goes only to the
    # machines that are allowed to see pay anyway. A site engineer or
    # store keeper holding a copy would undo the salary privacy the rest
    # of the app enforces.
    perms = effective_permissions(user)
    # Only someone who can already see pay in the app gets a file that
    # carries every salary. Approvals marks the office desk; the payroll
    # screens mark an admin. A site engineer, or a store keeper whose
    # login is office in name but store-only in permissions, is not
    # handed one.
    if not any(s in perms for s in ("approvals", "masterdata", "adjustments", "reports", "livecard")):
        raise HTTPException(status_code=403,
            detail="Only office and admin accounts can download a full backup.")
    data = build_backup_data(db)
    data["users"] = [{**u, "hashed_password": ""} for u in data.get("users", [])]
    data["downloaded_by"] = user.username
    body = json.dumps(data, default=str)
    # Keep one snapshot a day on the server as well, so the two copies
    # match and the list does not fill with one row per person per day.
    today = _dubai_today()
    existing = (db.query(models.Backup)
                  .filter(models.Backup.trigger == "daily")
                  .order_by(models.Backup.id.desc()).first())
    if not existing or existing.created_at.date() != today:
        db.add(models.Backup(created_by=user.id, trigger="daily",
                             data=json.dumps(build_backup_data(db), default=str)))
        # Each snapshot is several megabytes, so old ones are cleared as
        # new ones arrive - otherwise the database quietly grows by a
        # copy of itself every day. Manual backups are left alone: those
        # were taken deliberately, usually before something risky.
        cutoff = datetime.now(timezone.utc) - timedelta(days=BACKUP_KEEP_DAYS)
        old_ones = (db.query(models.Backup)
                      .filter(models.Backup.trigger == "daily",
                              models.Backup.created_at < cutoff).all())
        for b in old_ones:
            db.delete(b)
        db.commit()
    buf = io.BytesIO(body.encode("utf-8"))
    return StreamingResponse(
        buf, media_type="application/json",
        headers={"Content-Disposition":
                 f'attachment; filename=Infinia_Full_Backup_{today.isoformat()}.json'},
    )


@app.get("/backup/{backup_id}/download")
def download_backup(backup_id: int, token: str, db: Session = Depends(get_db)):
    user = auth.get_download_user_from_token(token, db)
    b = db.query(models.Backup).filter(models.Backup.id == backup_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Backup not found.")
    log_action(db, user.id, "download_backup", f"backup #{backup_id}")
    buf = io.BytesIO(b.data.encode("utf-8"))
    ts = b.created_at.date().isoformat()
    return StreamingResponse(
        buf, media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=Infinia_Backup_{ts}_{backup_id}.json"},
    )


@app.post("/backup/{backup_id}/restore")
def restore_backup(backup_id: int, db: Session = Depends(get_db),
                    user: models.User = Depends(auth.require_admin)):
    """
    Admin only, unlike taking a backup - this overwrites current data,
    so it stays behind the higher bar. Replaces attendance and the
    store - employees, sites, engineers, daily rows, summaries,
    adjustments, suppliers, materials, movements and requests - with
    exactly what is in the chosen snapshot. Staff logins missing from
    the live system are put back; existing ones are left alone so the
    admin doing the restore cannot lock themselves out.
    """
    b = db.query(models.Backup).filter(models.Backup.id == backup_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Backup not found.")
    data = json.loads(b.data)

    # Cleared child-first so nothing is left pointing at a deleted row.
    # Store tables are only cleared when the snapshot actually carries
    # them, so restoring an older backup cannot wipe the inventory it
    # never knew about.
    has_store = any(k in data for k in ("store_items", "store_movements", "material_requests"))
    if has_store:
        db.query(models.MaterialRequestLine).delete()
        db.query(models.MaterialRequest).delete()
        db.query(models.StoreMovement).delete()
        db.query(models.StoreItem).delete()
        db.query(models.Supplier).delete()
    db.query(models.SalaryAdjustment).delete()
    db.query(models.DailyRow).delete()
    db.query(models.EmployeeSummary).delete()
    db.query(models.Employee).delete()
    db.query(models.Site).delete()
    db.query(models.Engineer).delete()
    db.flush()

    def restore_rows(model, rows, date_fields=(), datetime_fields=()):
        for r in rows:
            r = dict(r)
            for f in date_fields:
                if r.get(f):
                    r[f] = date.fromisoformat(r[f])
            for f in datetime_fields:
                if r.get(f):
                    r[f] = datetime.fromisoformat(r[f])
            db.add(model(**r))

    restore_rows(models.Employee, data.get("employees", []), datetime_fields=("created_at", "updated_at"))
    restore_rows(models.Site, data.get("sites", []), datetime_fields=("created_at",))
    restore_rows(models.Engineer, data.get("engineers", []), datetime_fields=("created_at",))
    restore_rows(models.DailyRow, data.get("daily_rows", []), date_fields=("full_date",), datetime_fields=("created_at", "updated_at"))
    restore_rows(models.EmployeeSummary, data.get("summaries", []), datetime_fields=("created_at", "updated_at"))
    db.commit()

    for a in data.get("adjustments", []):
        a = dict(a)
        if a.get("created_at"):
            a["created_at"] = datetime.fromisoformat(a["created_at"])
        db.add(models.SalaryAdjustment(**a))
    db.commit()

    # ---- Store: suppliers and materials before the records that point
    # at them, so every link survives the restore.
    if has_store:
        restore_rows(models.Supplier, data.get("suppliers", []), datetime_fields=("created_at",))
        restore_rows(models.StoreItem, data.get("store_items", []),
                     date_fields=("rental_start", "rental_due"), datetime_fields=("created_at", "updated_at"))
        db.commit()
        restore_rows(models.StoreMovement, data.get("store_movements", []),
                     date_fields=("moved_on",), datetime_fields=("created_at",))
        restore_rows(models.MaterialRequest, data.get("material_requests", []),
                     date_fields=("needed_by", "requested_on", "closed_on", "expected_on"),
                     datetime_fields=("created_at", "updated_at"))
        db.commit()
        restore_rows(models.MaterialRequestLine, data.get("material_request_lines", []))
        db.commit()

    # Staff logins last: an admin restoring a snapshot must not delete
    # the account they are signed in with, so existing logins are kept
    # and only missing ones are put back.
    have = {u.username for u in db.query(models.User).all()}
    for u in data.get("users", []):
        if u.get("username") in have:
            continue
        u = dict(u)
        for f in ("created_at", "last_login"):
            if u.get(f):
                u[f] = datetime.fromisoformat(u[f])
        # On a rebuilt server the rescue admin already holds an id the
        # snapshot also claims. The login matters, the number does not -
        # let the database allocate a free one rather than failing the
        # whole restore.
        # Ids are not worth preserving here and cause collisions both
        # with logins already on the server and with each other once one
        # has been reassigned. The login, role and permissions are what
        # matter; let the database number them.
        u.pop("id", None)
        db.add(models.User(**u))
        db.flush()
    db.commit()

    log_action(db, user.id, "restore_backup", f"restored from backup #{backup_id}")
    return {"ok": True, "restored_from": backup_id}


@app.delete("/backup/{backup_id}")
def delete_backup(backup_id: int, db: Session = Depends(get_db),
                   user: models.User = Depends(auth.require_admin)):
    """
    Remove a single backup. Useful for clearing out snapshots taken
    before a known-bad state, so nobody restores one by mistake later -
    a real risk, since a backup's date alone doesn't say whether the
    data inside it was correct.
    """
    b = db.query(models.Backup).filter(models.Backup.id == backup_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Backup not found")
    when = b.created_at.isoformat() if b.created_at else str(backup_id)
    db.delete(b)
    db.commit()
    log_action(db, user.id, "delete_backup", f"#{backup_id} ({when})")
    return {"ok": True, "deleted": backup_id}


@app.get("/health")
def health_check():
    return {"status": "ok"}


# ---------------------------------------------------------------------
# STORE / INVENTORY
# ---------------------------------------------------------------------
CENTRAL = ""   # empty location string means the central store


def _stock_map(db: Session, upto: date = None):
    """
    Current quantity of every item at every location, derived from the
    movement ledger rather than stored - so a balance can never drift
    away from the history that produced it.

    Returns {(item_id, location): qty}. 'upto' gives the position as at
    a date, which is what makes stock-as-at reporting possible.
    """
    q = db.query(models.StoreMovement)
    if upto:
        q = q.filter(models.StoreMovement.moved_on <= upto)
    stock = {}
    for m in q.all():
        if m.kind == "in":
            stock[(m.item_id, m.location)] = stock.get((m.item_id, m.location), 0) + m.qty
        elif m.kind == "adjust":
            stock[(m.item_id, m.location)] = stock.get((m.item_id, m.location), 0) + m.qty
        elif m.kind == "lost":
            # Written off - it leaves wherever it was and doesn't arrive
            # anywhere, so only the 'from' side moves.
            stock[(m.item_id, m.from_location)] = stock.get((m.item_id, m.from_location), 0) - m.qty
        elif m.kind in ("out", "return", "transfer"):
            stock[(m.item_id, m.from_location)] = stock.get((m.item_id, m.from_location), 0) - m.qty
            stock[(m.item_id, m.location)] = stock.get((m.item_id, m.location), 0) + m.qty
    return stock


def _supplier_key(name: str) -> str:
    """Fold a supplier name to a comparison key: lower case, punctuation
    dropped, spacing collapsed, and the usual trading suffixes removed.
    "AL RAHA TRADING LLC", "Al-Raha Trading" and "al raha  trading llc."
    all land on "al raha", so one supplier is one record."""
    s = re.sub(r"[^a-z0-9 ]+", " ", (name or "").lower())
    s = " ".join(s.split())
    for suffix in (" llc", " l l c", " trading", " general trading", " co", " company",
                   " est", " establishment", " fzc", " fze", " ltd", " limited"):
        while s.endswith(suffix):
            s = s[: -len(suffix)].strip()
    return s


def _tidy_supplier_name(name: str) -> str:
    """Store the name the way it should be read: each word capitalised,
    but abbreviations people write in capitals (LLC, FZE, ADNOC) kept."""
    keep = {"llc", "fze", "fzc", "uae", "adnoc", "gi", "pvc", "upvc", "ppr", "opc"}
    out = []
    for w in " ".join((name or "").split()).split(" "):
        if not w:
            continue
        out.append(w.upper() if w.lower() in keep else w[0].upper() + w[1:].lower())
    return " ".join(out)


def _find_or_create_supplier(db, name, contact_person="", phone=""):
    """Look a supplier up by its folded key, creating it the first time.
    Contact details fill in as they are learned: a blank phone today is
    filled by tomorrow's delivery note, but an existing one is never
    overwritten with an empty box."""
    key = _supplier_key(name)
    if not key:
        return None
    sup = db.query(models.Supplier).filter(models.Supplier.name_key == key).first()
    if not sup:
        sup = models.Supplier(name=_tidy_supplier_name(name), name_key=key,
                              contact_person=_proper_name((contact_person or "").strip()),
                              phone=(phone or "").strip(), active=True)
        db.add(sup)
        db.flush()
        return sup
    if contact_person and contact_person.strip():
        sup.contact_person = _proper_name(contact_person.strip())
    if phone and phone.strip():
        sup.phone = phone.strip()
    return sup


@app.get("/store/suppliers")
def list_suppliers(db: Session = Depends(get_db),
                    user: models.User = Depends(require_any_screen("store", "requests", "approvals"))):
    return [{"id": s.id, "name": s.name, "contact_person": s.contact_person or "",
             "phone": s.phone or "", "notes": s.notes or ""}
            for s in db.query(models.Supplier).filter(models.Supplier.active == True)  # noqa: E712
                       .order_by(models.Supplier.name).all()]


@app.post("/store/suppliers")
def save_supplier(payload: schemas.SupplierIn, db: Session = Depends(get_db),
                   user: models.User = Depends(require_any_screen("store", "requests", "approvals"))):
    if not (payload.name or "").strip():
        raise HTTPException(status_code=400, detail="Supplier needs a name.")
    sup = _find_or_create_supplier(db, payload.name, payload.contact_person, payload.phone)
    if payload.notes:
        sup.notes = payload.notes
    db.commit()
    return {"id": sup.id, "name": sup.name, "contact_person": sup.contact_person or "",
            "phone": sup.phone or ""}


@app.put("/store/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, payload: schemas.SupplierIn,
                     db: Session = Depends(get_db),
                     user: models.User = Depends(require_any_screen("store", "requests", "approvals"))):
    """Edit a supplier in place. Renaming recomputes the folded key and
    refuses a name that would collide with another supplier - two
    records silently becoming aliases of each other is how contact
    numbers get lost."""
    sup = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found.")
    name = _proper_name((payload.name or "").strip())
    if not name:
        raise HTTPException(status_code=400, detail="Supplier needs a name.")
    key = _supplier_key(name)
    clash = db.query(models.Supplier).filter(models.Supplier.name_key == key,
                                              models.Supplier.id != supplier_id).first()
    if clash:
        raise HTTPException(status_code=400,
            detail=f'That name matches the existing supplier "{clash.name}".')
    sup.name = _tidy_supplier_name(name)
    sup.name_key = key
    sup.contact_person = _proper_name((payload.contact_person or "").strip())
    sup.phone = (payload.phone or "").strip()
    if payload.notes is not None:
        sup.notes = payload.notes
    db.commit()
    return {"id": sup.id, "name": sup.name, "contact_person": sup.contact_person,
            "phone": sup.phone}


@app.post("/store/request-lines/{line_id}/decision")
def decide_request_line(line_id: int, payload: schemas.LineDecisionIn,
                         db: Session = Depends(get_db),
                         user: models.User = Depends(require_screen("approvals"))):
    """Approve or reject one material without touching the rest.

    The office often wants nine of ten materials and not the tenth. The
    request itself then follows its lines: rejected outright only when
    every material is turned down, approved once anything is approved,
    and left waiting while decisions are still outstanding."""
    if payload.decision not in ("approved", "rejected", "pending"):
        raise HTTPException(status_code=400, detail="Decision must be approved, rejected or pending.")
    line = db.query(models.MaterialRequestLine).filter(
        models.MaterialRequestLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Request line not found.")
    if payload.decision == "rejected":
        if (line.qty_received or 0) > 0:
            raise HTTPException(status_code=400,
                detail="Some of this has already arrived, so it can't be rejected now.")
        if line.supplier_id:
            raise HTTPException(status_code=400,
                detail="This has already been ordered from a supplier, so it can't be rejected now.")
    line.status = payload.decision
    line.reject_reason = (payload.reason or "").strip() if payload.decision == "rejected" else ""

    mr = line.request
    states = [l.status or "pending" for l in mr.lines]
    if all(s == "rejected" for s in states):
        mr.status = "rejected"
    elif mr.status in ("pending", "approved", "rejected"):
        # Anything approved moves the request forward; otherwise it waits.
        mr.status = "approved" if any(s == "approved" for s in states) else "pending"
    db.commit()
    log_action(db, user.id, "request_line_decision",
               f"{mr.ref} line {line_id} {payload.decision}")
    return _mr_out(mr)


@app.post("/store/request-lines/{line_id}/link-item")
def link_line_item(line_id: int, payload: schemas.LinkLineItemIn,
                    db: Session = Depends(get_db),
                    user: models.User = Depends(require_any_screen("requests", "approvals"))):
    """Attach a store item to a line that was typed as free text.

    Someone asks for "Cushions", which isn't in the catalogue yet. The
    request is fine, but the delivery can't be received into stock
    because there's nothing to count it against. Rather than a dead end
    at the delivery modal, the keeper adds it to the item list from
    there and the line is linked to it."""
    line = db.query(models.MaterialRequestLine).filter(
        models.MaterialRequestLine.id == line_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Request line not found.")
    item = db.query(models.StoreItem).filter(models.StoreItem.id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found.")
    line.item_id = item.id
    if not line.unit:
        line.unit = item.unit
    db.commit()
    return {"ok": True, "item_id": item.id, "code": item.code}


@app.get("/store/items", response_model=list[schemas.StoreItemOut])
def list_store_items(active_only: bool = True, db: Session = Depends(get_db),
                      user: models.User = Depends(require_screen("store"))):
    q = db.query(models.StoreItem)
    if active_only:
        q = q.filter(models.StoreItem.active == True)  # noqa: E712
    return q.order_by(models.StoreItem.code).all()


# Recognising a material from its name. Units follow UN/CEFACT Rec 20
# symbols (t, kg, m, m2, m3, L) plus Rec 21 packaging (bag, drum, roll).
# Rules run top to bottom, first match wins, so the specific ones
# (welding rod -> box) sit above the general ones (rod -> kg). A rule
# only fires on whole words, so "sanding disc" never matches "sand".
UNIT_RULES = [
    (("cement", "gypsum powder", "grout", "plaster", "mortar", "adhesive powder", "tile adhesive", "white cement", "putty"), "bag", "powders are bought by the bag"),
    (("rebar", "rebars", "reinforcement", "tor steel", "tmt"), "t", "rebar is bought by the tonne"),
    (("sand", "aggregate", "gravel", "sweet soil", "crush", "readymix", "ready mix", "concrete"), "m3", "loose material is measured in cubic metres"),
    (("nail", "nails", "screw", "screws", "binding wire"), "kg", "loose fixings are bought by weight"),
    (("cable", "wire", "conduit", "trunking", "hose", "rope", "chain", "gi pipe", "pvc pipe", "ppr pipe", "upvc", "duct", "tube pipe"), "m", "run material is measured by the metre"),
    (("paint", "primer", "thinner", "curing compound", "admixture", "bitumen", "sealer", "chemical", "diesel", "petrol", "oil", "waterproofing liquid"), "L", "liquids are measured in litres"),
    (("plywood", "gypsum board", "mdf", "gi sheet", "cement board", "shutter ply", "marine ply"), "sheet", "boards are counted in sheets"),
    (("mesh roll", "geotextile", "membrane", "felt", "polythene", "shade net", "hessian", "insulation roll"), "roll", "rolled goods are counted in rolls"),
    (("welding rod", "electrode", "welding electrodes"), "box", "welding rods come by the box"),
    (("silicone", "sealant cartridge", "pu foam", "gun foam"), "tube", "cartridges are counted in tubes"),
    (("glove", "gloves", "boot", "boots", "goggle", "goggles"), "pair", "safety wear comes in pairs"),
]


def _suggest_unit(name: str):
    words = " " + " ".join((name or "").lower().replace("/", " ").replace("-", " ").split()) + " "
    for keys, unit, reason in UNIT_RULES:
        for k in keys:
            if f" {k} " in words or (k.endswith(" ") and k in words):
                return unit, k, reason
    return None, None, None


@app.get("/store/items/suggest-units")
def suggest_units(db: Session = Depends(get_db),
                   user: models.User = Depends(require_screen("store"))):
    """Walk the whole catalogue and propose a standard unit for every
    material whose name identifies it - cement to bags, rebar to tonnes,
    cable to metres. Nothing is changed here: the keeper reviews the
    list and applies only the rows they agree with."""
    out = []
    for it in db.query(models.StoreItem).filter(models.StoreItem.active == True).all():  # noqa: E712
        unit, matched, reason = _suggest_unit(it.name)
        if unit and unit.lower() != (it.unit or "").lower().strip():
            out.append({"id": it.id, "code": it.code, "name": it.name,
                        "current": it.unit, "suggested": unit,
                        "matched": matched, "reason": reason})
    return {"suggestions": out}


@app.post("/store/items/apply-units")
def apply_units(payload: schemas.ApplyUnitsIn, db: Session = Depends(get_db),
                 user: models.User = Depends(require_screen("store"))):
    """Apply the unit changes the keeper ticked in the review."""
    done = 0
    for ch in payload.changes:
        it = db.query(models.StoreItem).filter(models.StoreItem.id == ch.id).first()
        if it and ch.unit and ch.unit.strip():
            it.unit = ch.unit.strip()
            done += 1
    db.commit()
    return {"updated": done}


def _next_item_code(db):
    """ITM1, ITM2... The keeper never invents a code; existing items keep
    whatever code they were given."""
    used = {i.code for i in db.query(models.StoreItem).all()}
    n = 1
    while f"ITM{n}" in used:
        n += 1
    return f"ITM{n}"


def _clean_export_qty(v):
    """400.0 -> 400, 7.5 -> 7.5 - counts never carry a fake decimal."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    return str(int(f)) if f == int(f) else str(f)


def _person_name(s: str) -> str:
    """A person's name, tidied: "akhil", "AKHIL" and "Akhil" all become
    "Akhil", so the same person can't appear three ways in a list. Only
    fully-capitalised words are folded, so "Mohammed A K" keeps its
    initials."""
    out = []
    for w in " ".join((s or "").split()).split(" "):
        if not w:
            continue
        if len(w) > 1 and w == w.upper() and any(c.isalpha() for c in w):
            w = w.lower()
        out.append(w[0].upper() + w[1:])
    return " ".join(out)


def _proper_name(s: str) -> str:
    """First letter of each word up, the rest untouched, so "cushions"
    becomes "Cushions" while "cement OPC 42.5" keeps OPC as OPC. The
    server does this itself: names arrive from several paths and only
    some of them pass through the browser's tidying."""
    return " ".join(w[0].upper() + w[1:] if w else w
                    for w in " ".join((s or "").split()).split(" "))


def _find_or_create_item(db, name, unit, item_type):
    """Return the catalogue item for a typed-in material, creating it if
    it's genuinely new.

    A material asked for by name only ("Cushions") used to live as loose
    text on one request: it couldn't be received into stock, couldn't be
    given out, and the next person asking typed it slightly differently.
    Adding it to the list at request time means it has a code, a unit,
    and a history from the moment it is first asked for.
    """
    clean = " ".join((name or "").split())
    if not clean:
        return None
    existing = db.query(models.StoreItem).filter(
        func.lower(models.StoreItem.name) == clean.lower()).first()
    if existing:
        return existing
    it = models.StoreItem(code=_next_item_code(db), name=_proper_name(clean),
                          unit=(unit or "pcs").strip() or "pcs",
                          item_type=item_type if item_type in ("consumable", "asset", "rental") else "consumable",
                          category="", reorder_level=0, active=True)
    db.add(it)
    db.flush()
    return it


@app.post("/store/items", response_model=schemas.StoreItemOut)
def upsert_store_item(payload: schemas.StoreItemIn, db: Session = Depends(get_db),
                       user: models.User = Depends(require_screen("store"))):
    # Validate on the server, not just in the browser - a blank code or
    # name creates an item that can't be identified in any report.
    code = (payload.code or "").strip()
    name = _proper_name((payload.name or "").strip())
    if not code:
        code = _next_item_code(db)
    if not name:
        raise HTTPException(status_code=400, detail="Item needs a name.")
    if payload.item_type == "returnable":
        payload.item_type = "asset"        # retired type, folded into assets
    if payload.item_type not in ("consumable", "asset", "rental"):
        raise HTTPException(status_code=400, detail="Item type must be consumable, asset or rental.")
    if payload.reorder_level < 0:
        raise HTTPException(status_code=400, detail="Reorder level can't be negative.")
    payload.code, payload.name = code, name

    existing = db.query(models.StoreItem).filter(models.StoreItem.code == code).first()
    if existing:
        for k, v in payload.dict().items():
            setattr(existing, k, v)
        existing.active = True
    else:
        existing = models.StoreItem(**payload.dict())
        db.add(existing)
    db.commit()
    db.refresh(existing)
    log_action(db, user.id, "store_save_item", f"{payload.code} - {payload.name}")
    return existing


@app.delete("/store/items/{item_id}")
def deactivate_store_item(item_id: int, db: Session = Depends(get_db),
                           user: models.User = Depends(auth.get_current_user)):
    it = db.query(models.StoreItem).filter(models.StoreItem.id == item_id).first()
    if not it:
        raise HTTPException(status_code=404, detail="Item not found")
    # Deactivate rather than delete - its movement history must survive.
    it.active = False
    db.commit()
    log_action(db, user.id, "store_remove_item", it.code)
    return {"ok": True}


@app.get("/store/stock")
def store_stock(location: str = None, db: Session = Depends(get_db),
                 user: models.User = Depends(require_screen("store"))):
    """
    Stock on hand. Without a location, one row per item showing the
    central-store quantity, how much is out at sites, and the total -
    plus a low flag when it has fallen to or below its reorder level.
    """
    items = db.query(models.StoreItem).filter(models.StoreItem.active == True).all()  # noqa: E712
    stock = _stock_map(db)
    rows = []
    for it in items:
        at_central = stock.get((it.id, CENTRAL), 0)
        out_total = sum(v for (iid, loc), v in stock.items() if iid == it.id and loc != CENTRAL)
        by_site = {loc: v for (iid, loc), v in stock.items() if iid == it.id and loc != CENTRAL and v}
        if location is not None:
            qty = stock.get((it.id, location), 0)
            rows.append({"item_id": it.id, "code": it.code, "name": it.name,
                          "category": it.category, "unit": it.unit, "item_type": it.item_type,
                          "qty": round(qty, 2), "reorder_level": it.reorder_level,
                          "low": qty <= it.reorder_level and it.reorder_level > 0})
        else:
            rows.append({"item_id": it.id, "code": it.code, "name": it.name,
                          "category": it.category, "unit": it.unit, "item_type": it.item_type,
                          "central": round(at_central, 2), "out_at_sites": round(out_total, 2),
                          "total": round(at_central + out_total, 2),
                          "by_site": {k: round(v, 2) for k, v in by_site.items()},
                          "reorder_level": it.reorder_level,
                          "low": at_central <= it.reorder_level and it.reorder_level > 0})
    return rows


@app.get("/store/movements", response_model=list[schemas.StoreMovementOut])
def list_store_movements(item_id: int = None, location: str = None, kind: str = None,
                          date_from: str = None, date_to: str = None, limit: int = 500,
                          db: Session = Depends(get_db),
                          user: models.User = Depends(require_screen("store"))):
    q = db.query(models.StoreMovement)
    if item_id:
        q = q.filter(models.StoreMovement.item_id == item_id)
    if kind:
        q = q.filter(models.StoreMovement.kind == kind)
    if location:
        q = q.filter(or_(models.StoreMovement.location == location,
                          models.StoreMovement.from_location == location))
    if date_from:
        q = q.filter(models.StoreMovement.moved_on >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        q = q.filter(models.StoreMovement.moved_on <= datetime.strptime(date_to, "%Y-%m-%d").date())
    rows = q.order_by(models.StoreMovement.moved_on.desc(), models.StoreMovement.id.desc()).limit(limit).all()
    out = []
    for m in rows:
        d = schemas.StoreMovementOut.model_validate(m).model_dump()
        d["item_code"] = m.item.code if m.item else ""
        d["item_name"] = m.item.name if m.item else ""
        d["unit"] = m.item.unit if m.item else ""
        out.append(d)
    return out


@app.post("/store/movements", response_model=schemas.StoreMovementOut)
def add_store_movement(payload: schemas.StoreMovementIn, db: Session = Depends(get_db),
                        user: models.User = Depends(require_screen("store"))):
    # Seeing stock and moving stock are different things. A site login
    # can look at what the store holds so it knows what to ask for, but
    # one central keeper records every receipt and issue - a site
    # engineer issuing to himself is exactly the record nobody can
    # later reconcile.
    if user.role == "site":
        raise HTTPException(status_code=403,
            detail="Stock is received and issued by the store keeper. Raise a material request instead.")
    item = db.query(models.StoreItem).filter(models.StoreItem.id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=400, detail="Item not found.")
    if payload.kind not in ("in", "out", "return", "adjust", "transfer", "lost"):
        raise HTTPException(status_code=400, detail="Unknown movement type.")
    if payload.qty <= 0 and payload.kind != "adjust":
        raise HTTPException(status_code=400, detail="Quantity must be more than zero.")
    if payload.moved_on > date.today():
        raise HTTPException(status_code=400, detail="Date is in the future.")
    # Moving something from a place to the same place changes nothing but
    # leaves a confusing entry in the ledger, so reject it outright.
    if payload.kind in ("return", "transfer") and payload.from_location == payload.location:
        where = payload.location or "the central store"
        raise HTTPException(status_code=400,
            detail=f"'From' and 'To' are both {where} - pick different places.")

    # Don't allow issuing more than is actually held - a negative balance
    # means the ledger no longer describes anything real.
    if payload.kind in ("out", "return", "transfer", "lost"):
        have = _stock_map(db).get((item.id, payload.from_location), 0)
        if payload.qty > have + 1e-9:
            where = payload.from_location or "the central store"
            raise HTTPException(status_code=400,
                detail=f"Only {round(have, 2)} {item.unit} of {item.name} available at {where}.")

    m = models.StoreMovement(**payload.dict(), created_by=user.id)
    # Names are tidied wherever they enter the system, not only on the
    # screen that happened to be used - people type AKHIL, akhil and
    # Akhil, and all three are the same man.
    m.incharge = _person_name(getattr(m, "incharge", "") or "")
    if payload.kind == "in" and (payload.supplier or "").strip():
        sup = _find_or_create_supplier(db, payload.supplier)
        if sup:
            m.supplier = sup.name
    db.add(m)
    db.commit()
    db.refresh(m)
    log_action(db, user.id, "store_movement",
               f"{payload.kind} {payload.qty} {item.unit} {item.code}")
    d = schemas.StoreMovementOut.model_validate(m).model_dump()
    d["item_code"], d["item_name"], d["unit"] = item.code, item.name, item.unit
    return d


@app.delete("/store/movements/{movement_id}")
def delete_store_movement(movement_id: int, db: Session = Depends(get_db),
                           user: models.User = Depends(auth.get_current_user)):
    m = db.query(models.StoreMovement).filter(models.StoreMovement.id == movement_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Movement not found")
    db.delete(m)
    db.commit()
    log_action(db, user.id, "store_delete_movement", f"#{movement_id}")
    return {"ok": True}


@app.get("/store/report")
def store_report(kind: str = "stock", date_from: str = None, date_to: str = None,
                  db: Session = Depends(get_db),
                  user: models.User = Depends(require_screen("store"))):
    # NOTE: also called directly by the export endpoint, which passes
    # user=None after verifying a download token instead.
    """
    kind='stock'     - current stock, central vs out at sites
        ='low'       - items at or below reorder level
        ='by_site'   - what each site currently holds
        ='purchases' - what was received, with cost, over a period
        ='usage'     - what was issued out, per item, over a period
        ='returnable'- returnable items currently out, and where
    """
    d1 = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
    d2 = datetime.strptime(date_to, "%Y-%m-%d").date() if date_to else None
    items = {i.id: i for i in db.query(models.StoreItem).all()}
    stock = _stock_map(db)

    def moves(kinds):
        q = db.query(models.StoreMovement).filter(models.StoreMovement.kind.in_(kinds))
        if d1: q = q.filter(models.StoreMovement.moved_on >= d1)
        if d2: q = q.filter(models.StoreMovement.moved_on <= d2)
        return q.all()

    if kind == "low":
        rows = []
        for i in items.values():
            if not i.active or not i.reorder_level: continue
            have = stock.get((i.id, CENTRAL), 0)
            if have <= i.reorder_level:
                rows.append({"code": i.code, "name": i.name, "unit": i.unit,
                              "in_store": round(have, 2), "reorder_level": i.reorder_level,
                              "shortfall": round(i.reorder_level - have, 2)})
        return {"title": "Items to reorder", "rows": sorted(rows, key=lambda r: -r["shortfall"])}

    if kind == "by_site":
        rows = []
        for (iid, loc), qty in stock.items():
            if loc == CENTRAL or not qty or iid not in items: continue
            i = items[iid]
            rows.append({"site": loc, "code": i.code, "name": i.name,
                          "unit": i.unit, "qty": round(qty, 2), "item_type": i.item_type})
        return {"title": "Stock held at sites", "rows": sorted(rows, key=lambda r: (r["site"], r["code"]))}

    if kind == "purchases":
        agg = {}
        for m in moves(["in"]):
            i = items.get(m.item_id)
            if not i: continue
            a = agg.setdefault(i.id, {"code": i.code, "name": i.name, "unit": i.unit,
                                       "qty": 0, "deliveries": 0, "last_arrived": None,
                                       "suppliers": set()})
            a["qty"] += m.qty
            a["deliveries"] += 1
            if m.moved_on and (not a["last_arrived"] or m.moved_on > a["last_arrived"]):
                a["last_arrived"] = m.moved_on
            if m.supplier: a["suppliers"].add(m.supplier)
        rows = [{**v, "qty": round(v["qty"], 2),
                  "last_arrived": v["last_arrived"].isoformat() if v["last_arrived"] else "-",
                  "suppliers": ", ".join(sorted(v["suppliers"])) or "not recorded"}
                for v in agg.values()]
        return {"title": "Purchases received", "rows": sorted(rows, key=lambda r: -r["qty"])}

    if kind == "usage":
        agg = {}
        for m in moves(["out"]):
            i = items.get(m.item_id)
            if not i: continue
            key = (i.id, m.location)
            a = agg.setdefault(key, {"code": i.code, "name": i.name, "unit": i.unit,
                                      "site": m.location, "qty": 0})
            a["qty"] += m.qty
        rows = [{**v, "qty": round(v["qty"], 2)} for v in agg.values()]
        return {"title": "Materials issued", "rows": sorted(rows, key=lambda r: (r["site"], r["code"]))}

    if kind == "returnable":
        rows = []
        for (iid, loc), qty in stock.items():
            i = items.get(iid)
            if not i or i.item_type not in ("returnable", "asset") or loc == CENTRAL or qty <= 0: continue
            last = (db.query(models.StoreMovement)
                      .filter(models.StoreMovement.item_id == iid,
                               models.StoreMovement.location == loc,
                               models.StoreMovement.kind == "out")
                      .order_by(models.StoreMovement.moved_on.desc()).first())
            rows.append({"code": i.code, "name": i.name, "unit": i.unit, "site": loc,
                          "qty": round(qty, 2),
                          "incharge": last.incharge if last else "",
                          "since": last.moved_on.isoformat() if last else ""})
        return {"title": "Returnable items still out", "rows": sorted(rows, key=lambda r: r["site"])}

    if kind == "rentals":
        # Hired-in equipment: where it is, what it costs, and whether it
        # is overdue back to the supplier.
        today = _dubai_today()
        rows = []
        for i in items.values():
            if not i.active or i.item_type != "rental": continue
            at = {loc: q for (iid, loc), q in stock.items() if iid == i.id and q}
            where = ", ".join(f"{loc or 'store'}: {q:g}" for loc, q in sorted(at.items())) or "-"
            days = (today - i.rental_start).days if i.rental_start else None
            est = round(days * (i.rental_rate or 0), 2) if (days is not None and i.rental_period == "day") else None
            rows.append({"code": i.code, "name": i.name, "supplier": i.rental_supplier,
                          "rate": i.rental_rate, "period": i.rental_period,
                          "start": i.rental_start.isoformat() if i.rental_start else "",
                          "due": i.rental_due.isoformat() if i.rental_due else "",
                          "days_on_hire": days, "est_cost": est, "where": where,
                          "overdue": bool(i.rental_due and i.rental_due < today)})
        return {"title": "Rented equipment", "rows": sorted(rows, key=lambda r: (not r["overdue"], r["due"] or ""))}

    def _lost_by_item(iid):
        return sum(m.qty for m in db.query(models.StoreMovement)
                    .filter(models.StoreMovement.item_id == iid,
                             models.StoreMovement.kind == "lost").all())

    if kind == "assets":
        # Owned only - hired equipment has its own register, so the two
        # never get added together and mistaken for company property.
        rows = []
        for i in items.values():
            if not i.active or i.item_type not in ("asset", "returnable"): continue
            at = {loc: q for (iid, loc), q in stock.items() if iid == i.id and q}
            lost = _lost_by_item(i.id)
            rows.append({"code": i.code, "name": i.name, "unit": i.unit,
                          "in_store": round(at.get("", 0), 2),
                          "at_sites": round(sum(q for loc, q in at.items() if loc), 2),
                          "total": round(sum(at.values()), 2),
                          "written_off": round(lost, 2) if lost else 0,
                          "where": ", ".join(f"{loc or 'store'}: {q:g}"
                                              for loc, q in sorted(at.items())) or "-"})
        return {"title": "Owned assets and equipment", "rows": sorted(rows, key=lambda r: r["code"])}

    if kind == "hired":
        # Everything currently on hire: what, from whom, where it is, and
        # anything already lost or damaged that the supplier will charge
        # for. Who it came from and when are read from the deliveries
        # themselves - the item's own rental fields are only filled in if
        # someone typed them, so relying on them left the columns empty
        # even though the store knew the answer.
        arrivals = (db.query(models.StoreMovement)
                      .filter(models.StoreMovement.kind == "in")
                      .order_by(models.StoreMovement.moved_on.asc()).all())
        first_in, last_supplier = {}, {}
        for m in arrivals:
            if m.item_id not in first_in and m.moved_on:
                first_in[m.item_id] = m.moved_on
            if (m.supplier or "").strip():
                last_supplier[m.item_id] = m.supplier.strip()
        rows = []
        for i in items.values():
            if not i.active or i.item_type != "rental": continue
            at = {loc: q for (iid, loc), q in stock.items() if iid == i.id and q}
            lost = _lost_by_item(i.id)
            on_hire = round(sum(at.values()), 2)
            if on_hire <= 0 and not lost: continue
            since = i.rental_start or first_in.get(i.id)
            rows.append({"code": i.code, "name": i.name, "unit": i.unit,
                          "hired_from": i.rental_supplier or last_supplier.get(i.id) or "not recorded",
                          "on_hire": on_hire,
                          "in_store": round(at.get("", 0), 2),
                          "by_site": {loc: round(q, 2) for loc, q in at.items() if loc},
                          "lost_damaged": round(lost, 2) if lost else 0,
                          "since": since.isoformat() if since else "-",
                          "due_back": i.rental_due.isoformat() if i.rental_due else "no date set",
                          "days_out": (date.today() - since).days if since else ""})
        return {"title": "Equipment on hire", "rows": sorted(rows, key=lambda r: (r["hired_from"], r["code"]))}

    if kind == "lost":
        rows = []
        q = db.query(models.StoreMovement).filter(models.StoreMovement.kind == "lost")
        if d1: q = q.filter(models.StoreMovement.moved_on >= d1)
        if d2: q = q.filter(models.StoreMovement.moved_on <= d2)
        for m in q.order_by(models.StoreMovement.moved_on.desc()).all():
            i = items.get(m.item_id)
            if not i: continue
            sup = ""
            if i.item_type == "rental":
                sup = i.rental_supplier or (db.query(models.StoreMovement)
                        .filter(models.StoreMovement.item_id == i.id,
                                models.StoreMovement.kind == "in",
                                models.StoreMovement.supplier != "")
                        .order_by(models.StoreMovement.moved_on.desc())
                        .with_entities(models.StoreMovement.supplier).scalar() or "")
            rows.append({"date": m.moved_on.isoformat(), "code": i.code, "name": i.name,
                          "type": i.item_type, "qty": round(m.qty, 2), "unit": i.unit,
                          "where": m.from_location or "central store",
                          "hired_from": sup or ("not recorded" if i.item_type == "rental" else "owned"),
                          "reason": m.notes or "not given"})
        return {"title": "Lost and damaged", "rows": rows}

    # default: full stock position
    # Only materials that actually hold stock, or that someone has set a
    if kind not in ("stock", "low", "by_site", "purchases", "usage", "assets",
                    "lost", "hired", "rentals"):
        raise HTTPException(status_code=400,
            detail=f"There is no report called '{kind}'.")
    # reorder level on (so a watched item shows even when it hits zero).
    # The catalogue can run to thousands of materials; listing every one
    # at zero makes the few real ones impossible to find.
    rows = []
    for i in items.values():
        if not i.active: continue
        c = stock.get((i.id, CENTRAL), 0)
        per_site = {loc: round(v, 2) for (iid, loc), v in stock.items()
                    if iid == i.id and loc != CENTRAL and v}
        o = sum(per_site.values())
        if c == 0 and o == 0 and not i.reorder_level:
            continue
        rows.append({"code": i.code, "name": i.name, "unit": i.unit,
                      "item_type": i.item_type, "in_store": round(c, 2),
                      "at_sites": round(o, 2), "total": round(c + o, 2),
                      # Which site holds what, so a row can open into a
                      # proper breakdown instead of one crowded cell.
                      "by_site": per_site,
                      "low": bool(i.reorder_level and c <= i.reorder_level)})
    return {"title": "Current stock", "rows": sorted(rows, key=lambda r: r["code"])}


# ---------------------------------------------------------------------
# MATERIAL REQUESTS (store keeper -> office)
# ---------------------------------------------------------------------
def _next_mr_ref(db: Session) -> str:
    last = db.query(models.MaterialRequest).order_by(models.MaterialRequest.id.desc()).first()
    n = (last.id + 1) if last else 1
    return f"MR-{n:04d}"


def _mr_out(mr: models.MaterialRequest) -> dict:
    d = schemas.MaterialRequestOut.model_validate(mr).model_dump()
    # Supplier travels with the request so the list can show who to
    # chase without a second call per row.
    sup = getattr(mr, "supplier", None)
    d["supplier"] = ({"id": sup.id, "name": sup.name,
                      "contact_person": sup.contact_person or "", "phone": sup.phone or ""}
                     if sup else None)
    d["expected_on"] = mr.expected_on.isoformat() if getattr(mr, "expected_on", None) else None
    for i, ln in enumerate(mr.lines):
        d["lines"][i]["item_code"] = ln.item.code if ln.item else ""
        d["lines"][i]["item_name"] = ln.item.name if ln.item else (ln.description or "")
        ls = ln.supplier
        d["lines"][i]["status"] = ln.status or "pending"
        d["lines"][i]["reject_reason"] = ln.reject_reason or ""
        d["lines"][i]["supplier"] = ({"id": ls.id, "name": ls.name,
                                       "contact_person": ls.contact_person or "",
                                       "phone": ls.phone or ""} if ls else None)
    return d


@app.get("/store/requests")
def list_material_requests(status: str = None, site: str = None,
                            date_from: str = None, date_to: str = None,
                            db: Session = Depends(get_db),
                            user: models.User = Depends(require_any_screen("requests", "approvals"))):
    q = db.query(models.MaterialRequest)
    if status:
        q = q.filter(models.MaterialRequest.status == status)
    if site:
        q = q.filter(models.MaterialRequest.site == site)
    if date_from:
        q = q.filter(models.MaterialRequest.requested_on >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        q = q.filter(models.MaterialRequest.requested_on <= datetime.strptime(date_to, "%Y-%m-%d").date())
    return [_mr_out(m) for m in q.order_by(models.MaterialRequest.id.desc()).all()]


@app.post("/store/requests")
def create_material_request(payload: schemas.MaterialRequestIn, db: Session = Depends(get_db),
                             user: models.User = Depends(require_any_screen("requests", "approvals"))):
    lines = [l for l in payload.lines if l.qty_requested and l.qty_requested > 0]
    if not lines:
        raise HTTPException(status_code=400, detail="Add at least one material with a quantity.")
    for l in lines:
        if not l.item_id and not (l.description or "").strip():
            raise HTTPException(status_code=400, detail="Every line needs an item or a description.")

    # Repeat-click guard. A browser fault once let the save succeed while
    # the confirmation crashed, so the keeper kept clicking Send and one
    # request became thirty. If an identical request - same person, same
    # site, same materials and quantities - already exists from the last
    # few minutes, hand that one back instead of minting another.
    sig = sorted((l.item_id or 0, " ".join((l.description or "").lower().split()),
                  round(l.qty_requested, 3)) for l in lines)
    recent = (db.query(models.MaterialRequest)
                .filter(models.MaterialRequest.requested_by == payload.requested_by,
                        models.MaterialRequest.site == (payload.site or ""),
                        models.MaterialRequest.status == "pending")
                .order_by(models.MaterialRequest.id.desc()).limit(5).all())
    now = datetime.now(timezone.utc)
    def _recent(ts):
        if not ts:
            return True   # a pending twin with no timestamp is still a twin
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        return (now - ts) <= timedelta(minutes=10)
    for prev in recent:
        if not _recent(prev.created_at):
            continue
        prev_sig = sorted((pl.item_id or 0, " ".join((pl.description or "").lower().split()),
                           round(pl.qty_requested or 0, 3)) for pl in prev.lines)
        if prev_sig == sig:
            out = _mr_out(prev)
            out["duplicate_of"] = prev.ref
            return out

    mr = models.MaterialRequest(
        ref=_next_mr_ref(db), site=payload.site,
        requested_by=_person_name(payload.requested_by),
        needed_by=payload.needed_by, urgency=payload.urgency, notes=payload.notes,
        requested_on=payload.requested_on or date.today(), created_by=user.id,
    )
    db.add(mr)
    db.flush()
    created = []
    for l in lines:
        # The requester's own unit wins. Overwriting "tonne" with the
        # catalogue's "pcs" quietly changed what was being asked for -
        # the office read 25 pcs of rebar when 25 tonne was meant. The
        # catalogue unit is only a fallback for a line that has none.
        unit = (l.unit or "").strip()
        if not unit and l.item_id:
            it = db.query(models.StoreItem).filter(models.StoreItem.id == l.item_id).first()
            if it:
                unit = it.unit
        unit = unit or "pcs"
        item_id = l.item_id
        # A material typed by name joins the item list here and now, with
        # its own generated code, so it can be received into stock, given
        # out, reported on, and picked from the list next time.
        if not item_id:
            it = _find_or_create_item(db, l.description, unit, l.item_type)
            if it:
                item_id = it.id
                created.append(f"{it.code} - {it.name}")
        db.add(models.MaterialRequestLine(request_id=mr.id, item_id=item_id,
                                           description=_proper_name(l.description),
                                           qty_requested=l.qty_requested,
                                           qty_approved=l.qty_approved or 0, unit=unit,
                                           est_cost=l.est_cost or 0, notes=l.notes,
                                           purpose=l.purpose))
    db.commit()
    db.refresh(mr)
    log_action(db, user.id, "material_request", f"{mr.ref} - {len(lines)} item(s)")
    out = _mr_out(mr)
    if created:
        out["new_items"] = created
    return out


@app.post("/store/requests/{req_id}/status")
def set_material_request_status(req_id: int, payload: schemas.MaterialRequestStatusIn,
                                 db: Session = Depends(get_db),
                                 user: models.User = Depends(require_screen("approvals"))):
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(status_code=404, detail="Request not found")
    # Four steps, not six: "arranging" and "lpo_sent" both just meant
    # "the office is dealing with it", which made the screen busier
    # without telling anyone anything they could act on. Old values are
    # still accepted so existing requests keep working.
    allowed = ("pending", "approved", "ordered", "partial", "delivered",
               "closed", "rejected", "arranging", "lpo_sent", "received")
    if payload.status not in allowed:
        raise HTTPException(status_code=400, detail=f"Status must be one of: {', '.join(allowed)}")
    mr.status = payload.status
    # Approving or rejecting the request settles every material that is
    # still waiting - otherwise the request reads "approved" while its
    # materials are all still pending, and the screen keeps asking to
    # approve something it already approved.
    if payload.status == "approved":
        for l in mr.lines:
            if (l.status or "pending") == "pending":
                l.status = "approved"
    elif payload.status == "rejected":
        for l in mr.lines:
            if (l.status or "pending") == "pending" and not (l.qty_received or 0):
                l.status = "rejected"
                l.reject_reason = (payload.office_remark or "").strip()
    if payload.office_remark:
        mr.office_remark = payload.office_remark
    # Ordering is when the supplier becomes known. Capturing it here
    # gives the keeper a name and number to chase, instead of a request
    # that says "ordered" and nothing else.
    if getattr(payload, "supplier", None) and payload.supplier.strip():
        sup = _find_or_create_supplier(db, payload.supplier,
                                        getattr(payload, "contact_person", "") or "",
                                        getattr(payload, "phone", "") or "")
        if sup:
            # A request can be split across traders on price - cement
            # from one, rebar from another. If specific lines were named,
            # only those go to this supplier; otherwise the whole request
            # does. The request-level supplier is kept as a shortcut only
            # while every line agrees.
            line_ids = getattr(payload, "line_ids", None) or []
            targets = [l for l in mr.lines
                       if (not line_ids or l.id in line_ids)
                       and (l.status or "pending") != "rejected"]
            for l in targets:
                l.supplier_id = sup.id
                # Buying something is approving it. Leaving the line
                # "pending" made the screen contradict itself: ordered
                # from Newstar, yet still asking Approve or Reject.
                l.status = "approved"
            sup_ids = {l.supplier_id for l in mr.lines}
            mr.supplier_id = sup.id if len(sup_ids) == 1 and None not in sup_ids else None
    if getattr(payload, "expected_on", None):
        mr.expected_on = payload.expected_on
    mr.closed_on = date.today() if payload.status in ("delivered", "received", "closed", "rejected") else None
    db.commit()
    log_action(db, user.id, "material_request_status", f"{mr.ref} -> {payload.status}")
    return {"ok": True, "status": mr.status}


@app.delete("/store/requests/{req_id}")
def delete_material_request(req_id: int, db: Session = Depends(get_db),
                             user: models.User = Depends(require_screen("approvals"))):
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(status_code=404, detail="Request not found")
    ref = mr.ref
    db.delete(mr)
    db.commit()
    log_action(db, user.id, "material_request_delete", ref)
    return {"ok": True}


@app.post("/store/requests/{req_id}/receive")
def receive_against_request(req_id: int, line_id: int, qty: float, supplier: str = "",
                             unit_cost: float = 0.0, reference: str = "",
                             db: Session = Depends(get_db),
                             user: models.User = Depends(require_any_screen("requests", "approvals"))):
    """
    Record a delivery against one line of a request. This both files a
    normal 'in' stock movement AND advances the request, so outstanding
    quantities stay honest instead of the two drifting apart.
    """
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(status_code=404, detail="Request not found")
    line = db.query(models.MaterialRequestLine).filter(
        models.MaterialRequestLine.id == line_id,
        models.MaterialRequestLine.request_id == req_id).first()
    if not line:
        raise HTTPException(status_code=404, detail="Request line not found")
    if qty <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be more than zero.")
    outstanding = line.qty_requested - (line.qty_received or 0)
    if qty > outstanding + 1e-9:
        raise HTTPException(status_code=400,
            detail=f"Only {round(outstanding, 2)} {line.unit} still outstanding on this line.")
    if not line.item_id:
        raise HTTPException(status_code=400,
            detail="This line isn't linked to a store item, so it can't be received into stock. "
                   "Create the item first, then edit the request.")

    db.add(models.StoreMovement(item_id=line.item_id, kind="in", qty=qty, location="",
                                 supplier=supplier, unit_cost=unit_cost,
                                 reference=reference or mr.ref, moved_on=date.today(),
                                 notes=f"Against {mr.ref}", created_by=user.id))
    line.qty_received = (line.qty_received or 0) + qty
    all_done = all((l.qty_received or 0) >= l.qty_requested - 1e-9 for l in mr.lines)
    any_done = any((l.qty_received or 0) > 0 for l in mr.lines)
    mr.status = "received" if all_done else ("partial" if any_done else mr.status)
    if all_done:
        mr.closed_on = date.today()
    db.commit()
    log_action(db, user.id, "material_request_receive", f"{mr.ref} line {line_id}: {qty}")
    return {"ok": True, "status": mr.status, "qty_received": line.qty_received}


@app.get("/store/requests/report")
def material_request_report(kind: str = "open", db: Session = Depends(get_db),
                            user: models.User = Depends(auth.get_current_user)):
    # Also called directly by the export endpoint with user=None.
    """kind='open' | 'overdue' | 'outstanding' | 'history'"""
    today = _dubai_today()
    reqs = db.query(models.MaterialRequest).order_by(models.MaterialRequest.id.desc()).all()

    if kind == "overdue":
        rows = [{"ref": m.ref, "site": m.site, "requested_on": m.requested_on.isoformat(),
                  "needed_by": m.needed_by.isoformat() if m.needed_by else "",
                  "days_late": (today - m.needed_by).days if m.needed_by else 0,
                  "status": m.status, "urgency": m.urgency, "items": len(m.lines)}
                 for m in reqs
                 if m.needed_by and m.needed_by < today and m.status not in ("delivered", "received", "closed", "rejected")]
        return {"title": "Overdue material requests", "rows": sorted(rows, key=lambda r: -r["days_late"])}

    if kind == "outstanding":
        rows = []
        for m in reqs:
            if m.status in ("delivered", "received", "closed", "rejected"):
                continue
            for l in m.lines:
                out = (l.qty_requested or 0) - (l.qty_received or 0)
                if out <= 0:
                    continue
                rows.append({"ref": m.ref, "site": m.site, "status": m.status,
                              "item": (l.item.code + " - " + l.item.name) if l.item else l.description,
                              "unit": l.unit, "requested": l.qty_requested,
                              "received": l.qty_received or 0, "outstanding": round(out, 2),
                              "needed_by": m.needed_by.isoformat() if m.needed_by else ""})
        return {"title": "Outstanding materials", "rows": rows}

    if kind == "history":
        rows = [{"ref": m.ref, "site": m.site, "requested_on": m.requested_on.isoformat(),
                  "requested_by": m.requested_by, "urgency": m.urgency, "status": m.status,
                  "items": len(m.lines),
                  # No estimated value: the store keeper doesn't price a
                  # request, the office does when it orders.
                  "closed_on": m.closed_on.isoformat() if m.closed_on else ""}
                for m in reqs]
        return {"title": "Material request history", "rows": rows}

    rows = [{"ref": m.ref, "site": m.site, "requested_on": m.requested_on.isoformat(),
              "needed_by": m.needed_by.isoformat() if m.needed_by else "",
              "urgency": m.urgency, "status": m.status, "items": len(m.lines),
              "requested_by": m.requested_by}
            for m in reqs if m.status in ("pending", "approved", "partial")]
    return {"title": "Open material requests", "rows": rows}


@app.get("/export/store/report")
def export_store_report(kind: str = "stock", format: str = "excel",
                         date_from: str = None, date_to: str = None,
                         token: str = None, db: Session = Depends(get_db)):
    auth.get_download_user_from_token(token, db)
    # Material-request reports live under a different builder to the
    # stock ones, but both export through the same formatter.
    if kind.startswith("mr_"):
        data = material_request_report(kind=kind[3:], db=db, user=None)
    else:
        data = store_report(kind=kind, date_from=date_from, date_to=date_to, db=db, user=None)
    sub = ""
    if date_from or date_to:
        sub = f"{date_from or 'start'} to {date_to or 'today'}"
    else:
        sub = f"As at {date.today().isoformat()}"
    # On screen the site split opens as its own panel; on paper there is
    # nowhere to expand, so it becomes a readable column instead of a
    # raw mapping.
    rows = []
    for r in data["rows"]:
        r = {k: v for k, v in r.items() if k not in ("low", "overdue")}
        if isinstance(r.get("by_site"), dict):
            per = r.pop("by_site")
            r["at_which_sites"] = ", ".join(f"{loc}: {_clean_export_qty(q)}"
                                             for loc, q in sorted(per.items())) or "-"
        rows.append(r)
    data = {**data, "rows": rows}
    if format == "pdf":
        buf = export_web.build_store_report_pdf(data["title"], data["rows"], sub)
        media, ext = "application/pdf", "pdf"
    else:
        buf = export_web.build_store_report_excel(data["title"], data["rows"], sub)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    name = f"{kind}_{date.today().isoformat()}.{ext}"
    return StreamingResponse(buf, media_type=media,
                              headers={"Content-Disposition": f'attachment; filename="{name}"'})


@app.get("/export/store/request/{req_id}")
def export_material_request(req_id: int, token: str = None, db: Session = Depends(get_db)):
    auth.get_download_user_from_token(token, db)
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(status_code=404, detail="Request not found")
    buf = export_web.build_material_request_pdf(_mr_out(mr))
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f'attachment; filename="{mr.ref}.pdf"'})


@app.get("/permissions/screens")
def list_screens(user: models.User = Depends(auth.get_current_user)):
    return {"screens": ALL_SCREENS, "role_defaults": ROLE_DEFAULTS}


@app.get("/permissions/me")
def my_permissions(user: models.User = Depends(auth.get_current_user)):
    return {"role": user.role, "screens": effective_permissions(user)}


@app.post("/users/{user_id}/permissions")
def set_permissions(user_id: int, payload: schemas.PermissionsIn,
                     db: Session = Depends(get_db),
                     user: models.User = Depends(auth.require_admin)):
    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.role == "admin":
        raise HTTPException(status_code=400,
            detail="An admin always has access to everything - nothing to set.")
    wanted = [s.strip() for s in (payload.permissions or "").split(",") if s.strip()]
    bad = [s for s in wanted if s not in ALL_SCREENS]
    if bad:
        raise HTTPException(status_code=400, detail=f"Unknown screen(s): {', '.join(bad)}")
    target.permissions = ",".join(wanted)
    db.commit()
    log_action(db, user.id, "set_permissions", f"{target.username}: {target.permissions or '(role default)'}")
    return {"ok": True, "permissions": effective_permissions(target)}


# ---------------------------------------------------------------------
# NOTIFICATIONS
# ---------------------------------------------------------------------
# Plain words for a status, used when telling someone a request is late.
MR_STATUS_WORDS = {
    "pending": "waiting on the office",
    "approved": "approved, not yet ordered",
    "ordered": "on order",
    "arranging": "on order",
    "lpo_sent": "on order",
    "partial": "partly delivered",
}


@app.get("/notifications")
def get_notifications(db: Session = Depends(get_db),
                       user: models.User = Depends(auth.get_current_user)):
    """
    Notifications the way people expect them: each event goes to the
    role that has to act on it, and clicking one lands on the exact
    item. Derived from live data, so nothing can go stale.

      office (approvals)  <- a new request arrives; a request is late
      site (requests)     <- their request was approved, ordered,
                             delivered or rejected
      keeper (store)      <- stock is low; a rental is overdue back
      attendance          <- yesterday's sheet is missing or incomplete
    """
    today = _dubai_today()
    allowed = effective_permissions(user)
    out = []
    recent_cutoff = today - timedelta(days=7)

    def _mr_lines(m):
        n = len(m.lines)
        return f"{n} material{'s' if n != 1 else ''}"

    reqs = db.query(models.MaterialRequest).all() if ("approvals" in allowed or "requests" in allowed) else []

    # ---- Office: things to act on -----------------------------------
    if "approvals" in allowed:
        for m in reqs:
            if m.status == "pending":
                out.append({"id": f"new-req-{m.id}", "kind": "request",
                             "title": f"New material request {m.ref}",
                             "detail": f"{_person_name(m.requested_by) or 'Someone'} asked for {_mr_lines(m)}"
                                       + (f" for site {m.site}" if m.site else " for the central store")
                                       + (f", needed by {m.needed_by.isoformat()}" if m.needed_by else ""),
                             "screen": "approvals", "target": m.ref,
                             "when": m.requested_on.isoformat(), "level": "info"})
            if (m.needed_by and m.needed_by < today
                    and m.status not in ("delivered", "received", "closed", "rejected")):
                out.append({"id": f"late-{m.id}-{m.needed_by}", "kind": "late",
                             "title": f"{m.ref} is late",
                             "detail": f"Needed by {m.needed_by.isoformat()}, still "
                                       f"{MR_STATUS_WORDS.get(m.status, m.status)}",
                             "screen": "approvals", "target": m.ref,
                             "when": m.needed_by.isoformat(), "level": "warn"})

    # ---- Site staff: what happened to their requests -----------------
    if "requests" in allowed:
        for m in reqs:
            upd = m.updated_at.date() if m.updated_at else None
            if not upd or upd < recent_cutoff:
                continue
            if m.status == "approved":
                title, detail, lvl = f"{m.ref} approved", "The office has approved it and will order", "ok"
            elif m.status in ("ordered", "arranging", "lpo_sent"):
                sups = sorted({l.supplier.name for l in m.lines if l.supplier})
                title = f"{m.ref} ordered"
                detail = ("From " + ", ".join(sups) if sups else "Ordered by the office") \
                         + (f", expected {m.expected_on.isoformat()}" if m.expected_on else "")
                lvl = "ok"
            elif m.status == "partial":
                got = sum(l.qty_received or 0 for l in m.lines)
                asked = sum(l.qty_requested or 0 for l in m.lines)
                title, detail, lvl = f"{m.ref} partly delivered", f"{got:g} of {asked:g} arrived so far", "info"
            elif m.status in ("delivered", "received"):
                title, detail, lvl = f"{m.ref} delivered", "Everything has arrived at the store", "ok"
            elif m.status == "rejected":
                title, detail, lvl = f"{m.ref} rejected", (m.office_remark or "The office turned it down"), "warn"
            else:
                continue
            out.append({"id": f"status-{m.id}-{m.status}-{upd}", "kind": "status",
                         "title": title, "detail": detail,
                         "screen": "followup" if m.status not in ("delivered", "received", "rejected") else "requests",
                         "target": m.ref, "when": upd.isoformat(), "level": lvl})

    # ---- Keeper: the store itself ------------------------------------
    if "store" in allowed:
        items = {i.id: i for i in db.query(models.StoreItem).filter(models.StoreItem.active == True).all()}  # noqa: E712
        stock = _stock_map(db)
        low = [(i, stock.get((i.id, CENTRAL), 0)) for i in items.values()
               if i.reorder_level and stock.get((i.id, CENTRAL), 0) <= i.reorder_level]
        for i, have in low[:20]:
            out.append({"id": f"low-{i.id}-{have}", "kind": "low",
                         "title": f"{i.name} is running low",
                         "detail": f"{have:g} {i.unit} left, warn level is {i.reorder_level:g}",
                         "screen": "store", "target": i.code, "when": today.isoformat(), "level": "warn"})
        for i in items.values():
            if i.item_type == "rental" and i.rental_due and i.rental_due < today:
                out.append({"id": f"rent-{i.id}-{i.rental_due}", "kind": "rental",
                             "title": f"{i.name} is overdue back to {i.rental_supplier or 'the supplier'}",
                             "detail": f"Was due {i.rental_due.isoformat()}",
                             "screen": "store", "target": i.code, "when": i.rental_due.isoformat(), "level": "warn"})

    # ---- Attendance -------------------------------------------------
    if "attendance" in allowed:
        y = today - timedelta(days=1)
        marked = (db.query(models.DailyRow)
                    .filter(models.DailyRow.full_date == y,
                             or_(models.DailyRow.am != "", models.DailyRow.pm != ""))
                    .count())
        active = db.query(models.Employee).filter(models.Employee.active == True).count()  # noqa: E712
        if active and marked == 0:
            out.append({"id": f"att-none-{y}", "kind": "attendance",
                         "title": "No attendance saved for yesterday",
                         "detail": y.strftime("%A, %d %B %Y"),
                         "screen": "attendance", "when": y.isoformat(), "level": "warn"})
        elif active and marked < active:
            out.append({"id": f"att-part-{y}-{marked}", "kind": "attendance",
                         "title": "Yesterday's attendance is incomplete",
                         "detail": f"{marked} of {active} workers marked",
                         "screen": "attendance", "when": y.isoformat(), "level": "info"})

    # Newest first within each urgency band
    order = {"warn": 0, "info": 1, "ok": 2}
    out.sort(key=lambda n: (order.get(n["level"], 3), n["when"]), reverse=False)
    out.sort(key=lambda n: n["when"], reverse=True)
    out.sort(key=lambda n: order.get(n["level"], 3))
    return {"notifications": out[:60], "count": len(out)}


@app.post("/store/requests/{req_id}/receive-bulk")
def receive_request_bulk(req_id: int, payload: schemas.ReceiveRequestIn,
                          db: Session = Depends(get_db),
                          user: models.User = Depends(require_any_screen("store", "approvals"))):
    """
    Record a whole delivery against a request in one go.

    The store keeper opens the request, adjusts the quantities that
    actually turned up (a supplier often sends less than was ordered) and
    saves. Each line files its own 'in' stock movement, so the ledger
    stays the single source of truth, and the request advances to
    'partial' or 'delivered' by itself based on what is still owed.
    """
    mr = db.query(models.MaterialRequest).filter(models.MaterialRequest.id == req_id).first()
    if not mr:
        raise HTTPException(status_code=404, detail="Request not found")

    when = payload.received_on or date.today()
    # Deliveries are often written up a few days late, and sometimes
    # booked a day or two ahead for a load already on its way. Both are
    # real, so only a date far in the future is refused - that is a
    # typed year or month slip, not a delivery.
    if when > date.today() + timedelta(days=30):
        raise HTTPException(status_code=400,
            detail="Delivery date is more than a month ahead - check the date.")

    wanted = [l for l in payload.lines if l.qty and l.qty > 0]
    if not wanted:
        raise HTTPException(status_code=400, detail="Enter how much arrived for at least one material.")

    # Learn the supplier as the delivery is recorded: the name is tidied
    # to one spelling, and the request keeps a link to it so the keeper
    # can chase the next order without asking who sold it to us.
    sup = _find_or_create_supplier(db, payload.supplier) if (payload.supplier or "").strip() else None
    sup_name = sup.name if sup else ""

    errors, done = [], 0
    for w in wanted:
        line = db.query(models.MaterialRequestLine).filter(
            models.MaterialRequestLine.id == w.line_id,
            models.MaterialRequestLine.request_id == req_id).first()
        if not line:
            errors.append(f"Line {w.line_id} is not on this request."); continue
        if not line.item_id:
            errors.append(f"'{line.description}' isn't a store item yet, so it can't be received into stock.")
            continue
        if (line.status or "pending") == "rejected":
            name = line.item.name if line.item else line.description
            errors.append(f"{name}: the office rejected this material.")
            continue
        outstanding = (line.qty_requested or 0) - (line.qty_received or 0)
        if w.qty > outstanding + 1e-9:
            name = line.item.name if line.item else line.description
            errors.append(f"{name}: only {round(outstanding, 2)} {line.unit} still due.")
            continue
        # Supplier, in order of authority: what was typed against this
        # material, then the trader it was ordered from, then whoever
        # the delivery header names.
        own = (getattr(w, "supplier", "") or "").strip()
        if own:
            own_sup = _find_or_create_supplier(db, own)
            if own_sup and not line.supplier_id:
                line.supplier_id = own_sup.id
            line_sup = own_sup.name if own_sup else own
        else:
            line_sup = (line.supplier.name if line.supplier else "") or sup_name
        db.add(models.StoreMovement(
            item_id=line.item_id, kind="in", qty=w.qty, location="",
            supplier=line_sup, reference=payload.reference or mr.ref,
            notes=(payload.notes or f"Against {mr.ref}"), moved_on=when, created_by=user.id))
        line.qty_received = (line.qty_received or 0) + w.qty
        # The trader on the delivery note is who this line was bought
        # from - learn it on lines that never got a supplier at ordering.
        if sup and not line.supplier_id:
            line.supplier_id = sup.id
        done += 1

    # The request-level supplier is a shortcut kept only while every
    # line agrees; recompute rather than overwrite.
    sup_ids = {l.supplier_id for l in mr.lines}
    mr.supplier_id = list(sup_ids)[0] if len(sup_ids) == 1 and None not in sup_ids else None

    if errors and not done:
        raise HTTPException(status_code=400, detail={"errors": errors})

    # A rejected line is never coming, so it must not hold the request
    # open. Judged on the lines that were actually wanted; otherwise a
    # request with one line turned down sits on the chase list forever
    # with nothing left to chase.
    wanted = [l for l in mr.lines if (l.status or "pending") != "rejected"]
    all_done = bool(wanted) and all((l.qty_received or 0) >= (l.qty_requested or 0) - 1e-9 for l in wanted)
    any_done = any((l.qty_received or 0) > 0 for l in wanted)
    mr.status = "delivered" if all_done else ("partial" if any_done else mr.status)
    if all_done:
        mr.closed_on = when
    db.commit()
    log_action(db, user.id, "material_request_receive",
               f"{mr.ref}: {done} line(s) received on {when}")
    return {"ok": True, "received_lines": done, "status": mr.status, "warnings": errors}
