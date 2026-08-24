"""
Infinia Labour Tool - Data Engine
==================================
Reads Labour Ecard .xlsx files, consolidates them, and produces every report.

ACCURACY DESIGN NOTE (important - read before modifying):
Every payroll figure (Final Salary, OT Amount, BH Amount, Deduction) is
INDEPENDENTLY RECALCULATED here from the raw daily attendance grid, using
the exact same formulas as the source card:

    Present = (COUNTIF(AM,"Present") + COUNTIF(PM,"Present")) / 2   [same for
               Absent/Sick/Medical/Friday/Holiday/Leave]
    OT hours   = SUM(OT column)
    BH hours   = SUM(BH column)
    Total Salary (pay component) = (Present+Sick+Medical+Friday+Holiday) * (Salary/30)
    Deduction                    = (Salary/30) * Absent
    OT Amount                    = (Salary/30/8) * OT hours
    BH Amount                    = (Salary/30/8) * BH hours
    Final Salary = ROUND(TotalSalary + OT Amount + BH Amount + Allowances
                          - Deduction - OtherDeduction, 0)

We do NOT simply trust the cached formula result stored in the file (cell
H44 etc.), because that cache can be stale or the file may not have been
recalculated. Instead we recompute from source, AND read the cached value
if present, and compare the two. Any mismatch is written to the Import
Issues log so it can be checked by a human before it's used for payroll.
"""

import os
import math
import re
from datetime import date, datetime
from dataclasses import dataclass, field

import openpyxl

# ---------------------------------------------------------------------------
# Fixed cell map for the Labour Ecard template
# ---------------------------------------------------------------------------
CELL_NAME = "C2"
CELL_EMPNO = "C3"
CELL_TRADE = "F3"
CELL_MONTHYEAR = "C4"
CELL_SALARY = "F4"
GRID_HEADER_ROW = 6
GRID_FIRST_ROW = 7
GRID_LAST_ROW = 37
COL_DAY, COL_AM, COL_PM, COL_OT, COL_BH, COL_SITE, COL_ENGINEER, COL_COMMENTS = range(1, 9)
CELL_BASIC_PAY = "F42"
CELL_ALLOWANCES = "F47"
CELL_OTHER_DEDUCTION = "F48"
CELL_CACHED_FINAL_SALARY = "H44"

STATUS_KEYS = ["Present", "Absent", "Sick", "Medical", "Friday", "Holiday", "Leave"]

_CELL_REF_RE_DE = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')


def _find_label_cell_de(ws, candidates, min_row, max_row, min_col=1, max_col=8):
    """
    Same purpose as export_engine.py's _find_exact_label_cell: real cards
    vary their row layout by a row or two (confirmed directly across
    several real cards - "Basic Pay" doesn't sit at a fixed row on every
    one of them). Searches by matching text instead of trusting a fixed
    cell coordinate.
    """
    wanted = {c.strip().lower() for c in candidates}
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() in wanted:
                return r, c
    return None, None


def _find_final_salary_formula_refs(path, min_row=38, max_row=50, col=8):
    """
    Opens the file a SECOND time with data_only=False (the primary
    import load uses data_only=True and read_only=True for speed, which
    discards formula text) just to read the Final Salary formula's own
    cell references - e.g. "=ROUND((SUM(F42+F44+F45+F46)-F43-F47),0)"
    yields [F42,F44,F45,F46,F43,F47] in that order.

    Confirmed directly, on every real card checked, this formula always
    has the same structure: 4 references summed (Total Salary, OT
    Amount, BH Amount, Allowances, in that order) minus 2 references
    subtracted (Deduction, then Any-other-Deduction). This lets Basic
    Pay/Allowances/Other Deduction be found by their POSITION in the
    formula when a card's custom label for that row ("Security
    Deduction", "Two hrs BH from June 26", etc.) doesn't match any
    known label text - confirmed as a real, not hypothetical, problem:
    without this a stale or wrong-cell value silently produced the
    wrong Final Salary for 3 of 4 real cards checked in one batch.

    Returns a list of (row, col) tuples in formula order, or [] if no
    formula could be found/read.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return []
    try:
        ws = wb.worksheets[0]
        for r in range(min_row, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.startswith("=ROUND("):
                refs = _CELL_REF_RE_DE.findall(v)
                out = []
                for _, c_letter, _, r_str in refs:
                    try:
                        c_idx = sum((ord(ch) - 64) * (26 ** i) for i, ch in enumerate(reversed(c_letter)))
                        out.append((int(r_str), c_idx))
                    except ValueError:
                        continue
                if out:
                    return out
        return []
    finally:
        wb.close()

MONTH_LOOKUP = {m.lower(): i for i, m in enumerate(
    ["", "January", "February", "March", "April", "May", "June",
     "July", "August", "September", "October", "November", "December"])}


@dataclass
class DailyRow:
    emp_no: str
    emp_name: str
    trade: str
    month_year: str
    day: object
    am: str
    pm: str
    ot: float
    bh: float
    site: str
    engineer: str
    comments: str
    full_date: object
    source_file: str


@dataclass
class SalaryAdjustment:
    """
    A manual line-item adjustment applied AFTER the independently
    recalculated Final Salary - e.g. a deduction for an incident, or an
    addition for overtime carried from a previous period. These are
    tracked separately from Final Salary (never silently baked into it),
    the same way mismatch flags work: the recalculated figure stays the
    accuracy anchor, and adjustments are a visible, itemized, removable
    layer on top of it.
    """
    description: str
    amount: float
    is_deduction: bool


@dataclass
class EmployeeSummary:
    emp_no: str
    emp_name: str
    trade: str
    month_year: str
    total_salary: float
    present_days: float = 0
    absent_days: float = 0
    sick_days: float = 0
    medical_days: float = 0
    friday_days: float = 0
    holiday_days: float = 0
    leave_days: float = 0
    ot_hours: float = 0
    bh_hours: float = 0
    basic_pay_input: float = 0
    total_salary_component: float = 0.0
    deduction: float = 0.0
    ot_amount: float = 0.0
    bh_amount: float = 0.0
    allowances: float = 0.0
    other_deduction: float = 0.0
    final_salary: float = 0.0
    cached_final_salary: object = None
    mismatch: bool = False
    mismatch_amount: float = 0.0
    resolved_by_user: bool = False
    source_file: str = ""
    source_path: str = ""
    adjustments: list = field(default_factory=list)   # list[SalaryAdjustment]
    file_bytes: object = field(default=None, repr=False)   # raw original file, so a
                                                             # saved session still works
                                                             # for exact-copy exports even
                                                             # if the original file is
                                                             # later moved or deleted
    is_manual_entry: bool = False   # True when this summary was created via Daily
                                     # Attendance Entry rather than an imported file -
                                     # its file_bytes (if any) is a borrowed TEMPLATE
                                     # from a different month, so exports must regenerate
                                     # the daily grid and derived cells from this
                                     # summary's own data rather than copying the
                                     # template's cells verbatim (which would silently
                                     # show a different month's numbers)

    def adjustments_total(self):
        """Net effect of all adjustments: additions minus deductions."""
        total = 0.0
        for a in self.adjustments:
            total += -a.amount if a.is_deduction else a.amount
        return round(total, 2)

    def adjusted_final_salary(self):
        """Final Salary plus/minus every manual adjustment - the number that
        actually gets paid and printed, once adjustments have been entered."""
        return round(self.final_salary + self.adjustments_total(), 2)


@dataclass
class ImportResult:
    daily_rows: list = field(default_factory=list)
    summaries: list = field(default_factory=list)
    issues: list = field(default_factory=list)   # list of (filename, message)
    files_ok: int = 0
    files_failed: int = 0


def _nz(v):
    try:
        if v is None:
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _txt(v):
    """
    Converts a cell value to display text. Numeric cells that are whole
    numbers (e.g. a Site No stored as 913.0 in Excel) render as "913", not
    "913.0" - this matters everywhere: report tables, exports, AND filter
    matching (typing "913" into the Site No box needs to actually match).
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def parse_month_year(month_year_text, fallback_source=""):
    """Best-effort parse of the free-text Month & Year cell, e.g. 'August-2026'."""
    if not month_year_text:
        return None
    s = month_year_text.strip().replace("_", "-").replace("/", "-")
    m = re.match(r"([A-Za-z]+)\s*-?\s*(\d{4})", s)
    if m:
        mon = MONTH_LOOKUP.get(m.group(1).lower())
        if mon:
            return (int(m.group(2)), mon)
    m = re.match(r"(\d{1,2})\s*-\s*(\d{4})", s)
    if m:
        mon = int(m.group(1))
        if 1 <= mon <= 12:
            return (int(m.group(2)), mon)
    return None


def full_date_for_day(month_year_text, day_value):
    """
    Returns a python date for a given day-of-month number, anchored to the
    parsed Month & Year cell. Returns None if it can't be determined safely
    (in which case the row still imports, it just won't appear in
    date-filtered reports).
    """
    if not isinstance(day_value, (int, float)):
        return None
    my = parse_month_year(month_year_text)
    if my is None:
        return None
    year, month = my
    day = int(day_value)
    if not (1 <= day <= 31):
        return None
    try:
        return date(year, month, day)
    except ValueError:
        # e.g. day 31 in a 30-day month - the card is using a rolling pay
        # cycle rather than a calendar month; roll into the next month.
        try:
            if month == 12:
                return date(year + 1, 1, day - 31)
            return date(year, month + 1, day - _days_in_month(year, month))
        except ValueError:
            return None


def _days_in_month(year, month):
    if month == 12:
        nxt = date(year + 1, 1, 1)
    else:
        nxt = date(year, month + 1, 1)
    return (nxt - date(year, month, 1)).days


def recalculate_from_daily_rows(daily_rows, total_salary, basic_pay_input=0.0,
                                 allowances=0.0, other_deduction=0.0):
    """
    THE single source of truth for turning a set of daily attendance rows
    into payroll figures - present/absent/etc day-counts, OT/BH hours and
    amounts, and Final Salary. Used by BOTH import_one_file() (recalculating
    an imported card independently of its cached value) AND manual Daily
    Attendance Entry (recalculating as new days get entered by hand) -
    there must only ever be ONE place these formulas are written, or the
    two paths will silently drift apart and produce different salaries for
    the same underlying data. If you need to change a payroll formula,
    this is the only function that should need editing.

    daily_rows: any iterable of DailyRow (or objects with matching
    .am/.pm/.ot/.bh attributes) for ONE worker's ONE card/cycle.

    Returns a dict of every derived field, ready to unpack into an
    EmployeeSummary (either building a new one or updating an existing one).
    """
    status_counts = {k: 0.0 for k in STATUS_KEYS}
    ot_total = 0.0
    bh_total = 0.0

    for r in daily_rows:
        am = (r.am or "").strip()
        pm = (r.pm or "").strip()
        for status in STATUS_KEYS:
            half = 0.0
            if am.lower() == status.lower():
                half += 0.5
            if pm.lower() == status.lower():
                half += 0.5
            status_counts[status] += half
        ot_total += _nz(r.ot)
        bh_total += _nz(r.bh)

    daily_rate = total_salary / 30.0 if total_salary else 0.0

    total_salary_component = (
        status_counts["Present"] + status_counts["Sick"] + status_counts["Medical"]
        + status_counts["Friday"] + status_counts["Holiday"]
    ) * daily_rate
    deduction = daily_rate * status_counts["Absent"]
    ot_amount = (daily_rate / 8.0) * ot_total if daily_rate else 0.0
    bh_amount = (daily_rate / 8.0) * bh_total if daily_rate else 0.0

    # Round half-up with a tiny epsilon correction, not plain round().
    # Confirmed directly on a real card: Python's floating-point
    # arithmetic computed this sum as 1237.4999999999998 where the
    # mathematically exact result is precisely 1237.5 (the daily-rate
    # thirds/eighths involved should cancel out exactly) - round()
    # rounds that a hair below .5 down to 1237, while Excel's own
    # ROUND() on the same real numbers correctly produces 1238. The
    # epsilon nudges a value that's within floating-point noise of a
    # true .5 boundary back onto it before rounding, matching how Excel
    # (and everyday "round half up" expectations) actually behaves,
    # instead of Python's round-half-to-even on a value already
    # slightly corrupted by floating-point representation error.
    raw_total = (total_salary_component + ot_amount + bh_amount + allowances
                 - deduction - other_deduction)
    # Sign-aware: Excel's ROUND() rounds half AWAY FROM ZERO, not
    # toward +infinity - a naive floor(x+0.5) would round a negative
    # half-value (e.g. -175.5) up to -175 instead of the correct -176.
    if raw_total >= 0:
        final_salary = math.floor(raw_total + 0.5 + 1e-9)
    else:
        final_salary = math.ceil(raw_total - 0.5 - 1e-9)

    return {
        "present_days": status_counts["Present"], "absent_days": status_counts["Absent"],
        "sick_days": status_counts["Sick"], "medical_days": status_counts["Medical"],
        "friday_days": status_counts["Friday"], "holiday_days": status_counts["Holiday"],
        "leave_days": status_counts["Leave"], "ot_hours": ot_total, "bh_hours": bh_total,
        "basic_pay_input": basic_pay_input, "total_salary_component": total_salary_component,
        "deduction": deduction, "ot_amount": ot_amount, "bh_amount": bh_amount,
        "allowances": allowances, "other_deduction": other_deduction,
        "final_salary": final_salary,
    }


def import_one_file(path):
    """
    Reads a single Labour Ecard .xlsx file.
    Returns (daily_rows, summary, issue_message_or_None)
    Raises no exceptions - all failures are reported via issue_message so a
    batch import of 100 files never aborts on one bad file.
    """
    filename = os.path.basename(path)
    full_path = os.path.abspath(path)
    try:
        with open(path, "rb") as fh:
            raw_bytes = fh.read()
    except Exception as e:
        return [], None, f"Could not read file ({e})"
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return [], None, f"Could not open file ({e})"

    try:
        ws = wb.worksheets[0]

        emp_name = _txt(ws[CELL_NAME].value)
        emp_no = _txt(ws[CELL_EMPNO].value)
        trade = _txt(ws[CELL_TRADE].value)
        month_year = _txt(ws[CELL_MONTHYEAR].value)
        salary = _nz(ws[CELL_SALARY].value)

        if not emp_no:
            return [], None, "Employee No cell (C3) is blank - this file was skipped"

        # ---- Daily grid ----
        daily_rows = []

        for r in range(GRID_FIRST_ROW, GRID_LAST_ROW + 1):
            am = _txt(ws.cell(r, COL_AM).value)
            pm = _txt(ws.cell(r, COL_PM).value)
            ot = _nz(ws.cell(r, COL_OT).value)
            bh = _nz(ws.cell(r, COL_BH).value)
            site = _txt(ws.cell(r, COL_SITE).value)
            engineer = _txt(ws.cell(r, COL_ENGINEER).value)
            comments = _txt(ws.cell(r, COL_COMMENTS).value)
            day_val = ws.cell(r, COL_DAY).value

            if am == "" and pm == "" and site == "" and comments == "":
                continue  # unfilled future day - skip entirely, don't count

            fdate = full_date_for_day(month_year, day_val)

            daily_rows.append(DailyRow(
                emp_no=emp_no, emp_name=emp_name, trade=trade, month_year=month_year,
                day=day_val, am=am, pm=pm, ot=ot, bh=bh, site=site, engineer=engineer,
                comments=comments, full_date=fdate, source_file=filename,
            ))

        # ---- Recalculate summary independently from the grid, using the
        # ONE shared formula function (see recalculate_from_daily_rows) ----
        # Basic Pay/Allowances/Other Deduction found by dynamic label
        # search first (real cards vary this row by 1-2 rows - confirmed
        # directly), falling back to the SAME position in the card's own
        # Final Salary formula when the label doesn't match anything
        # known (a custom one-off label like "Security Deduction" or
        # "Two hrs BH from June 26" - also confirmed directly, on real
        # cards, not hypothetical). The old fixed CELL_BASIC_PAY/
        # CELL_ALLOWANCES/CELL_OTHER_DEDUCTION coordinates are the last
        # resort only, since trusting them blindly silently produced the
        # wrong Final Salary on 3 of 4 real cards checked in one batch -
        # missing a real Basic Pay entry entirely on one, and missing a
        # real ~AED 9-276 monthly note on two others.
        formula_refs = _find_final_salary_formula_refs(path)
        # Formula order is [total_salary, ot_amount, bh_amount, allowances,
        # deduction, other_deduction] - see _find_final_salary_formula_refs.
        formula_allowances_cell = formula_refs[3] if len(formula_refs) >= 4 else None
        formula_other_deduction_cell = formula_refs[5] if len(formula_refs) >= 6 else None

        def _read_input_field(label_candidates, fixed_coord, formula_cell):
            label_row, label_col = _find_label_cell_de(ws, label_candidates, 38, 50)
            if label_row is not None:
                return _nz(ws.cell(label_row, label_col + 3).value)
            if formula_cell is not None:
                return _nz(ws.cell(formula_cell[0], formula_cell[1]).value)
            return _nz(ws[fixed_coord].value)

        basic_pay_input = _read_input_field(["basic pay"], CELL_BASIC_PAY, None)
        allowances = _read_input_field(["allowances"], CELL_ALLOWANCES, formula_allowances_cell)
        other_deduction = _read_input_field(
            ["any other deduction(s)", "any other deduction"], CELL_OTHER_DEDUCTION, formula_other_deduction_cell)

        recalced = recalculate_from_daily_rows(
            daily_rows, salary, basic_pay_input, allowances, other_deduction)
        final_salary = recalced["final_salary"]

        cached_final = ws[CELL_CACHED_FINAL_SALARY].value
        cached_final_num = None
        mismatch = False
        mismatch_amount = 0.0
        if isinstance(cached_final, (int, float)):
            cached_final_num = float(cached_final)
            mismatch_amount = abs(cached_final_num - final_salary)
            mismatch = mismatch_amount > 1.0  # more than AED 1 difference

        summary = EmployeeSummary(
            emp_no=emp_no, emp_name=emp_name, trade=trade, month_year=month_year,
            total_salary=salary,
            cached_final_salary=cached_final_num,
            mismatch=mismatch, mismatch_amount=mismatch_amount, source_file=filename,
            source_path=full_path, file_bytes=raw_bytes,
            **recalced,
        )

        issue = None
        if mismatch:
            issue = (f"Recalculated final salary (AED {final_salary:,.0f}) does not match "
                      f"the value already on the card (AED {cached_final_num:,.0f}). "
                      f"Difference: AED {mismatch_amount:,.2f}. Verify this card manually.")
        if not month_year:
            note = "Month & Year cell (C4) is blank - this card's rows will be excluded from date-filtered reports."
            issue = (issue + " ALSO: " + note) if issue else note

        return daily_rows, summary, issue

    finally:
        wb.close()


def import_files(paths, progress_callback=None):
    """
    Imports a list of file paths. progress_callback(done, total, filename) is
    called after each file if provided (for a progress bar in the UI).
    """
    result = ImportResult()
    total = len(paths)
    for i, p in enumerate(paths, start=1):
        filename = os.path.basename(p)
        try:
            daily_rows, summary, issue = import_one_file(p)
        except Exception as e:
            daily_rows, summary, issue = [], None, f"Unexpected error: {e}"

        if summary is not None:
            result.daily_rows.extend(daily_rows)
            result.summaries.append(summary)
            result.files_ok += 1
        else:
            result.files_failed += 1

        if issue:
            result.issues.append((filename, issue))

        if progress_callback:
            progress_callback(i, total, filename)

    return result
