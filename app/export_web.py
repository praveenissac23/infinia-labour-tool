"""
Export logic for the web app - generates Excel and PDF salary cards
directly from EmployeeSummary + DailyRow database records. Unlike the
desktop app's export_engine.py, this does NOT try to preserve an
originally-uploaded Excel template, since every record in the web app
is entered directly through the API - there is no "source file" to
borrow formatting from. Structure mirrors the desktop app's own
regenerated-card layout (Employee info fields, full 1-31 day grid,
Total Days block, Salary Summary block, Final Salary box) so a card
produced here reads the same way a desktop-generated one does.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import payroll_cycle as pcyc

BRAND_RED = "C0392B"
BRAND_BLACK = "2E3238"
GREEN_FILL = "C6EFCE"

# A colour per attendance status. Pale enough to print legibly and to
# read black text on: green means the day was worked, red means it cost
# the worker money, the rest are calm - sick and medical are not faults
# and a rest day is normal.
STATUS_FILLS = {
    "Present": "E3F4E4", "Absent": "FBE0DE", "Sick": "FFF2D6",
    "Medical": "FDE8D7", "Sunday": "E4EDFA", "Friday": "E4EDFA",
    "Holiday": "E9E4F7", "Leave": "EFEFEF",
}
GREY_FILL = "D8D8D8"

DAILY_HEADERS = ["Date", "A.M", "P.M", "Site", "Engineer", "OT", "BH", "Comments"]

TOTAL_DAYS_FIELDS = [
    ("Present", "present_days"), ("Absent", "absent_days"), ("Sick", "sick_days"),
    ("Medical", "medical_days"), ("Friday", "friday_days"), ("Sunday", "sunday_days"),
    ("Holiday", "holiday_days"),
    ("Leave", "leave_days"), ("OT", "ot_hours"), ("BH", "bh_hours"),
]


def _total_days_fields(summary):
    """The day rows worth printing for this worker.

    Friday was the paid rest day before Sunday replaced it. Cycles
    recorded back then still hold Friday days and must still print them,
    but a card for a recent cycle should not carry a Friday row that
    will always read zero."""
    return [(label, attr) for label, attr in TOTAL_DAYS_FIELDS
            if attr != "friday_days" or (getattr(summary, attr, 0) or 0)]

SUMMARY_FIELDS = [
    ("Basic Pay", "basic_pay_input", None),
    ("Total Salary", "total_salary_component", None),
    ("Absence or Leave", "deduction", "-"),
    ("OT Amount", "ot_amount", "+"),
    ("BH Amount", "bh_amount", "+"),
]


def _adjusted_final_salary(summary):
    total = summary.final_salary
    for adj in summary.adjustments:
        total += (-adj.amount if adj.is_deduction else adj.amount)
    return round(total, 2)


def _rows_by_date(daily_rows):
    return {r.full_date: r for r in daily_rows if r.full_date}


def _num(v):
    """2.0 -> '2', 2.5 -> '2.5', whole numbers show clean with no trailing
    .0. Zero/blank/None all show as blank, matching the original 'row.ot
    or ""' convention - a day with attendance but no overtime shouldn't
    show a distracting '0' in every single row."""
    if not v:
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    return str(int(f)) if f == int(f) else str(f)


# The web app and the exports must speak the same language: the column is
# called "In central store" on screen, so the Excel and PDF say the same,
# not a machine-ish "In Store". One map, used by every store export.
STORE_LABELS = {
    "in_store": "In central store", "at_sites": "Out at sites", "total": "Total held",
    "by_site": "Where at sites", "item_type": "Type", "reorder_level": "Warn under",
    "hired_from": "Hired from", "on_hire": "On hire", "due_back": "Due back",
    "lost_damaged": "Lost / damaged", "written_off": "Written off",
    "qty": "Qty", "name": "Material", "code": "Code", "item": "Material",
    "ref": "Request", "requested_on": "Asked on", "needed_by": "Needed by",
    "days_late": "Days late", "outstanding": "Still to come",
}


def _store_label(k):
    return STORE_LABELS.get(k, k.replace("_", " ").title())


def _clean_qty(v):
    """380.0 -> '380', 7.5 -> '7.5'. Counts of things never show a fake
    decimal - 400 bags, not 400.0 - and genuinely fractional values keep
    only the decimals they actually have."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    return f"{int(f):,}" if f == int(f) else f"{f:,.2f}".rstrip("0").rstrip(".")


def _is_texty(k):
    """Columns that read as words, so they sit flush left like prose;
    numbers go flush right so digits line up down the column."""
    return any(w in k.lower() for w in
               ("name", "item", "category", "note", "description", "supplier",
                "purpose", "site", "status", "urgency", "hired", "person"))


def _cycle_dates(month_year):
    """
    Every actual calendar date in this cycle, in order - 26th of the
    prior month through the 25th of this one, matching how the cycle
    genuinely runs. Falls back to a plain 1-31 range only if month_year
    can't be parsed (shouldn't normally happen).
    """
    try:
        parsed = datetime.strptime(f"25 {month_year}", "%d %B %Y").date()
        start, end, _ = pcyc.cycle_bounds_for(parsed)
        dates = []
        cur = start
        while cur <= end:
            dates.append(cur)
            cur = cur.fromordinal(cur.toordinal() + 1)
        return dates
    except ValueError:
        return list(range(1, 32))


def _write_worker_card(ws, summary, rows, border, start_row):
    """
    Writes one worker's full card starting at start_row and returns the
    row number right after it (before the blank separator row) - lets
    the caller stack many cards in one worksheet, one blank row apart,
    instead of one sheet per worker.
    """
    thin_grey = Side(style="thin", color="BFBFBF")
    box_border = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)
    r = start_row

    # Employee info block - a modest-width panel (columns C-G) centered
    # within the 8-column grid below it, with blank margin columns on
    # both sides - not stretched edge-to-edge, which just left a big
    # bare grey band with no content in most of it.
    info = [
        ("Employee Name:", summary.emp_name), ("Employee No:", summary.emp_no),
        ("Trade:", summary.trade), ("Month & Year:", summary.month_year),
        ("Salary (AED):", summary.total_salary),
    ]
    for label, value in info:
        ws.row_dimensions[r].height = 15
        lbl = ws.cell(row=r, column=3, value=label)
        lbl.font = Font(bold=True, size=10.5)
        lbl.border = box_border
        lbl.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        val = ws.cell(row=r, column=4, value=value)
        val.fill = PatternFill("solid", fgColor=GREY_FILL)
        val.font = Font(size=10.5)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = box_border
        for col in (5, 6, 7):
            ws.cell(row=r, column=col).border = box_border
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=GREY_FILL)
        r += 1

    # Daily grid header
    header_row = r
    ws.row_dimensions[header_row].height = 13
    for i, h in enumerate(DAILY_HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=BRAND_RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    r += 1

    by_date = _rows_by_date(rows)
    cycle_dates = _cycle_dates(summary.month_year)
    for idx, d in enumerate(cycle_dates):
        row = by_date.get(d)
        label = d.strftime("%d %b") if hasattr(d, "strftime") else d
        vals = [label, row.am if row else "", row.pm if row else "", row.site if row else "",
                row.engineer if row else "", (_num(row.ot) if row else ""),
                (_num(row.bh) if row else ""), (row.comments if row else "")]
        stripe = "F7F7F7" if idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r].height = 15.5
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            # A.M and P.M carry their status colour; the rest of the row
            # keeps the plain stripe so the colour draws the eye.
            fill = STATUS_FILLS.get(v) if i in (2, 3) else None
            c.fill = PatternFill("solid", fgColor=fill or stripe)
            c.font = Font(size=11.5)
            if i == 4:  # Site column - values like "704" look numeric but
                c.number_format = "@"  # aren't; "@" stops Excel's green
                # triangle "number stored as text" warning on them.
        r += 1

    # OFFICE USE ONLY bar
    ws.row_dimensions[r].height = 14
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    office_cell = ws.cell(row=r, column=1, value="OFFICE USE ONLY")
    office_cell.font = Font(bold=True, size=9.5)
    office_cell.fill = PatternFill("solid", fgColor="BFBFBF")
    office_cell.alignment = Alignment(horizontal="center", vertical="center")
    office_cell.border = box_border
    r += 1

    block_start = r
    # Total Days (columns A-B), Salary Summary (columns D-E), Final
    # Salary box (columns G-H) - C and F are left as narrow, unbordered
    # spacer columns so the three blocks read as distinct panels rather
    # than one continuous, cluttered table.
    for label, attr in _total_days_fields(summary):
        ws.row_dimensions[r].height = 13
        status_fill = STATUS_FILLS.get(label)
        val = getattr(summary, attr, 0) or 0
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = Font(bold=True, size=11)
        lbl.fill = PatternFill("solid", fgColor=status_fill or GREY_FILL)
        lbl.border = box_border
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vcell = ws.cell(row=r, column=2, value=round(val, 2) if val else 0)
        vcell.fill = PatternFill("solid", fgColor=status_fill or GREY_FILL)
        vcell.border = box_border
        vcell.alignment = Alignment(horizontal="center", vertical="center")
        vcell.font = Font(size=11)
        r += 1
    total_days_end = r

    r = block_start
    for label, attr, sign in SUMMARY_FIELDS:
        val = getattr(summary, attr, 0) or 0
        prefix = f"{sign} " if sign else ""
        lbl = ws.cell(row=r, column=4, value=label)
        lbl.font = Font(bold=True, size=11)
        lbl.fill = PatternFill("solid", fgColor=GREY_FILL)
        lbl.border = box_border
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vcell = ws.cell(row=r, column=5, value=f"{prefix}AED {val:,.2f}")
        vcell.fill = PatternFill("solid", fgColor=GREY_FILL)
        vcell.border = box_border
        vcell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        vcell.font = Font(size=11)
        r += 1
    for adj in summary.adjustments:
        sign = "-" if adj.is_deduction else "+"
        lbl = ws.cell(row=r, column=4, value=adj.description)
        lbl.font = Font(bold=True, size=11)
        lbl.fill = PatternFill("solid", fgColor=GREY_FILL)
        lbl.border = box_border
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vcell = ws.cell(row=r, column=5, value=f"{sign} AED {adj.amount:,.2f}")
        vcell.fill = PatternFill("solid", fgColor=GREY_FILL)
        vcell.border = box_border
        vcell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        vcell.font = Font(size=8.5, color="C0392B" if adj.is_deduction else "2E7D32")
        r += 1
    summary_end = r

    # Final Salary to Process - its own box (columns G-H)
    head = ws.cell(row=block_start, column=7, value="FINAL SALARY TO PROCESS")
    head.font = Font(bold=True, color="FFFFFF", size=8.5)
    head.fill = PatternFill("solid", fgColor=BRAND_BLACK)
    head.border = box_border
    head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=block_start, start_column=7, end_row=block_start, end_column=8)
    ws.cell(row=block_start, column=8).border = box_border

    final_cell = ws.cell(row=block_start + 1, column=7, value=f"AED {_adjusted_final_salary(summary):,.2f}")
    final_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
    final_cell.font = Font(bold=True, size=13)
    final_cell.border = box_border
    final_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=block_start + 1, start_column=7, end_row=block_start + 1, end_column=8)
    ws.cell(row=block_start + 1, column=8).border = box_border

    card_end = max(total_days_end, summary_end, block_start + 2)

    # A card gets signed and handed over, so it needs somewhere to write
    # who checked it, the worker's signature, and anything said about it.
    # Printed boxes with room to write in, not data fields.
    sign_row = card_end + 1
    for col, label, width in ((1, "VERIFIED BY", 3), (4, "EMPLOYEE SIGNATURE", 2), (6, "REMARKS", 3)):
        head = ws.cell(row=sign_row, column=col, value=label)
        head.font = Font(bold=True, size=7.5, color="5B6167")
        head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=sign_row, start_column=col,
                       end_row=sign_row, end_column=col + width - 1)
        # A tall empty box underneath, ruled on all sides so it reads as
        # somewhere to sign rather than a gap in the sheet.
        for c in range(col, col + width):
            ws.cell(row=sign_row, column=c).border = box_border
            ws.cell(row=sign_row + 1, column=c).border = box_border
        ws.merge_cells(start_row=sign_row + 1, start_column=col,
                       end_row=sign_row + 1, end_column=col + width - 1)
    ws.row_dimensions[sign_row].height = 14
    ws.row_dimensions[sign_row + 1].height = 34

    return sign_row + 1


def build_combined_excel(summaries_with_rows):
    """
    summaries_with_rows: list of (EmployeeSummary, [DailyRow]) tuples.
    All workers stacked in ONE worksheet, one blank row between each
    card. Each card gets an explicit page break right after it so
    printing gives exactly one worker per physical page - the full
    31-day cycle (26th to 25th) is always shown, so every card is a
    consistent height that fits one page cleanly.
    """
    from openpyxl.worksheet.pagebreak import Break

    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Cards"
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 17
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 7
    ws.column_dimensions["H"].width = 26
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.freeze_panes = "A1"

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    last_row = 1
    for idx, (summary, rows) in enumerate(summaries_with_rows):
        next_row = _write_worker_card(ws, summary, rows, border, row)
        last_row = next_row
        if idx < len(summaries_with_rows) - 1:
            ws.row_breaks.append(Break(id=next_row))
        row = next_row + 1  # one blank row between cards

    ws.print_area = f"A1:H{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_separate_excel_files(summaries_with_rows):
    """
    One standalone .xlsx per worker instead of one workbook with many
    sheets - returns [(filename, BytesIO), ...] for the caller to zip.
    """
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    files = []
    for summary, rows in summaries_with_rows:
        wb = Workbook()
        ws = wb.active
        ws.title = "Card"
        ws.column_dimensions["A"].width = 17
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 19
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 30
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        _write_worker_card(ws, summary, rows, border, 1)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(summary.emp_no))
        files.append((f"{safe_name}.xlsx", buf))
    return files


def _build_pdf_card_elements(summary, rows, doc_width, styles):
    label_style = ParagraphStyle("InfoLabel", parent=styles["Normal"], fontSize=11.5, fontName="Helvetica-Bold", alignment=TA_LEFT)
    value_style = ParagraphStyle("InfoValue", parent=styles["Normal"], fontSize=11.5, alignment=TA_LEFT)
    grey = colors.HexColor(f"#{GREY_FILL}")
    grid_color = colors.HexColor("#B0B0B0")
    elements = []

    info_rows = [
        [Paragraph("Employee Name:", label_style), Paragraph(summary.emp_name or "", value_style)],
        [Paragraph("Employee No:", label_style), Paragraph(summary.emp_no or "", value_style)],
        [Paragraph("Trade:", label_style), Paragraph(summary.trade or "", value_style)],
        [Paragraph("Month & Year:", label_style), Paragraph(summary.month_year or "", value_style)],
        [Paragraph("Salary (AED):", label_style), Paragraph(f"{summary.total_salary:,.0f}", value_style)],
    ]
    info_tbl = Table(info_rows, colWidths=[doc_width * 0.24, doc_width * 0.40])
    info_tbl.hAlign = "CENTER"
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), grey),
        ("GRID", (0, 0), (-1, -1), 0.5, grid_color),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 6))

    cell_style = ParagraphStyle("CardCell", parent=styles["Normal"], fontSize=9, leading=10.3, alignment=TA_CENTER)
    head_style = ParagraphStyle("CardHead", parent=styles["Normal"], fontSize=9, leading=10.3,
                                 textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER)
    headers = ["Date", "A.M", "P.M", "OT", "BH", "Site", "Engineer", "Comments"]
    by_date = _rows_by_date(rows)
    cycle_dates = _cycle_dates(summary.month_year)
    data = [[Paragraph(h, head_style) for h in headers]]
    status_cells = []
    for d in cycle_dates:
        row = by_date.get(d)
        label = d.strftime("%d %b") if hasattr(d, "strftime") else str(d)
        if row is not None:
            vals = [label, row.am, row.pm, _num(row.ot), _num(row.bh), row.site, row.engineer, row.comments]
        else:
            vals = [label, "", "", "", "", "", "", ""]
        data.append([Paragraph(v or "", cell_style) for v in vals])
        # Remember which status each half-day carried, so the A.M and
        # P.M cells can be tinted once the table is built.
        status_cells.append((len(data) - 1, vals[1], vals[2]))

    col_widths = [doc_width * w for w in (0.10, 0.10, 0.10, 0.06, 0.06, 0.10, 0.14, 0.34)]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    row_bg_cmds = [("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F7F7F7")) for i in range(2, len(data), 2)]
    # Status colours go on after the stripes so they win.
    for ri, am, pm in status_cells:
        if STATUS_FILLS.get(am):
            row_bg_cmds.append(("BACKGROUND", (1, ri), (1, ri), colors.HexColor("#" + STATUS_FILLS[am])))
        if STATUS_FILLS.get(pm):
            row_bg_cmds.append(("BACKGROUND", (2, ri), (2, ri), colors.HexColor("#" + STATUS_FILLS[pm])))
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_RED}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, grid_color),
        ("TOPPADDING", (0, 0), (-1, -1), 1), ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ] + row_bg_cmds))
    elements.append(tbl)
    elements.append(Spacer(1, 4))

    office_tbl = Table([[Paragraph("OFFICE USE ONLY", ParagraphStyle(
        "OfficeUse", parent=styles["Normal"], fontSize=10.5, fontName="Helvetica-Bold", alignment=TA_CENTER))]],
        colWidths=[doc_width])
    office_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#BFBFBF")),
        ("GRID", (0, 0), (-1, -1), 0.5, grid_color),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(office_tbl)
    elements.append(Spacer(1, 6))

    value_right_style = ParagraphStyle("ValueRight", parent=styles["Normal"], fontSize=10.5, alignment=TA_RIGHT)
    day_fields = _total_days_fields(summary)
    days_data = [[Paragraph(label, label_style), Paragraph(f"{(getattr(summary, attr, 0) or 0):g}", value_right_style)]
                 for label, attr in day_fields]
    days_tbl = Table(days_data, colWidths=[doc_width * 0.16, doc_width * 0.09])
    days_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), grey),
        # Each day-status row takes its own colour, matching the grid
        # above and the app on screen.
        *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#" + STATUS_FILLS[label]))
          for i, (label, _) in enumerate(day_fields) if label in STATUS_FILLS],
        ("GRID", (0, 0), (-1, -1), 0.5, grid_color),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    summary_rows = []
    summary_row_colors = []
    for label, attr, sign in SUMMARY_FIELDS:
        val = getattr(summary, attr, 0) or 0
        prefix = f"{sign} " if sign else ""
        summary_rows.append([Paragraph(label, label_style), Paragraph(f"{prefix}AED {val:,.2f}", value_right_style)])
        summary_row_colors.append(None)
    for adj in summary.adjustments:
        sign = "-" if adj.is_deduction else "+"
        adj_value_style = ParagraphStyle("AdjValue", parent=value_right_style,
                                          textColor=colors.HexColor("#C0392B") if adj.is_deduction else colors.HexColor("#2E7D32"))
        summary_rows.append([Paragraph(adj.description, label_style),
                              Paragraph(f"{sign} AED {adj.amount:,.2f}", adj_value_style)])
    summary_tbl = Table(summary_rows, colWidths=[doc_width * 0.20, doc_width * 0.16])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), grey),
        ("GRID", (0, 0), (-1, -1), 0.5, grid_color),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    final_header = Paragraph("FINAL SALARY TO PROCESS", ParagraphStyle(
        "FinalHeader", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER))
    final_value = Paragraph(f"AED {_adjusted_final_salary(summary):,.2f}", ParagraphStyle(
        "FinalValue", parent=styles["Normal"], fontSize=13.5, fontName="Helvetica-Bold", alignment=TA_CENTER))
    final_tbl = Table([[final_header], [final_value]], colWidths=[doc_width * 0.20])
    final_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(f"#{BRAND_BLACK}")),
        ("BACKGROUND", (0, 1), (0, 1), colors.HexColor(f"#{GREEN_FILL}")),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("BOX", (0, 0), (-1, -1), 0.5, grid_color),
        ("LINEBELOW", (0, 0), (0, 0), 0.5, grid_color),
    ]))

    side_by_side = Table([[days_tbl, summary_tbl, final_tbl]],
                          colWidths=[doc_width * 0.27, doc_width * 0.38, doc_width * 0.22])
    side_by_side.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(side_by_side)

    # Somewhere to sign. A printed card is checked by someone, signed by
    # the worker, and often carries a note - without ruled boxes those
    # end up scrawled across the figures.
    label = ParagraphStyle("signlabel", parent=styles["Normal"], fontSize=6.5,
                            textColor=colors.HexColor("#5B6167"), spaceAfter=0, leading=8)
    sign_tbl = Table(
        [[Paragraph("VERIFIED BY", label), Paragraph("EMPLOYEE SIGNATURE", label),
          Paragraph("REMARKS", label)],
         ["", "", ""]],
        colWidths=[doc_width * 0.28, doc_width * 0.28, doc_width * 0.44],
        rowHeights=[10, 34])
    sign_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (0, 1), 0.5, grid_color),
        ("BOX", (1, 0), (1, 1), 0.5, grid_color),
        ("BOX", (2, 0), (2, 1), 0.5, grid_color),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, 0), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(Spacer(1, 8))
    elements.append(sign_tbl)
    return elements


def build_combined_pdf(summaries_with_rows):
    """
    Each worker's full 31-day cycle card gets its own page - one card
    per page keeps a consistent, predictable layout for a card sized
    to always show the full 26th-25th cycle.
    """
    from reportlab.platypus import PageBreak
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=8 * mm, bottomMargin=8 * mm,
                             leftMargin=8 * mm, rightMargin=8 * mm)
    styles = getSampleStyleSheet()
    elements = []
    for idx, (summary, rows) in enumerate(summaries_with_rows):
        elements.extend(_build_pdf_card_elements(summary, rows, doc.width, styles))
        if idx < len(summaries_with_rows) - 1:
            elements.append(PageBreak())
    doc.build(elements)
    buf.seek(0)
    return buf


def build_separate_pdf_files(summaries_with_rows):
    """One standalone .pdf per worker instead of one combined document."""
    styles = getSampleStyleSheet()
    files = []
    for summary, rows in summaries_with_rows:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=8 * mm, bottomMargin=8 * mm,
                                 leftMargin=8 * mm, rightMargin=8 * mm)
        elements = _build_pdf_card_elements(summary, rows, doc.width, styles)
        doc.build(elements)
        buf.seek(0)
        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(summary.emp_no))
        files.append((f"{safe_name}.pdf", buf))
    return files


def zip_files(files):
    """files: [(filename, BytesIO), ...] -> a single zip file as BytesIO."""
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, filebuf in files:
            zf.writestr(filename, filebuf.getvalue())
    buf.seek(0)
    return buf


REPORT_COLUMNS_META = {
    "emp_no": ("Emp No", "text"), "emp_name": ("Name", "text"), "trade": ("Trade", "text"),
    "sites": ("Site", "text"), "total_salary": ("Total Salary", "money"),
    "present_days": ("Present", "num"), "absent_days": ("Absent", "num"), "sick_days": ("Sick", "num"),
    "medical_days": ("Medical", "num"), "friday_days": ("Friday", "num"),
    "sunday_days": ("Sunday", "num"), "holiday_days": ("Holiday", "num"),
    "leave_days": ("Leave", "num"), "ot_hours": ("OT Hours", "num"), "bh_hours": ("BH Hours", "num"),
    "basic_pay_input": ("Basic Pay", "money"), "deduction": ("Absence/Leave Deduction", "money"),
    "ot_amount": ("OT Amount", "money"), "bh_amount": ("BH Amount", "money"),
    "final_salary": ("Final Salary", "money"), "adjustments": ("Adjustments", "text"),
    "adjusted_final_salary": ("Adjusted Final Salary", "money"),
}


def _report_row_value(item, key):
    if key == "adjusted_final_salary":
        return item.final_salary + sum(-a.amount if a.is_deduction else a.amount for a in item.adjustments)
    if key == "adjustments":
        return "; ".join(f"{a.description}: {'-' if a.is_deduction else '+'}{a.amount}" for a in item.adjustments) or "-"
    return getattr(item, key, "")


def build_report_table_excel(items, column_keys, cycle_label):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    cols = [(k, *REPORT_COLUMNS_META.get(k, (k, "text"))) for k in column_keys if k in REPORT_COLUMNS_META]
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (key, label, kind) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND_RED)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = 22 if kind == "text" else 16

    r = 2
    totals = {key: 0.0 for key, _, kind in cols if kind in ("money", "num")}
    for item in items:
        for i, (key, label, kind) in enumerate(cols, start=1):
            val = _report_row_value(item, key)
            if kind == "money":
                val_num = val or 0
                totals[key] += val_num
                display = f"AED {val_num:,.2f}"
            elif kind == "num":
                val_num = val or 0
                totals[key] += val_num
                display = val_num
            else:
                display = val
            c = ws.cell(row=r, column=i, value=display)
            c.alignment = Alignment(horizontal="center")
            c.border = border
        r += 1

    # Totals row
    for i, (key, label, kind) in enumerate(cols, start=1):
        if i == 1:
            c = ws.cell(row=r, column=i, value="TOTAL")
            c.font = Font(bold=True)
        elif kind == "money":
            c = ws.cell(row=r, column=i, value=f"AED {totals[key]:,.2f}")
            c.font = Font(bold=True)
        elif kind == "num":
            c = ws.cell(row=r, column=i, value=round(totals[key], 2))
            c.font = Font(bold=True)
        else:
            c = ws.cell(row=r, column=i, value="")
        c.fill = PatternFill("solid", fgColor=GREEN_FILL)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_report_table_pdf(items, column_keys, cycle_label):
    cols = [(k, *REPORT_COLUMNS_META.get(k, (k, "text"))) for k in column_keys if k in REPORT_COLUMNS_META]
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("TblCell", parent=styles["Normal"], fontSize=7, leading=9)
    head_style = ParagraphStyle("TblHead", parent=styles["Normal"], fontSize=7.5, leading=9,
                                 textColor=colors.white, fontName="Helvetica-Bold")
    bold_style = ParagraphStyle("TblBold", parent=styles["Normal"], fontSize=7.5, leading=9, fontName="Helvetica-Bold")

    data = [[Paragraph(label, head_style) for _, label, _ in cols]]
    totals = {key: 0.0 for key, _, kind in cols if kind in ("money", "num")}
    for item in items:
        row = []
        for key, label, kind in cols:
            val = _report_row_value(item, key)
            if kind == "money":
                val_num = val or 0
                totals[key] += val_num
                row.append(Paragraph(f"AED {val_num:,.2f}", cell_style))
            elif kind == "num":
                val_num = val or 0
                totals[key] += val_num
                row.append(Paragraph(str(val_num), cell_style))
            else:
                row.append(Paragraph(str(val) if val else "-", cell_style))
        data.append(row)

    total_row = []
    for i, (key, label, kind) in enumerate(cols):
        if i == 0:
            total_row.append(Paragraph("TOTAL", bold_style))
        elif kind == "money":
            total_row.append(Paragraph(f"AED {totals[key]:,.2f}", bold_style))
        elif kind == "num":
            total_row.append(Paragraph(str(round(totals[key], 2)), bold_style))
        else:
            total_row.append(Paragraph("", bold_style))
    data.append(total_row)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=10 * mm, bottomMargin=10 * mm,
                             leftMargin=8 * mm, rightMargin=8 * mm)
    col_width = doc.width / max(len(cols), 1)
    tbl = Table(data, colWidths=[col_width] * len(cols), repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_RED}")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D0D0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(f"#{GREEN_FILL}")),
    ]
    tbl.setStyle(TableStyle(style_cmds))

    title = Paragraph(f"Report - {cycle_label}", ParagraphStyle(
        "Title", parent=styles["Normal"], fontSize=13, fontName="Helvetica-Bold", spaceAfter=8))
    doc.build([title, tbl])
    buf.seek(0)
    return buf


def build_generic_result_excel(result_dict, cycle_label):
    """
    Exports an already-computed report result (columns/rows/totals, the
    shape build_custom_report or site_cost_center return) directly - no
    re-aggregation needed, just formatting what's already there.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    cols = result_dict["columns"]
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c["label"])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND_RED)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = 22

    # Money columns get Excel's thousands-separator format so a figure
    # like 21344.92 reads as 21,344.92 rather than a wall of digits.
    # Applied as a cell number_format (not a pre-formatted string) so the
    # value stays a real number Excel can still sum and chart.
    MONEY_FMT = '#,##0.00'
    COUNT_FMT = '#,##0.##'   # separators, but no forced decimals
    def is_money(key):
        k = key.lower()
        return "cost" in k or "amount" in k or "salary" in k or "pay" in k

    r = 2
    for row in result_dict["rows"]:
        for i, c in enumerate(cols, start=1):
            v = row.get(c["key"], "")
            cell = ws.cell(row=r, column=i, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if isinstance(v, (int, float)):
                cell.number_format = MONEY_FMT if is_money(c["key"]) else COUNT_FMT
        r += 1

    totals = result_dict.get("totals") or {}
    if totals:
        for i, c in enumerate(cols, start=1):
            v = totals.get(c["key"])
            cell = ws.cell(row=r, column=i, value="TOTAL" if i == 1 else (v if v is not None else ""))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if i > 1 and isinstance(v, (int, float)):
                cell.number_format = MONEY_FMT if is_money(c["key"]) else COUNT_FMT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_generic_result_pdf(result_dict, cycle_label):
    cols = result_dict["columns"]
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("TblCell", parent=styles["Normal"], fontSize=7, leading=9)
    head_style = ParagraphStyle("TblHead", parent=styles["Normal"], fontSize=7.5, leading=9,
                                 textColor=colors.white, fontName="Helvetica-Bold")
    bold_style = ParagraphStyle("TblBold", parent=styles["Normal"], fontSize=7.5, leading=9, fontName="Helvetica-Bold")

    def is_money(key):
        k = key.lower()
        return "cost" in k or "amount" in k or "salary" in k or "pay" in k

    def fmt(key, v):
        """Money gets 2 decimals (21,344.92); other numbers get thousands
        separators but keep their natural precision (1,776 stays whole,
        12.5 stays 12.5)."""
        if v is None or v == "":
            return ""
        if isinstance(v, (int, float)):
            if is_money(key):
                return f"{v:,.2f}"
            return f"{v:,.10g}" if v != int(v) else f"{int(v):,}"
        return str(v)

    data = [[Paragraph(c["label"], head_style) for c in cols]]
    for row in result_dict["rows"]:
        data.append([Paragraph(fmt(c["key"], row.get(c["key"], "")), cell_style) for c in cols])

    totals = result_dict.get("totals") or {}
    if totals:
        trow = []
        for i, c in enumerate(cols):
            v = totals.get(c["key"])
            trow.append(Paragraph("TOTAL" if i == 0 else fmt(c["key"], v), bold_style))
        data.append(trow)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=10 * mm, bottomMargin=10 * mm,
                             leftMargin=8 * mm, rightMargin=8 * mm)
    col_width = doc.width / max(len(cols), 1)
    tbl = Table(data, colWidths=[col_width] * len(cols), repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_RED}")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0D0D0")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    if totals:
        style_cmds.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(f"#{GREEN_FILL}")))
    tbl.setStyle(TableStyle(style_cmds))

    title = Paragraph(f"Report - {cycle_label}", ParagraphStyle(
        "Title", parent=styles["Normal"], fontSize=13, fontName="Helvetica-Bold", spaceAfter=8))
    doc.build([title, tbl])
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------
# STORE / INVENTORY EXPORTS
# ---------------------------------------------------------------------
def build_store_report_excel(title, rows, subtitle=""):
    """
    Any store report as a formatted sheet: company header, report title,
    the period it covers, bordered auto-width columns, and a totals row
    for numeric money columns. Column set is taken from the data, so one
    function serves every report rather than one per report drifting apart.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(rows[0]) if rows else 2))
    h = ws.cell(row=1, column=1, value="INFINIA CONTRACTING LLC")
    h.font = Font(bold=True, size=13)
    h.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(2, len(rows[0]) if rows else 2))
    t = ws.cell(row=2, column=1, value=title)
    t.font = Font(bold=True, size=11)
    t.alignment = Alignment(horizontal="center")
    if subtitle:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(2, len(rows[0]) if rows else 2))
        s = ws.cell(row=3, column=1, value=subtitle)
        s.font = Font(size=9, italic=True, color="777777")
        s.alignment = Alignment(horizontal="center")

    r = 5
    if not rows:
        ws.cell(row=r, column=1, value="Nothing to show.").font = Font(italic=True, color="999999")
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    cols = list(rows[0].keys())
    money_like = lambda k: any(w in k.lower() for w in ("cost", "value", "amount", "rate", "price"))
    header_row = r
    for i, k in enumerate(cols, start=1):
        c = ws.cell(row=r, column=i, value=_store_label(k))
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=BRAND_RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    r += 1

    numeric_totals = {k: 0 for k in cols if money_like(k)}
    for row in rows:
        for i, k in enumerate(cols, start=1):
            v = row.get(k, "")
            if isinstance(v, dict):
                # Site breakdowns are counts: "704: 380", never "704: 380.0".
                v = ", ".join(f"{a}: {_clean_qty(b)}" for a, b in v.items()) or "-"
            elif isinstance(v, bool):
                v = "Yes" if v else ""
            elif k == "item_type" and v:
                v = str(v).title()
            c = ws.cell(row=r, column=i, value=v)
            c.border = border
            c.font = Font(size=10)
            if isinstance(v, (int, float)):
                # Numbers flush right with thousands separators; counts
                # keep no fake decimals, money always shows two.
                c.number_format = '#,##0.00' if money_like(k) else '#,##0.##'
                c.alignment = Alignment(horizontal="right", vertical="center")
                if k in numeric_totals:
                    numeric_totals[k] += v
            else:
                c.alignment = Alignment(horizontal="left" if _is_texty(k) else "center",
                                        vertical="center")
        r += 1

    if numeric_totals:
        for i, k in enumerate(cols, start=1):
            v = round(numeric_totals[k], 2) if k in numeric_totals else ("TOTAL" if i == 1 else "")
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=GREEN_FILL)
            c.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "center")
            c.border = border
            if isinstance(v, (int, float)):
                c.number_format = '#,##0.00'

    # Long reports stay usable: headers stay put while scrolling, and the
    # filter arrows let the office slice by site or type right in Excel.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(cols))}{header_row + len(rows)}"

    for i, k in enumerate(cols, start=1):
        width = max(len(str(_store_label(k))) + 4, *(len(str(row.get(k, ""))) + 3 for row in rows))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, 10), 40)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def build_store_report_pdf(title, rows, subtitle=""):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=10 * mm, bottomMargin=10 * mm,
                             leftMargin=8 * mm, rightMargin=8 * mm)
    styles = getSampleStyleSheet()
    head = ParagraphStyle("H", parent=styles["Normal"], fontSize=8, leading=10,
                           textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER)
    cell = ParagraphStyle("C", parent=styles["Normal"], fontSize=8, leading=10, alignment=TA_CENTER)
    el = [Paragraph("<b>INFINIA CONTRACTING LLC</b>",
                     ParagraphStyle("T", parent=styles["Normal"], fontSize=13, alignment=TA_CENTER)),
          Spacer(1, 3),
          Paragraph(f"<b>{title}</b>",
                     ParagraphStyle("S", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER))]
    if subtitle:
        el += [Spacer(1, 2), Paragraph(subtitle,
                ParagraphStyle("Sub", parent=styles["Normal"], fontSize=8,
                                textColor=colors.HexColor("#777777"), alignment=TA_CENTER))]
    el.append(Spacer(1, 8))

    if not rows:
        el.append(Paragraph("Nothing to show.", cell))
        doc.build(el); buf.seek(0); return buf

    cols = list(rows[0].keys())
    money_like = lambda k: any(w in k.lower() for w in ("cost", "value", "amount", "rate", "price"))
    cellL = ParagraphStyle("CL", parent=cell, alignment=TA_LEFT)
    cellR = ParagraphStyle("CR", parent=cell, alignment=TA_RIGHT)
    data = [[Paragraph(_store_label(k), head) for k in cols]]
    totals = {k: 0 for k in cols if money_like(k)}
    for row in rows:
        line = []
        for k in cols:
            v = row.get(k, "")
            sty = cellL if _is_texty(k) else cell
            if isinstance(v, dict):
                # Site breakdowns are counts: "704: 380", never "704: 380.0".
                v = ", ".join(f"{a}: {_clean_qty(b)}" for a, b in v.items()) or "-"
                sty = cellL
            elif isinstance(v, bool):
                v = "Yes" if v else ""
            elif isinstance(v, (int, float)):
                if k in totals: totals[k] += v
                v = f"{v:,.2f}" if money_like(k) else _clean_qty(v)
                sty = cellR
            elif k == "item_type" and v:
                v = str(v).title()
            line.append(Paragraph(str(v) if v not in (None, "") else "-", sty))
        data.append(line)
    if totals:
        data.append([Paragraph(f"<b>{f'{totals[k]:,.2f}' if k in totals else ('TOTAL' if i == 0 else '')}</b>",
                               cellR if k in totals else cell)
                     for i, k in enumerate(cols)])

    w = doc.width / len(cols)
    tbl = Table(data, colWidths=[w] * len(cols), repeatRows=1)
    style = [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BRAND_RED)),
             ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
             ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
             ("ROWBACKGROUNDS", (0, 1), (-1, -2 if totals else -1),
              [colors.white, colors.HexColor("#F7F7F7")])]
    if totals:
        style.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#" + GREEN_FILL)))
    tbl.setStyle(TableStyle(style))
    el.append(tbl)
    doc.build(el); buf.seek(0); return buf


def build_material_request_pdf(mr: dict):
    """The request itself as a document the store keeper can send to the office."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                             leftMargin=14 * mm, rightMargin=14 * mm)
    styles = getSampleStyleSheet()
    cen = lambda s, sz, b=False: Paragraph(
        (f"<b>{s}</b>" if b else s),
        ParagraphStyle("x", parent=styles["Normal"], fontSize=sz, alignment=TA_CENTER))
    cell = ParagraphStyle("c", parent=styles["Normal"], fontSize=9, leading=11)
    head = ParagraphStyle("h", parent=styles["Normal"], fontSize=9, leading=11,
                           textColor=colors.white, fontName="Helvetica-Bold", alignment=TA_CENTER)

    el = [cen("INFINIA CONTRACTING LLC", 14, True), Spacer(1, 5),
          cen("MATERIAL REQUEST", 11, True), Spacer(1, 12)]

    cellR = ParagraphStyle("cr", parent=cell, alignment=TA_RIGHT)
    urg = mr["urgency"]
    urg_p = (Paragraph('<font color="#B26A00"><b>Urgent</b></font>', cell)
             if urg == "urgent" else Paragraph(str(urg).title(), cell))
    info = [["Request No:", mr["ref"], "Date:", mr["requested_on"]],
            ["Site:", mr["site"] or "-", "Needed by:", mr.get("needed_by") or "-"],
            ["Requested by:", mr["requested_by"] or "-", "Urgency:", urg_p],
            ["Status:", mr["status"].title(), "", ""]]
    t = Table([[Paragraph(f"<b>{a}</b>", cell),
                b if isinstance(b, Paragraph) else Paragraph(str(b), cell),
                Paragraph(f"<b>{c}</b>", cell),
                d if isinstance(d, Paragraph) else Paragraph(str(d), cell)] for a, b, c, d in info],
              colWidths=[doc.width * 0.16, doc.width * 0.34, doc.width * 0.16, doc.width * 0.34])
    t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
                            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F4F5F7")),
                            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F4F5F7")),
                            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    el += [t, Spacer(1, 12)]

    # "What it is for" travels with the request - it's the one thing the
    # office needs to judge whether to order, so the printed copy carries
    # it just like the screen does.
    lhead = ParagraphStyle("lh", parent=head, fontSize=8)
    data = [[Paragraph(h, lhead) for h in
             ["#", "Material", "Unit", "Requested", "Approved", "Received", "Still due", "What it is for", "Notes"]]]
    for i, ln in enumerate(mr["lines"], start=1):
        out = (ln["qty_requested"] or 0) - (ln["qty_received"] or 0)
        name = (f'{ln.get("item_code")} - {ln.get("item_name")}' if ln.get("item_code") else
                (ln.get("item_name") or ln.get("description") or "-"))
        data.append([Paragraph(str(i), cell), Paragraph(name, cell), Paragraph(ln["unit"], cell),
                     Paragraph(_clean_qty(ln["qty_requested"]), cellR),
                     Paragraph(_clean_qty(ln["qty_approved"]) if ln["qty_approved"] else "-", cellR),
                     Paragraph(_clean_qty(ln["qty_received"]) if ln["qty_received"] else "-", cellR),
                     Paragraph(_clean_qty(out) if out > 0 else "-", cellR),
                     Paragraph(ln.get("purpose") or "-", cell),
                     Paragraph(ln.get("notes") or "-", cell)])
    w = doc.width
    tbl = Table(data, colWidths=[w*.03, w*.22, w*.08, w*.11, w*.11, w*.10, w*.10, w*.14, w*.11], repeatRows=1)
    tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BRAND_RED)),
                              ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
                              ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                              ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")])]))
    el += [tbl, Spacer(1, 12)]

    if mr.get("notes"):
        el += [Paragraph(f"<b>Notes:</b> {mr['notes']}", cell), Spacer(1, 6)]
    if mr.get("office_remark"):
        el += [Paragraph(f"<b>Office remark:</b> {mr['office_remark']}", cell), Spacer(1, 6)]

    el += [Spacer(1, 26),
           Table([[Paragraph("Requested by", cell), Paragraph("Approved by", cell), Paragraph("Received by", cell)]],
                 colWidths=[w/3]*3,
                 style=TableStyle([("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.HexColor("#999999")),
                                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                                    ("ALIGN", (0, 0), (-1, -1), "CENTER")]))]
    doc.build(el); buf.seek(0); return buf
