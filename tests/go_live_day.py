"""Day one, as it will actually happen.

Run: cd app && DATABASE_URL=sqlite:////tmp/golive.db python3 ../tests/go_live_day.py

Starts from a Start Fresh - master lists in, nothing else - and walks
the first real day in the order the company will do it: the admin
checks the setup, a site engineer marks attendance and raises a
request, the office approves and orders it, the store keeper receives
and issues stock, the office runs the reports and the error check, and
the admin takes a backup. Every figure is checked against arithmetic
done by hand, not against what the app says elsewhere.
"""
import sys, io, json
sys.path.insert(0, '.')
from datetime import date, timedelta, datetime
from fastapi.testclient import TestClient
import database, models, auth
models.Base.metadata.create_all(database.engine)
import main

FAIL = []
def ck(label, ok, extra=''):
    print(('PASS ' if ok else 'FAIL ') + label + ('' if ok else f'   [{extra}]'))
    if not ok: FAIL.append(label)
def section(t): print(f"\n=== {t} ===")

TODAY = date(2026, 9, 3)
CYCLE = "September 2026"          # 26 Aug - 25 Sep

# ---------------------------------------------------------------- setup
db = database.SessionLocal()
db.add(models.User(username='admin', hashed_password=auth.hash_password('Admin@123'), full_name='Praveen', role='admin'))
db.add(models.User(username='office', hashed_password=auth.hash_password('Office@123'), full_name='Office Desk', role='office'))
db.add(models.User(username='amal', hashed_password=auth.hash_password('Site@123'), full_name='Amal', role='site'))
db.add(models.User(username='sreekanth', hashed_password=auth.hash_password('Store@123'), full_name='Sreekanth',
                   role='office', permissions='dashboard,store,requests,settings'))
workers = [('F-001', 'RAJAN KUMAR', 'Mason', 'Infinia', 3000, 1800),
           ('F-002', 'SURESH BABU', 'Helper', 'Infinia', 1500, 900),
           ('P-001', 'KUMAR SINGH', 'Carpenter', 'Prime Infinia', 2500, 1500),
           ('D-01', 'MOHAMED SALIM', 'Driver', 'Prime Infinia', 3200, 2000)]
for no, nm, tr, co, tot, bas in workers:
    db.add(models.Employee(emp_no=no, name=nm, trade=tr, company=co, total_salary=tot, basic_salary=bas, active=True))
for s in ['902', '904', '905']: db.add(models.Site(code=s, active=True))
for e in ['Amal', 'Muhsina', 'Febiyan']: db.add(models.Engineer(name=e, active=True))
for i, (nm, unit) in enumerate([('Cement OPC 50kg', 'bag'), ('Steel Bar 12mm', 'pcs'), ('Plywood 18mm', 'sheet'),
                                 ('Scaffold Ledger', 'pcs'), ('Safety Helmet', 'pcs')]):
    db.add(models.StoreItem(code=f'ITM{i+1}', name=nm, unit=unit,
                            item_type='rental' if 'Scaffold' in nm else 'consumable', active=True))
db.commit(); db.close()

c = TestClient(main.app)
def login(u, p):
    r = c.post('/auth/login', data={'username': u, 'password': p})
    return {'Authorization': 'Bearer ' + r.json()['access_token']} if r.status_code == 200 else None
A, O, S, K = login('admin', 'Admin@123'), login('office', 'Office@123'), login('amal', 'Site@123'), login('sreekanth', 'Store@123')

# ------------------------------------------------- 1. Admin checks setup
section("1. Admin signs in and checks the setup")
ck('all four logins work', all([A, O, S, K]))
emps = c.get('/employees', headers=A).json()
ck('workers present with company', len(emps) == 4 and {e['company'] for e in emps} == {'Infinia', 'Prime Infinia'})
ck('material list present', len(c.get('/store/items', headers=A).json()) == 5)
ck('no attendance yet', len(c.get('/attendance/2026-09-03', headers=A).json()) == 0)
ck('no requests yet', len(c.get('/store/requests', headers=A).json()) == 0)
ck('error check is empty of pay warnings', not any(r.get('severity') == 'legal' for r in c.get(f'/error-check/{CYCLE}', headers=A).json()['rows']))
me = c.get('/permissions/me', headers=S).json()
ck('site engineer sees only his screens', set(me['screens']) == {'dashboard', 'attendance', 'store', 'requests', 'settings'}, me['screens'])
ck('store keeper sees only his screens', set(c.get('/permissions/me', headers=K).json()['screens']) == {'dashboard', 'store', 'requests', 'settings'})

# -------------------------------------------------- 2. Site marks attendance
section("2. Site engineer marks today's attendance")
rows = [
    {'emp_no': 'F-001', 'full_date': '2026-09-03', 'am': 'Present', 'pm': 'Present', 'site': '904', 'engineer': 'Amal', 'ot': 2, 'bh': 0, 'comments': ''},
    {'emp_no': 'F-002', 'full_date': '2026-09-03', 'am': 'Present', 'pm': 'Absent',  'site': '904', 'engineer': 'Amal', 'ot': 0, 'bh': 0, 'comments': 'left at noon'},
    {'emp_no': 'P-001', 'full_date': '2026-09-03', 'am': 'Absent',  'pm': 'Absent',  'site': '',    'engineer': '',     'ot': 0, 'bh': 0, 'comments': ''},
    {'emp_no': 'D-01',  'full_date': '2026-09-03', 'am': 'Present', 'pm': 'Present', 'site': '902', 'engineer': 'Amal', 'ot': 0, 'bh': 3, 'comments': 'night load'},
]
r = c.post('/attendance/save', json={'month_year': CYCLE, 'rows': rows}, headers=S)
ck('attendance saved by the site engineer', r.status_code == 200, r.text[:150])
saved = {x['emp_no']: x for x in c.get('/attendance/2026-09-03', headers=S).json()}
ck('all four rows came back', len(saved) == 4)
ck('half day recorded as such', saved['F-002']['am'] == 'Present' and saved['F-002']['pm'] == 'Absent')
ck('comment kept', saved['D-01']['comments'] == 'night load')
r = c.post('/attendance/save', json={'month_year': CYCLE, 'rows': [
    {'emp_no': 'F-001', 'full_date': '2026-09-20', 'am': 'Present', 'pm': 'Present', 'site': '904', 'engineer': 'Amal', 'ot': 0, 'bh': 0, 'comments': ''}]}, headers=S)
ck('a future date is refused', r.status_code == 400, f'{r.status_code} {r.text[:80]}')
r = c.post('/attendance/save', json={'month_year': CYCLE, 'rows': [
    {'emp_no': 'F-001', 'full_date': '2026-09-02', 'am': 'Present', 'pm': 'Present', 'site': '', 'engineer': '', 'ot': 0, 'bh': 0, 'comments': ''}]}, headers=S)
ck('present without a site is refused', r.status_code == 400, f'{r.status_code} {r.text[:80]}')
r = c.post('/attendance/save', json={'month_year': CYCLE, 'rows': [
    {'emp_no': 'F-001', 'full_date': '2026-08-30', 'am': 'Sunday', 'pm': 'Sunday', 'site': '', 'engineer': '', 'ot': 0, 'bh': 0, 'comments': ''}]}, headers=S)
ck('Sunday accepted as a status', r.status_code == 200, r.text[:100])
# Yesterday for everyone, so summaries have two days
r = c.post('/attendance/save', json={'month_year': CYCLE, 'rows': [
    {'emp_no': n, 'full_date': '2026-09-02', 'am': 'Present', 'pm': 'Present', 'site': '904', 'engineer': 'Muhsina', 'ot': 1, 'bh': 0, 'comments': ''}
    for n, *_ in workers]}, headers=S)
ck('yesterday backfilled', r.status_code == 200)
ck('a site engineer cannot see salaries', all(e['total_salary'] == 0 for e in c.get('/employees', headers=S).json()))

# --------------------------------------------- 3. Live card arithmetic
section("3. Live card - checked by hand")
lc = c.get(f'/live-card/F-001/{CYCLE}', headers=A).json()
s = lc['summary']
# F-001: 2 present days (2 Sep, 3 Sep) + 1 Sunday (30 Aug); OT 1+2 = 3
ck('present days = 2', s['present_days'] == 2, s['present_days'])
ck('sunday days = 1', s['sunday_days'] == 1, s['sunday_days'])
ck('OT hours = 3', s['ot_hours'] == 3, s['ot_hours'])
daily = 3000 / 30
ck('salary component = 3 paid days x 100', abs(s['total_salary_component'] - 3 * daily) < 0.01, s['total_salary_component'])
hourly = 3000 / 30 / 8
ck('OT amount = 3 x hourly rate', abs(s['ot_amount'] - 3 * hourly) < 0.01, (s['ot_amount'], 3 * hourly))
lc2 = c.get(f'/live-card/F-002/{CYCLE}', headers=A).json()['summary']
ck('half day counts 0.5 present, 0.5 absent', lc2['present_days'] == 1.5 and lc2['absent_days'] == 0.5, (lc2['present_days'], lc2['absent_days']))
ck('half-day absence deducted at half a day', abs(lc2['deduction'] - 0.5 * 1500 / 30) < 0.01, lc2['deduction'])
ck('live card is refused to a site engineer', c.get(f'/live-card/F-001/{CYCLE}', headers=S).status_code == 403)

# ----------------------------------------------- 4. Site raises a request
section("4. Site engineer raises a material request")
items = {i['name']: i for i in c.get('/store/items', headers=S).json()}
mr = c.post('/store/requests', json={
    'site': '904', 'requested_by': 'Amal', 'needed_by': '2026-09-06', 'urgency': 'urgent', 'notes': 'slab pour Saturday',
    'lines': [
        {'item_id': items['Cement OPC 50kg']['id'], 'qty_requested': 100, 'unit': 'bag', 'purpose': 'slab', 'item_type': 'consumable', 'description': '', 'est_cost': 0, 'notes': ''},
        {'item_id': items['Steel Bar 12mm']['id'], 'qty_requested': 200, 'unit': 'pcs', 'purpose': 'slab', 'item_type': 'consumable', 'description': '', 'est_cost': 0, 'notes': ''},
        {'item_id': None, 'qty_requested': 5, 'unit': 'pcs', 'purpose': 'site office', 'item_type': 'consumable', 'description': 'Extension reel 30m', 'est_cost': 0, 'notes': 'not in list'},
    ]}, headers=S)
ck('request raised', mr.status_code == 200, mr.text[:150])
mr = mr.json()
ck('first request is MR-0001', mr['ref'] == 'MR-0001', mr['ref'])
ck('three lines including a free-text one', len(mr['lines']) == 3)
ck('engineer cannot approve his own request',
   c.post(f"/store/requests/{mr['id']}/status", json={'status': 'approved'}, headers=S).status_code == 403)
notes = c.get('/notifications', headers=O).json()
ck('office is notified of the new request', any('MR-0001' in n['title'] for n in notes['notifications']), [n['title'] for n in notes['notifications']])

# ------------------------------------------ 5. Office approves and orders
section("5. Office approves, orders from two suppliers")
lines = mr['lines']
cement, steel, reel = lines[0], lines[1], lines[2]
r = c.post(f"/store/request-lines/{cement['id']}/decision", json={'decision': 'approved'}, headers=O)
ck('office approves cement line', r.status_code == 200, r.text[:120])
r = c.post(f"/store/request-lines/{steel['id']}/decision", json={'decision': 'approved'}, headers=O)
ck('office approves steel line', r.status_code == 200, r.text[:120])
r = c.post(f"/store/request-lines/{reel['id']}/decision", json={'decision': 'rejected', 'reason': 'buy locally from petty cash'}, headers=O)
ck('office rejects the reel with a reason', r.status_code == 200, r.text[:120])
req = [x for x in c.get('/store/requests', headers=O).json() if x['id'] == mr['id']][0]
ck('rejection reason stored', any(l.get('reject_reason') == 'buy locally from petty cash' for l in req['lines']))
r = c.post(f"/store/requests/{mr['id']}/status", json={'status': 'ordered', 'line_ids': [cement['id']],
    'supplier': 'al raha trading', 'contact_person': 'bijuam', 'phone': '0500908900', 'expected_on': '2026-09-05'}, headers=O)
ck('office orders cement from Al Raha', r.status_code == 200, r.text[:150])
r = c.post(f"/store/requests/{mr['id']}/status", json={'status': 'ordered', 'line_ids': [steel['id']],
    'supplier': 'GATEWAY STEEL', 'contact_person': 'ravi', 'phone': '0551234567', 'expected_on': '2026-09-05'}, headers=O)
ck('office orders steel from Gateway', r.status_code == 200, r.text[:150])

# The follow-up screen is built from open requests: a line still owed
# something (received < requested) that is not rejected.
def followup(h):
    out = []
    for rq in c.get('/store/requests', headers=h).json():
        if rq['status'] in ('delivered', 'received', 'closed', 'rejected'): continue
        for l in rq['lines']:
            if (l.get('status') or 'pending') == 'rejected': continue
            if (l['qty_requested'] - (l.get('qty_received') or 0)) <= 0: continue
            out.append(l)
    return out
sup = {x['name'] for x in c.get('/store/suppliers', headers=O).json()}
ck('supplier names normalised', sup == {'Al Raha Trading', 'Gateway Steel'}, sup)
fu = followup(O)
ck('both orders on the follow-up list', len(fu) == 2, len(fu))
ck('follow-up carries the phone to call', all((x.get('supplier') or {}).get('phone') or x.get('phone') for x in fu), [x.get('supplier') for x in fu])
ck('site engineer sees the request status', any(x['status'] in ('ordered', 'arranging', 'lpo_sent') for x in c.get('/store/requests', headers=S).json()))

# --------------------------------------------- 6. Store keeper receives
section("6. Store keeper receives the deliveries")
r = c.post(f"/store/requests/{mr['id']}/receive-bulk", json={
    'supplier': 'Al Raha Trading', 'reference': 'DN-4471', 'notes': '', 'received_on': '2026-09-03',
    'lines': [{'line_id': cement['id'], 'qty': 100}]}, headers=K)
ck('keeper receives the cement', r.status_code == 200, r.text[:150])
r = c.post(f"/store/requests/{mr['id']}/receive-bulk", json={
    'supplier': 'Gateway Steel', 'reference': 'INV-882', 'notes': 'short by 20', 'received_on': '2026-09-03',
    'lines': [{'line_id': steel['id'], 'qty': 180}]}, headers=K)
ck('keeper receives steel, short delivery', r.status_code == 200, r.text[:150])
req = [x for x in c.get('/store/requests', headers=K).json() if x['id'] == mr['id']][0]
st = {l['id']: l for l in req['lines']}
ck('cement line fully received', st[cement['id']]['qty_received'] == 100, st[cement['id']]['qty_received'])
ck('steel line partly received', st[steel['id']]['qty_received'] == 180, st[steel['id']]['qty_received'])
ck('request marked partial while short', req['status'] == 'partial', req['status'])
ck('request stays open while steel is short', req['status'] != 'delivered' and req['status'] != 'closed', req['status'])
fu = followup(O)
ck('cement drops off follow-up, steel stays', len(fu) == 1 and (fu[0].get('supplier') or {}).get('name') == 'Gateway Steel', [(x.get('supplier') or {}).get('name') for x in fu])
stock = {x['name']: x for x in c.get('/store/stock', headers=K).json()}
ck('stock: 100 cement in the store', stock['Cement OPC 50kg']['central'] == 100, stock.get('Cement OPC 50kg'))
ck('stock: 180 steel in the store', stock['Steel Bar 12mm']['central'] == 180)
r = c.post(f"/store/requests/{mr['id']}/receive-bulk", json={
    'supplier': 'Gateway Steel', 'reference': 'INV-883', 'notes': '', 'received_on': '2026-09-03',
    'lines': [{'line_id': steel['id'], 'qty': 20}]}, headers=K)
ck('balance delivered', r.status_code == 200)
req = [x for x in c.get('/store/requests', headers=K).json() if x['id'] == mr['id']][0]
ck('request closes when everything has arrived', req['status'] in ('delivered', 'closed'), req['status'])
ck('follow-up list is empty', len(followup(O)) == 0, len(followup(O)))

# ------------------------------------------------ 7. Issue to site
section("7. Store keeper issues stock to site, and rentals")
r = c.post('/store/movements', json={'item_id': items['Cement OPC 50kg']['id'], 'kind': 'out', 'qty': 60, 'location': '904', 'incharge': 'Amal', 'moved_on': '2026-09-03'}, headers=K)
ck('60 bags issued to site 904', r.status_code == 200, r.text[:120])
r = c.post('/store/movements', json={'item_id': items['Cement OPC 50kg']['id'], 'kind': 'out', 'qty': 500, 'location': '904', 'incharge': 'Amal', 'moved_on': '2026-09-03'}, headers=K)
ck('issuing more than is in stock is refused', r.status_code == 400, f'{r.status_code} {r.text[:80]}')
r = c.post('/store/movements', json={'item_id': items['Scaffold Ledger']['id'], 'kind': 'in', 'qty': 300, 'location': '', 'supplier': 'Hire Co', 'moved_on': '2026-09-03', 'rental_due': '2026-10-03'}, headers=K)
ck('rental received', r.status_code == 200, r.text[:120])
r = c.post('/store/movements', json={'item_id': items['Scaffold Ledger']['id'], 'kind': 'out', 'qty': 120, 'location': '905', 'incharge': 'Febiyan', 'moved_on': '2026-09-03'}, headers=K)
ck('rental sent to site', r.status_code == 200)
stock = {x['name']: x for x in c.get('/store/stock', headers=K).json()}
ck('cement: 40 in store, 60 at sites', stock['Cement OPC 50kg']['central'] == 40 and stock['Cement OPC 50kg']['out_at_sites'] == 60, stock['Cement OPC 50kg'])
ck('scaffold: 180 in store, 120 at sites', stock['Scaffold Ledger']['central'] == 180 and stock['Scaffold Ledger']['out_at_sites'] == 120, stock['Scaffold Ledger'])
mv = c.get('/store/movements', headers=K).json()
ck('every movement is recorded', len(mv) >= 6, len(mv))
ck('a site engineer cannot issue stock', c.post('/store/movements', json={'item_id': items['Cement OPC 50kg']['id'], 'kind': 'out', 'qty': 1, 'location': '904', 'incharge': 'Amal', 'moved_on': '2026-09-03'}, headers=S).status_code == 403)

# -------------------------------------------- 8. Salary adjustment
section("8. Office records a salary advance and a bonus")
sums = c.get(f'/summaries?month_year={CYCLE}', headers=A).json() if c.get(f'/summaries?month_year={CYCLE}', headers=A).status_code == 200 else []
d2 = database.SessionLocal()
sid = {s.emp_no: s.id for s in d2.query(models.EmployeeSummary).filter(models.EmployeeSummary.month_year == CYCLE).all()}
d2.close()
r = c.post(f"/summaries/{sid['D-01']}/adjustments", json={'description': 'Advance', 'amount': 1500, 'is_deduction': True}, headers=A)
ck('advance recorded', r.status_code == 200, r.text[:120])
r = c.post(f"/summaries/{sid['F-001']}/adjustments", json={'description': 'Site bonus', 'amount': 200, 'is_deduction': False}, headers=A)
ck('a site engineer cannot add an adjustment', c.post(f"/summaries/{sid['F-001']}/adjustments", json={'description': 'x', 'amount': 1, 'is_deduction': False}, headers=S).status_code == 403)
ck('bonus recorded', r.status_code == 200, r.text[:120])
lc = c.get(f'/live-card/D-01/{CYCLE}', headers=A).json()
ck('advance appears on the live card', any(a['description'] == 'Advance' for a in lc['summary']['adjustments']))

# -------------------------------------------------- 9. Error check
section("9. Error check on day one")
ec = c.get(f'/error-check/{CYCLE}', headers=O).json()['rows']
by = {}
for r in ec: by.setdefault(r['emp_no'], []).append(r)
# D-01: 2 present days = 213.33 pay, minus 1500 advance -> far below 40% of 3200
ck('D-01 flagged for the 40% rule', any(x.get('severity') == 'legal' for x in by.get('D-01', [])), [x['kind'] for x in by.get('D-01', [])])
ck('every worker shows missing days (only 2-3 entered)', all(any('missing' in x['kind'] for x in by.get(n, [])) for n, *_ in workers))
ck('no contradiction flagged (no OT on absent days)', not any(x.get('severity') == 'contradiction' for x in ec))
ck('site engineer cannot run error check', c.get(f'/error-check/{CYCLE}', headers=S).status_code == 403)
# Now create a contradiction and confirm it is caught
c.post('/attendance/save', json={'month_year': CYCLE, 'rows': [
    {'emp_no': 'P-001', 'full_date': '2026-09-01', 'am': 'Absent', 'pm': 'Absent', 'site': '', 'engineer': '', 'ot': 4, 'bh': 0, 'comments': ''}]}, headers=S)
ec = c.get(f'/error-check/{CYCLE}', headers=O).json()['rows']
ck('OT on an absent day is caught', any(x.get('severity') == 'contradiction' and x['emp_no'] == 'P-001' for x in ec))

# ---------------------------------------------------- 10. Reports
section("10. Reports and exports")
cat = c.get('/reports/builder-catalog', headers=O).json()
ck('catalog offers company on both sources', all(any(d['key'] == 'company' for d in cat[s]['dimensions']) for s in ('daily', 'summary')))
r = c.get(f'/reports/custom?month_year={CYCLE}&data_source=daily&dimensions=company&measures=worker_count,days_present,ot_hours', headers=O).json()
rows = {x['dim_0']: x for x in r['rows']}
ck('daily report groups by company', set(rows) == {'Infinia', 'Prime Infinia'}, list(rows))
ck('Infinia headcount 2', rows['Infinia']['worker_count'] == 2)
r = c.get(f'/reports/custom?month_year={CYCLE}&data_source=summary&dimensions=name&measures=additions,addition_reasons,deductions,deduction_reasons&company=Prime%20Infinia', headers=O).json()
names = {x['dim_0']: x for x in r['rows']}
ck('filtered to Prime Infinia only', set(names) == {'KUMAR SINGH', 'MOHAMED SALIM'}, list(names))
ck('advance shown with its reason', names['MOHAMED SALIM']['deductions'] == 1500 and 'Advance 1,500' in names['MOHAMED SALIM']['deduction_reasons'], names['MOHAMED SALIM'])
ck('title names the company', 'Prime Infinia' in r['title'])
tok = c.post('/auth/download-token', headers=O).json()['token']
for name, path in [('combined excel', f'/export/{CYCLE}/excel'), ('separate excel', f'/export/{CYCLE}/excel-separate'),
                   ('combined pdf', f'/export/{CYCLE}/pdf'), ('separate pdf', f'/export/{CYCLE}/pdf-separate'),
                   ('custom report excel', f'/export/{CYCLE}/custom-report?data_source=summary&dimensions=name&measures=present_days,final_salary&format=excel'),
                   ('custom report pdf', f'/export/{CYCLE}/custom-report?data_source=summary&dimensions=name&measures=present_days,final_salary&format=pdf')]:
    sep = '&' if '?' in path else '?'
    x = c.get(f'{path}{sep}token={tok}', headers=O)
    ck(f'{name} builds', x.status_code == 200 and len(x.content) > 1500, f'{x.status_code} {len(x.content)}')
from openpyxl import load_workbook
x = c.get(f'/export/{CYCLE}/excel?token={tok}', headers=O)
ws = load_workbook(io.BytesIO(x.content)).active
cells = [str(v.value) for row in ws.iter_rows() for v in row if v.value]
ck('printed card has sign-off boxes', 'VERIFIED BY' in cells and 'EMPLOYEE SIGNATURE' in cells and 'REMARKS' in cells)
ck('printed card shows Sunday row for F-001', 'Sunday' in cells)
for name, path in [('stock report excel', '/export/store/report?kind=stock&format=excel'), ('stock report pdf', '/export/store/report?kind=stock&format=pdf'),
                   ('by-site excel', '/export/store/report?kind=by_site&format=excel'), ('purchases pdf', '/export/store/report?kind=purchases&format=pdf'),
                   ('usage excel', '/export/store/report?kind=usage&format=excel'), ('returnable pdf', '/export/store/report?kind=returnable&format=pdf'),
                   ('low stock excel', '/export/store/report?kind=low&format=excel'),
                   ('open requests excel', '/export/store/report?kind=mr_open&format=excel'), ('all requests pdf', '/export/store/report?kind=mr_all&format=pdf')]:
    x = c.get(f'{path}&token={tok}', headers=K)
    ck(f'{name} builds', x.status_code == 200 and len(x.content) > 1000, f'{x.status_code} {len(x.content)}')

# ------------------------------------------------- 11. Backups
section("11. Backups at the end of day one")
r = c.post('/backup/create', headers=K)
ck('store keeper can take a backup', r.status_code == 200)
tok_k = c.post('/auth/download-token', headers=K).json()['token']
ck('store keeper is NOT given the full daily download (salaries)', c.get(f'/backup/latest/download?token={tok_k}', headers=K).status_code == 403)
tok_o = c.post('/auth/download-token', headers=O).json()['token']
x = c.get(f'/backup/latest/download?token={tok_o}', headers=O)
ck('office gets the full daily download', x.status_code == 200)
d = json.loads(x.content)
ck('backup holds today\'s work', len(d['daily_rows']) >= 10 and len(d['material_requests']) == 1 and len(d['store_movements']) >= 6 and len(d['adjustments']) == 2,
   {k: len(v) for k, v in d.items() if isinstance(v, list)})
ck('backup carries no password hashes', all(not u['hashed_password'] for u in d['users']))
bid = c.post('/backup/create', headers=A).json()['id']
ck('office cannot restore', c.post(f'/backup/{bid}/restore', headers=O).status_code == 403)
ck('office cannot start fresh', c.post('/backup/fresh-start', json={'confirm': 'CLEAR EVERYTHING'}, headers=O).status_code == 403)

# ------------------------------------------------ 12. Settings
section("12. Settings for each role")
r = c.post('/auth/change-password', json={'current_password': 'Site@123', 'new_password': 'NewSite@456'}, headers=S)
ck('site engineer changes his own password', r.status_code == 200, r.text[:100])
ck('old password no longer works', c.post('/auth/login', data={'username': 'amal', 'password': 'Site@123'}).status_code == 401)
ck('new password works', login('amal', 'NewSite@456') is not None)
ck('office cannot list logins', c.get('/users', headers=O).status_code == 403)
ck('admin can', c.get('/users', headers=A).status_code == 200)
ck('a new login can only be office/site/admin', c.post('/users', json={'username': 'x1', 'password': 'pw123456', 'full_name': 'X', 'role': 'staff'}, headers=A).status_code == 400)

# -------------------------------------------------- 13. Master data
section("13. Master data changes during the day")
r = c.post('/employees', json={'emp_no': 'F-003', 'name': 'new joiner', 'trade': 'helper', 'company': 'Prime Infinia', 'total_salary': 1400, 'basic_salary': 800, 'active': True}, headers=A)
ck('new worker added', r.status_code == 200, r.text[:100])
e = [x for x in c.get('/employees', headers=A).json() if x['emp_no'] == 'F-003'][0]
ck('name cased properly', e['name'] == 'New Joiner', e['name'])
ck('company stored', e['company'] == 'Prime Infinia')
r = c.delete('/employees/F-003', headers=A).json()
ck('a worker with no history is deleted outright', r['action'] == 'deleted', r)
r = c.delete('/employees/F-001', headers=A).json()
ck('a worker with attendance is deactivated instead', r['action'] == 'deactivated', r)
ck('his attendance survives', len([x for x in c.get('/attendance/2026-09-03', headers=A).json() if x['emp_no'] == 'F-001']) == 1)
ck('he leaves the active list', all(x['emp_no'] != 'F-001' for x in c.get('/employees?active_only=true', headers=A).json()))
ck('site engineer cannot touch master data', c.post('/employees', json={'emp_no': 'Z', 'name': 'Z', 'trade': '', 'total_salary': 1, 'basic_salary': 1, 'active': True}, headers=S).status_code == 403)

print()
print('GO-LIVE DAY CLEAN' if not FAIL else f'{len(FAIL)} PROBLEM(S):')
for f in FAIL: print('   -', f)
sys.exit(1 if FAIL else 0)
